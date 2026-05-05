#!/usr/bin/env python3
"""
Scraper Pharmacie Cap 3000
==========================
Cible : https://www.pharmacie-cap3000.com/meilleures-ventes
Moteur : requests + BeautifulSoup (HTML server-rendered)
Sortie : benchmark_cap3000.xlsx

Chaque li.ajax_block_product contient un <pre> caché avec les données PHP :
ean13, price, name, etc. → extraction directe, pas de page produit nécessaire.
"""

import re
import time
import unicodedata
import logging
import sys
from pathlib import Path
from typing import Optional, List, Dict

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "https://www.pharmacie-cap3000.com"
LIST_URL = f"{BASE_URL}/meilleures-ventes"
DELAY       = 0.5
MAX_RETRIES = 3
TIMEOUT     = 20

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("cap3000")


def normalize(text: str) -> str:
    if not text:
        return ""
    text = text.lower().strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def extract_php_field(pre_text: str, field: str) -> str:
    """Extrait un champ depuis le dump PHP array dans <pre>.
    Format : ["field"]=>\nstring(N) "value"  |  float(N.NN)  |  int(N)
    """
    lines = pre_text.split('\n')
    for i, line in enumerate(lines):
        if f'["{field}"]' in line and i + 1 < len(lines):
            val = lines[i + 1].strip()
            m = re.search(r'string\(\d+\) "(.+)"', val)
            if m:
                return m.group(1)
            m = re.search(r'float\(([\d.]+)\)', val)
            if m:
                return m.group(1)
            m = re.search(r'int\((\d+)\)', val)
            if m:
                return m.group(1)
    return ""


def get_page(session: requests.Session, url: str) -> Optional[BeautifulSoup]:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except requests.RequestException as e:
            wait = attempt * 3
            log.warning(f"Tentative {attempt}/{MAX_RETRIES} ({e}) — attente {wait}s")
            time.sleep(wait)
    log.error(f"Échec : {url}")
    return None


def get_last_page(soup: BeautifulSoup) -> int:
    nums = []
    for a in soup.find_all("a", href=True):
        m = re.search(r'[?&]p=(\d+)', a["href"])
        if m:
            nums.append(int(m.group(1)))
        m2 = re.search(r'pharmaciecap3000\[page\]=(\d+)', a["href"])
        if m2:
            nums.append(int(m2.group(1)))
    return max(nums) if nums else 1


def parse_listing(soup: BeautifulSoup) -> List[Dict]:
    products = []
    for li in soup.find_all("li", class_="ajax_block_product"):
        # ── Données PHP dans le <pre> caché ──────────────────────
        pre = li.find("pre")
        pre_text = pre.get_text() if pre else ""

        ean    = extract_php_field(pre_text, "ean13") or extract_php_field(pre_text, "reference")
        prix_s = extract_php_field(pre_text, "price")
        try:
            prix = float(prix_s) if prix_s else None
        except ValueError:
            prix = None

        # ── Nom et URL depuis le lien ─────────────────────────────
        link = li.find("a", title=True)
        if not link:
            link = li.find("a", href=True)
        name = (link.get("title") or link.get_text(strip=True)) if link else ""
        url  = link["href"] if link else ""
        if url and not url.startswith("http"):
            url = BASE_URL + url

        # Fallback nom via <pre>
        if not name:
            name = extract_php_field(pre_text, "name")

        if not name:
            continue

        # Prix affiché depuis le HTML
        price_el = li.find(class_=re.compile(r"price"))
        prix_aff = price_el.get_text(strip=True) if price_el else (f"{prix:.2f}€" if prix else "")

        # Marque
        brand_el = li.find(class_=re.compile(r"manufacturer|brand"))
        marque = brand_el.get_text(strip=True) if brand_el else ""

        products.append({
            "nom":          name.strip(),
            "nom_norm":     normalize(name),
            "marque":       marque,
            "ean":          ean if ean and len(ean) >= 8 else "",
            "prix_affiche": prix_aff,
            "prix":         prix,
            "url":          url,
        })
    return products


def export_excel(products: List[Dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cap 3000"

    cols = [
        ("Nom",            "nom"),
        ("Marque",         "marque"),
        ("EAN",            "ean"),
        ("Prix affiché",   "prix_affiche"),
        ("Prix numérique", "prix"),
        ("Nom normalisé",  "nom_norm"),
        ("URL",            "url"),
    ]

    hdr_fill = PatternFill("solid", fgColor="2D6A4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (label, _) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    for row_i, p in enumerate(products, 2):
        for col_i, (_, key) in enumerate(cols, 1):
            val = p.get(key)
            if key == "url" and val:
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.hyperlink = val
                cell.font = Font(color="0057FF", underline="single")
            elif key == "prix" and val is not None:
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.number_format = '#,##0.00 "€"'
            else:
                ws.cell(row=row_i, column=col_i, value=val)

    widths = [55, 25, 16, 14, 16, 55, 80]
    for col_i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col_i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(path)
    sz = path.stat().st_size / 1024
    log.info(f"Export : {path}  ({sz:.0f} KB, {len(products)} produits)")


def main():
    out = Path(__file__).parent / "benchmark_cap3000.xlsx"

    log.info("=" * 60)
    log.info("  Scraper Pharmacie Cap 3000 — Meilleures ventes")
    log.info(f"  Cible : {LIST_URL}")
    log.info("=" * 60)

    session = requests.Session()

    soup = get_page(session, LIST_URL)
    if not soup:
        log.error("Impossible de charger la page 1.")
        return

    last_page = get_last_page(soup)
    log.info(f"Pages détectées : {last_page}")

    all_products: List[Dict] = []
    seen_urls: set = set()

    for page_num in range(1, last_page + 1):
        if page_num > 1:
            url = f"{LIST_URL}?p={page_num}"
            soup = get_page(session, url)
            if not soup:
                continue

        products = parse_listing(soup)
        new = 0
        for p in products:
            key = p["url"] or p["nom_norm"]
            if key not in seen_urls:
                seen_urls.add(key)
                all_products.append(p)
                new += 1

        if page_num % 100 == 0 or page_num == 1 or page_num == last_page:
            log.info(f"Page {page_num}/{last_page} : {new} nouveaux | Total : {len(all_products)}")

        time.sleep(DELAY)

    log.info("\n" + "=" * 60)
    log.info(f"TOTAL : {len(all_products)} produits uniques")
    avec_prix = sum(1 for p in all_products if p["prix"] is not None)
    avec_ean  = sum(1 for p in all_products if p["ean"])
    log.info(f"Avec prix : {avec_prix}")
    log.info(f"Avec EAN  : {avec_ean}")

    if all_products:
        export_excel(all_products, out)
        log.info(f"\n✓ Fichier : {out}")
    else:
        log.error("Aucun produit extrait.")


if __name__ == "__main__":
    main()
