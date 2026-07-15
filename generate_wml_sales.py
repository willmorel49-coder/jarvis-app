"""
Génère les données de l'app OPSO Santé depuis les stats mensuelles (STATS/*_0X_2026.xlsx) :
  - opso/wml-sales-data.js  (WML_STATIC_SALES : lignes produit par officine — vue détaillée)
  - opso/opso-adherents.js  (OPSO_ADHERENTS : adhérents OPSO ACTIFS = avec ventes)
  - opso/wml-data.js        (WML_MONTHS + WML_DATA : agrégat par officine — onglet CA & Commandes)

Périmètre = adhérents OPSO officiels (opso/opso-listing-2026.js), toutes sources commerciales
confondues (William=WML, Karine=KV, …). Un adhérent OPSO apparaît dès qu'il a des ventes.

mg (marge/remise obtenue) = (PLVPUBRUT − PLVPUNET) × PLVQTE  [dispo dans tous les exports].
Backfill EAN / ARTNATURE / AFMCODE via table ARTCODE pour les exports bruts (KV de Karine).
"""
import json, os, re, glob, warnings
from collections import defaultdict
warnings.filterwarnings('ignore')
import openpyxl

BASE = '/Users/williammorel/JARVIS/APP'
STATS = os.path.join(BASE, 'STATS')
LISTING_JS = os.path.join(BASE, 'opso', 'opso-listing-2026.js')
OUT_SALES = os.path.join(BASE, 'opso', 'wml-sales-data.js')
OUT_ADH = os.path.join(BASE, 'opso', 'opso-adherents.js')
OUT_DATA = os.path.join(BASE, 'opso', 'wml-data.js')
MONTHS = [1, 2, 3, 4, 5, 6]
WML_MONTHS = ['2026-%02d' % m for m in MONTHS]


def classify(nature, afm, pu_net):
    n = (nature or '').lower()
    a = (afm or '').lower()
    if 'biosimilaire' in n:
        return 'biosim'
    if 'generique' in n:
        return 'generique'
    if a in ('para', 'dm', 'dm_20'):
        return 'nr'
    try:
        p = float(pu_net or 0)
        if p > 468:
            return 'ch'
        if p > 4.33:
            return 'mi'
    except (TypeError, ValueError):
        pass
    return 'pp'


def tranche(remise_pct):
    if remise_pct < 3:
        return 't1'
    if remise_pct < 5:
        return 't2'
    if remise_pct < 8:
        return 't3'
    return 't4'


def fv(x):
    if x is None:
        return 'null'
    try:
        f = float(x)
        if f == int(f):
            return str(int(f))
        return f'{f:.4f}'.rstrip('0').rstrip('.')
    except (TypeError, ValueError):
        return 'null'


def r2(x):
    return round(float(x or 0), 2)


def js_str(s):
    if not s:
        return 'null'
    return json.dumps(str(s), ensure_ascii=False)


def code_str(v):
    if v is None:
        return ''
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return str(v).strip()


# ── 1. Adhérents OPSO officiels (périmètre) ──
txt = open(LISTING_JS, encoding='utf-8').read()
OPSO_INFO = {}
for m in re.finditer(r'\{[^{}]*?"cip":\s*"(\d+)"[^{}]*?\}', txt, re.S):
    blob, cip = m.group(0), m.group(1)

    def field(name, _b=blob):
        mm = re.search(r'"%s":\s*"([^"]*)"' % name, _b)
        return mm.group(1) if mm else ''
    OPSO_INFO[cip] = {'nom': field('nom'), 'cp': field('cp'), 'ville': field('ville'), 'uga': field('uga')}
OPSO_CIPS = set(OPSO_INFO.keys())
print('Adhérents OPSO officiels :', len(OPSO_CIPS))

# ── 2. Table ARTCODE/PROCODE -> {ean, nature, afm, sub} depuis les fichiers enrichis ──
ARTMAP = {}


def _put(key, ean, nat, afm, sub):
    if key in (None, ''):
        return
    key = str(key).split('.')[0]
    d = ARTMAP.setdefault(key, {})
    if ean and not d.get('ean'):
        d['ean'] = ean
    if nat and not d.get('nature'):
        d['nature'] = nat
    if afm and not d.get('afm'):
        d['afm'] = afm
    if sub and not d.get('sub'):
        d['sub'] = sub


for p in sorted(glob.glob(os.path.join(STATS, '*_0[1-6]_2026*.xlsx'))):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    hi = {n: i for i, n in enumerate(next(it))}
    if 'ARTNATURE' not in hi and 'ARTCODEBARRE' not in hi:
        wb.close()
        continue
    ia, ip = hi.get('ARTCODE'), hi.get('PROCODE')
    ib, inat, iafm, isub = hi.get('ARTCODEBARRE'), hi.get('ARTNATURE'), hi.get('AFMCODE'), hi.get('ARTSOUSFAMILLE')
    for r in it:
        def gv(i):
            return r[i] if (i is not None and i < len(r)) else None
        ean = code_str(gv(ib)) if ib is not None else ''
        for k in (gv(ia), gv(ip)):
            _put(k, ean, str(gv(inat) or ''), str(gv(iafm) or ''), str(gv(isub) or ''))
    wb.close()
print('Table ARTCODE -> attributs :', len(ARTMAP), 'entrées')

# ── 2bis. Benchmark CRM (catalogue IP) : affinage catégorie « finition JARVIS » ──
# Set des désignations NR (non remboursées SS) pour reclasser un princeps prix -> NR
# (ex. Mounjaro, Wegovy) exactement comme le CRM (classifyProduct / getBenchNrSet).
BENCH_NR = set()
try:
    btxt = open(os.path.join(BASE, 'crm', 'benchmark-data.js'), encoding='utf-8').read()
    for m in re.finditer(r'\{designation:"((?:[^"\\]|\\.)*)"[^}]*?categorie:"([a-z]+)"[^}]*?has_ameli:(true|false)', btxt):
        desig, cat, ha = m.group(1), m.group(2), m.group(3)
        if ha == 'false' and cat not in ('biosim', 'generique'):
            BENCH_NR.add(re.sub(r'\s+', ' ', desig).strip().upper())
    print('Benchmark NR (désignations) :', len(BENCH_NR))
except Exception as e:
    print('benchmark non chargé (affinage NR ignoré) :', e)


def cat_famille(designation, famille):
    """Catégorie CATS finale (finition CRM) : affine princeps prix -> NR via benchmark."""
    if famille in ('pp', 'mi', 'ch'):
        k = re.sub(r'\s+', ' ', designation).strip().upper()
        if k in BENCH_NR:
            return 'nr'
    return famille

# ── 3. Lecture des sources mensuelles, filtrées aux adhérents OPSO ──
all_sales = []       # pour wml-sales-data.js
agg = {}             # cip -> agrégat pour wml-data.js
idx = 0


def new_ph(cip):
    return {'tc': cip, 'nom': OPSO_INFO.get(cip, {}).get('nom', 'Officine ' + cip),
            'ca': 0.0, 'mg': 0.0, 'qt': 0.0,
            'ca_m': [0.0] * len(MONTHS), 'mg_m': [0.0] * len(MONTHS), 'qt_m': [0.0] * len(MONTHS),
            'afm': defaultdict(lambda: [0.0, 0.0, 0.0]),
            'cat': defaultdict(lambda: [0.0, 0.0, 0.0]),
            'sf': defaultdict(lambda: [0.0, 0.0, 0.0]),
            'tr': {'t1': [0, 0.0, 0.0], 't2': [0, 0.0, 0.0], 't3': [0, 0.0, 0.0], 't4': [0, 0.0, 0.0]},
            'prod': defaultdict(lambda: [0.0, 0.0, 0.0, ''])}  # nom -> [ca,mg,qt,ean]


for path in sorted(glob.glob(os.path.join(STATS, '*_0[1-6]_2026.xlsx'))):
    base = os.path.basename(path)
    mm = re.match(r'[A-Za-z]+_(0[1-6])_2026\.xlsx$', base)
    if not mm:
        continue
    month = int(mm.group(1))
    if month not in MONTHS:
        continue
    mi = MONTHS.index(month)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    col = {n: i for i, n in enumerate(next(it)) if n}

    def g(row, k, _c=col):
        i = _c.get(k)
        return row[i] if (i is not None and i < len(row)) else None

    added = 0
    for row in it:
        cip = code_str(g(row, 'TIRCODE'))
        if not cip or cip not in OPSO_CIPS:
            continue
        designation = str(g(row, 'PLVDESIGNATION') or '').strip()
        if not designation:
            continue
        art_code = code_str(g(row, 'ARTCODE'))
        pro_code = code_str(g(row, 'PROCODE'))
        amap = ARTMAP.get(art_code) or ARTMAP.get(pro_code) or {}

        def num(k):
            try:
                return float(g(row, k) or 0)
            except (TypeError, ValueError):
                return 0.0
        qte, pu_brut, pu_net, mnt = num('PLVQTE'), num('PLVPUBRUT'), num('PLVPUNET'), num('PLVMNTNETHT')
        mg_line = (pu_brut - pu_net) * qte
        ean = code_str(g(row, 'ARTCODEBARRE')) or amap.get('ean', '')
        nature = g(row, 'ARTNATURE') or amap.get('nature') or ''
        afm = str(g(row, 'AFMCODE') or amap.get('afm') or '').strip() or 'AUTRE'
        subfamily = str(g(row, 'ARTSOUSFAMILLE') or amap.get('sub') or '')
        famille = classify(str(nature), afm, pu_net)
        art_famille = 'froid' if subfamily.lower() == 'froid' else famille

        # -- wml-sales-data.js --
        all_sales.append({'id': f'wml_{cip}_{month}_2026_{idx}', 'pharmacyCode': cip, 'month': month,
                          'artDesignation': designation, 'artCode': art_code, 'artFamille': art_famille,
                          'qte': qte, 'puBrut': pu_brut, 'puNet': pu_net, 'mntNetHt': mnt})
        idx += 1

        # -- wml-data.js (agrégat) --
        a = agg.get(cip) or agg.setdefault(cip, new_ph(cip))
        a['ca'] += mnt; a['mg'] += mg_line; a['qt'] += qte
        a['ca_m'][mi] += mnt; a['mg_m'][mi] += mg_line; a['qt_m'][mi] += qte
        a['afm'][afm][0] += mnt; a['afm'][afm][1] += mg_line; a['afm'][afm][2] += qte
        fc = cat_famille(designation, famille)   # catégorie produit CATS (finition CRM)
        a['cat'][fc][0] += mnt; a['cat'][fc][1] += mg_line; a['cat'][fc][2] += qte
        if subfamily:
            a['sf'][subfamily][0] += mnt; a['sf'][subfamily][1] += mg_line; a['sf'][subfamily][2] += qte
        t = tranche(mg_line / mnt * 100 if mnt > 0 else 0)
        a['tr'][t][0] += 1; a['tr'][t][1] += mnt; a['tr'][t][2] += mg_line
        pr = a['prod'][designation]
        pr[0] += mnt; pr[1] += mg_line; pr[2] += qte
        if ean and not pr[3]:
            pr[3] = ean
        added += 1
    wb.close()
    if added:
        print('  %-26s : %d lignes OPSO' % (base, added))

print('\nTotal : %d lignes · %d adhérents OPSO actifs' % (len(all_sales), len(agg)))

# ── 4. wml-sales-data.js ──
lines = [
    '// WML Sales Data — app OPSO Santé (généré par generate_wml_sales.py)',
    '// %d lignes · %d adhérents actifs · Jan-Juin 2026 (William + Karine…)' % (len(all_sales), len(agg)),
    '// pharmacyCode = CIP, mappé à pharmacyId au chargement (initApp)',
    'const WML_STATIC_SALES = [',
]
for s in all_sales:
    lines.append('{'
                 f'id:{js_str(s["id"])},pharmacyCode:{js_str(s["pharmacyCode"])},importId:null,'
                 f'month:{s["month"]},year:2026,'
                 f'artDesignation:{js_str(s["artDesignation"])},artCode:{js_str(s["artCode"])},artId:null,'
                 f'artFamille:{js_str(s["artFamille"])},qte:{fv(s["qte"])},puBrut:{fv(s["puBrut"])},'
                 f'puNet:{fv(s["puNet"])},mntNetHt:{fv(s["mntNetHt"])}' '},')
lines.append('];')
open(OUT_SALES, 'w', encoding='utf-8').write('\n'.join(lines))
print('→ %s (%d Ko)' % (os.path.relpath(OUT_SALES, BASE), os.path.getsize(OUT_SALES) // 1024))

# ── 5. wml-data.js (agrégat par officine, onglet CA & Commandes) ──
def obj_num_map(d):
    return '{' + ','.join("%s:[%s,%s,%s]" % (js_str(k), r2(v[0]), r2(v[1]), r2(v[2])) for k, v in d.items()) + '}'


order = sorted(agg.keys(), key=lambda c: -agg[c]['ca'])
dlines = [
    '// WML Commandes — app OPSO Santé (généré par generate_wml_sales.py)',
    '// %d officines · %d mois (%s → %s) · William + Karine' % (len(agg), len(MONTHS), WML_MONTHS[0], WML_MONTHS[-1]),
    '// Champs : tc nom ca mg qt ca_m/mg_m/qt_m[%d]  afm{code:[ca,mg,qt]}  sf{sf:[ca,mg,qt]}  pr[[nom,ca,mg,qt,ean]]  tr{t1..t4:[n,ca,mg]}' % len(MONTHS),
    'const WML_MONTHS = %s;' % json.dumps(WML_MONTHS),
    'const WML_DATA = [',
]
for cip in order:
    a = agg[cip]
    prods = sorted(([nom] + v for nom, v in a['prod'].items()), key=lambda x: -x[1])[:30]
    pr_js = '[' + ','.join('[%s,%s,%s,%s,%s]' % (js_str(p[0]), r2(p[1]), r2(p[2]), r2(p[3]), js_str(p[4])) for p in prods) + ']'
    tr = a['tr']
    tr_js = '{' + ','.join("%s:[%d,%s,%s]" % (t, tr[t][0], r2(tr[t][1]), r2(tr[t][2])) for t in ('t1', 't2', 't3', 't4')) + '}'
    dlines.append(
        '{tc:%s,nom:%s,ca:%s,mg:%s,qt:%s,ca_m:%s,mg_m:%s,qt_m:%s,afm:%s,cat:%s,sf:%s,pr:%s,tr:%s},' % (
            cip, js_str(a['nom']), r2(a['ca']), r2(a['mg']), r2(a['qt']),
            json.dumps([r2(x) for x in a['ca_m']]), json.dumps([r2(x) for x in a['mg_m']]), json.dumps([r2(x) for x in a['qt_m']]),
            obj_num_map(a['afm']), obj_num_map(a['cat']), obj_num_map(a['sf']), pr_js, tr_js))
dlines.append('];')
dlines.append('try{window.WML_MONTHS=WML_MONTHS;window.WML_DATA=WML_DATA;}catch(e){}')
open(OUT_DATA, 'w', encoding='utf-8').write('\n'.join(dlines))
print('→ %s (%d Ko)' % (os.path.relpath(OUT_DATA, BASE), os.path.getsize(OUT_DATA) // 1024))

# ── 6. opso-adherents.js (actifs, triés par CA) ──
adh = [{'cip': c, **{k: OPSO_INFO.get(c, {}).get(k, '') for k in ('nom', 'cp', 'ville', 'uga')}} for c in order]
with open(OUT_ADH, 'w', encoding='utf-8') as f:
    f.write('// Listing adhérents OPSO Santé ACTIFS (avec ventes) — généré par generate_wml_sales.py\n')
    f.write('// %d pharmacies actives (William + Karine…)\n' % len(adh))
    f.write('const OPSO_ADHERENTS = [\n')
    for a in adh:
        f.write('  {cip:%s,nom:%s,cp:%s,ville:%s,uga:%s},\n' % (
            js_str(a['cip']), js_str(a['nom'] or 'Officine ' + a['cip']), js_str(a['cp']), js_str(a['ville']), js_str(a['uga'])))
    f.write('];\n')
    f.write('try{window.OPSO_ADHERENTS=OPSO_ADHERENTS;}catch(e){}\n')
print('→ %s (%d adhérents actifs)' % (os.path.relpath(OUT_ADH, BASE), len(adh)))
print('\nCA groupement total : %.0f € · marge (remise) : %.0f € · %d officines' % (
    sum(a['ca'] for a in agg.values()), sum(a['mg'] for a in agg.values()), len(agg)))
for cip in order:
    a = agg[cip]
    print('  %s  %-38s  CA %10.0f €  mg %8.0f €' % (cip, a['nom'][:38], a['ca'], a['mg']))
print('Terminé.')
