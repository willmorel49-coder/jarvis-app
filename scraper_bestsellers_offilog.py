#!/usr/bin/env python3
"""
Scraper Offilog — Meilleures ventes
Connexion : identifiants hors dépôt (~/.config/jarvis/offilog.json)
Préserve l'ordre d'affichage = rang de vente décroissant
Output    : bestsellers_offilog.json  +  bestsellers_offilog.xlsx
"""

# ══════════════════════════════════════════════════════════════════════════
# ⚠️ 03/09/2026 — CE SCRIPT EST REMPLACÉ, ET IL EST DANGEREUX DE LE LANCER.
#
# Il écrit les PRIX B2B directement dans un fichier du dépôt, qui est PUBLIC
# et servi par GitHub Pages. C'est exactement la fuite qu'on vient de fermer :
# le relancer la rouvrirait, sans le moindre message d'erreur.
#
# À utiliser à la place — même travail, mais par navigateur réel (Playwright,
# le site rend sa liste en JavaScript) ET séparant le public du protégé :
#   node scraper_offilog_playwright.js
# ══════════════════════════════════════════════════════════════════════════
import sys as _sys
print(__doc__ or '')
print("ARRÊT : ce script réécrirait les prix B2B dans un fichier PUBLIC.")
print("Utiliser :  node scraper_offilog_playwright.js")
_sys.exit(2)

import re
import sys
import json
import time
import openpyxl
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from datetime import datetime

BASE     = Path(__file__).parent
OUT_JSON = BASE / 'bestsellers_offilog.json'
OUT_XLSX = BASE / 'bestsellers_offilog.xlsx'

LOGIN_URL = 'https://offilog.fr/connexion?back=my-account'
# ── Identifiants ────────────────────────────────────────────────────────────
# ⚠️ 03/09/2026 — ils étaient ÉCRITS EN CLAIR ici, dans un dépôt PUBLIC servi
# par GitHub Pages. N'importe qui pouvait se connecter à la plateforme Offilog.
# Ils se lisent désormais hors du dépôt : ~/.config/jarvis/offilog.json
#   {"email": "...", "password": "..."}
# ou, à défaut, dans les variables d'environnement OFFILOG_EMAIL / OFFILOG_PWD.
import os, json as _json
from pathlib import Path as _Path

def _identifiants():
    f = _Path.home() / '.config' / 'jarvis' / 'offilog.json'
    if f.exists():
        d = _json.loads(f.read_text(encoding='utf-8'))
        return d.get('email', ''), d.get('password', '')
    return os.environ.get('OFFILOG_EMAIL', ''), os.environ.get('OFFILOG_PWD', '')

EMAIL, PASSWORD = _identifiants()
if not EMAIL or not PASSWORD:
    raise SystemExit(
        "Identifiants Offilog absents. Crée ~/.config/jarvis/offilog.json "
        '({"email": "...", "password": "..."}) ou exporte OFFILOG_EMAIL / OFFILOG_PWD. '
        "Ne JAMAIS les réécrire dans ce fichier : le dépôt est public.")

BEST_URL  = 'https://offilog.fr/meilleures-ventes'
PER_PAGE  = 100
DELAY     = 0.5


def make_session():
    s = requests.Session()
    s.headers.update({'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'})
    r = s.post(LOGIN_URL, data={'back': 'my-account', 'email': EMAIL, 'password': PASSWORD, 'submitLogin': '1'}, timeout=20)
    if 'Déconnexion' not in r.text and 'mon-compte' not in r.url:
        print('ERREUR : login échoué')
        sys.exit(1)
    print(f'  ✓ Connecté en tant que {EMAIL}')
    return s


def parse_price(text):
    if not text:
        return None
    text = text.strip().replace('\xa0', '').replace(' ', '').replace(',', '.').replace('€', '')
    try:
        return round(float(text), 4)
    except ValueError:
        return None


def scrape_page(s, url, rank_offset=0):
    """Retourne (produits avec rang, total_count)."""
    time.sleep(DELAY)
    try:
        r = s.get(url, timeout=20)
    except Exception as e:
        print(f'    ERREUR requête {url}: {e}')
        return [], 0

    soup = BeautifulSoup(r.text, 'html.parser')
    articles = soup.find_all('article', class_='js-product-miniature')

    total = 0
    total_el = soup.find(class_='total-products')
    if total_el:
        m = re.search(r'(\d[\d\s]*)\s*article', total_el.get_text())
        if m:
            total = int(m.group(1).replace(' ', '').replace('\xa0', ''))

    products = []
    for i, art in enumerate(articles):
        rang = rank_offset + i + 1
        prod_id = art.get('data-id-product', '')

        name_a = art.find('a', class_='product_name')
        nom = name_a.get_text(strip=True) if name_a else ''
        prod_url = name_a['href'] if name_a and name_a.get('href') else ''

        ean = ''
        m = re.search(r'-(\d{8,13})\.html', prod_url)
        if m:
            ean = m.group(1)

        marque_el = art.find(class_='manufacturer')
        marque = marque_el.get_text(strip=True) if marque_el else ''

        prix_el = art.find(itemprop='price') or art.find(class_='price')
        prix = parse_price(prix_el.get_text()) if prix_el else None

        prix_barre_el = art.find(class_='regular-price') or art.find(class_='old-price')
        prix_barre = parse_price(prix_barre_el.get_text()) if prix_barre_el else None

        img_el = art.find('img')
        img_url = img_el.get('src', '') if img_el else ''

        flags = [f.get_text(strip=True) for f in art.find_all(class_='product-flag')]

        products.append({
            'rang':       rang,
            'id_product': prod_id,
            'nom':        nom,
            'marque':     marque,
            'ean':        ean,
            'prix':       prix,
            'prix_barre': prix_barre,
            'en_promo':   prix_barre is not None and prix_barre > 0,
            'flags':      ','.join(flags),
            'url':        prod_url,
            'img':        img_url,
        })

    return products, total


def main():
    print('=' * 60)
    print('SCRAPER OFFILOG — MEILLEURES VENTES')
    print(f'Démarré le {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 60)

    s = make_session()

    print(f'\n  Scraping : {BEST_URL}')

    # Page 1
    url1 = f'{BEST_URL}?resultsPerPage={PER_PAGE}&page=1'
    prods, total = scrape_page(s, url1, rank_offset=0)
    if not prods:
        print('  → Page vide ou inaccessible')
        sys.exit(1)

    # Calcul pages : si total détecté → ceil(total/PER_PAGE), sinon on scrape jusqu'à page vide
    if total > 0:
        n_pages = -(-total // PER_PAGE)
        print(f'  → {total} produits détectés · {n_pages} page(s)')
    else:
        n_pages = None  # inconnu, mode "until empty"
        print(f'  → Total non détecté — scraping jusqu\'à page vide (mode exhaustif)')

    all_prods = list(prods)
    p = 2
    while True:
        if n_pages is not None and p > n_pages:
            break
        url_p = f'{BEST_URL}?resultsPerPage={PER_PAGE}&page={p}'
        print(f'    Page {p}{f"/{n_pages}" if n_pages else ""}…  ({len(all_prods)} produits)', end='', flush=True)
        pp, _ = scrape_page(s, url_p, rank_offset=len(all_prods))
        if not pp:
            print(f'  → Page vide, fin.')
            break
        all_prods.extend(pp)
        print(f'  +{len(pp)}')
        p += 1

    print(f'\n  → {len(all_prods)} produits récupérés (ordre = rang de vente décroissant)')

    # ── JSON ──────────────────────────────────────────────────────
    meta = {
        'generated':   datetime.now().isoformat(),
        'source':      'offilog.fr/meilleures-ventes',
        'compte':      EMAIL,
        'count':       len(all_prods),
        'description': 'Produits classés par rang de vente décroissant (rang 1 = meilleure vente)',
    }
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump({'meta': meta, 'products': all_prods}, f, ensure_ascii=False, indent=2)
    print(f'JSON → {OUT_JSON}')

    # ── Excel ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Meilleures Ventes'
    headers = ['rang', 'id_product', 'nom', 'marque', 'ean', 'prix', 'prix_barre', 'en_promo', 'flags', 'url', 'img']
    ws.append(headers)
    for prod in all_prods:
        ws.append([prod.get(h, '') for h in headers])

    # Style entête
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='FF6B35')
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['J'].width = 60
    ws.column_dimensions['K'].width = 60
    wb.save(OUT_XLSX)
    print(f'Excel → {OUT_XLSX}')

    # ── Stats ─────────────────────────────────────────────────────
    avec_prix = sum(1 for p in all_prods if p['prix'])
    avec_ean  = sum(1 for p in all_prods if p['ean'])
    avec_img  = sum(1 for p in all_prods if p['img'])
    en_promo  = sum(1 for p in all_prods if p['en_promo'])
    print(f'\n{"="*60}')
    print(f'TOTAL       : {len(all_prods)} produits')
    print(f'Avec prix   : {avec_prix} ({avec_prix/len(all_prods)*100:.0f}%)')
    print(f'Avec EAN    : {avec_ean}  ({avec_ean/len(all_prods)*100:.0f}%)')
    print(f'Avec image  : {avec_img}  ({avec_img/len(all_prods)*100:.0f}%)')
    print(f'En promo    : {en_promo}')
    print(f'\nTop 10 :')
    for p in all_prods[:10]:
        print(f'  #{p["rang"]:>3}  {p["marque"]:<20} {p["nom"][:45]:<45}  {p["prix"]} €')


if __name__ == '__main__':
    main()
