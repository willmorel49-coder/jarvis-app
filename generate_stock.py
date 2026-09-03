#!/usr/bin/env python3
"""
generate_stock.py
=================
Lit : STATS/stock01 06 2026.xlsx
Filtre : stockdispo > 0
Écrit : crm/stock.js

Format JS :
  window.STOCK = {
    "artcode": { nom, ean, marque, dispo, commande, mol, sousfamille, nature, categorie, collection, classe },
    ...
  }
  window.STOCK_EAN = { "ean": "artcode", ... }   // index secondaire EAN → artcode
  window.STOCK_DATE = "2026-06-01"

Colonnes source :
  artcode, artdesignation, artsousfamille, artnature, artcategorie,
  artcollection, artcodebarre, afmcode, stockcde, stockdispo, artclasse, artmol
"""

from pathlib import Path
import datetime
import openpyxl
import json
import re

BASE   = Path(__file__).parent
OUT    = BASE / "crm" / "stock.js"


def dernier_export():
    """Le fichier de stock LE PLUS RÉCENT, et sa date lue dans son NOM.

    ⚠️ Le chemin était écrit en dur (« STATS/stock01 06 2026.xlsx »). Le 03/09/2026 Will a
    déposé un export frais dans STOCKS/ : le robot ne l'aurait jamais vu et aurait continué
    à publier un inventaire vieux de 94 jours, sans la moindre erreur à l'écran. Un chemin
    en dur est un robot qui ment en silence.

    ⚠️ Et la date se lit dans le NOM du fichier, jamais dans sa date de modification : un
    fichier recopié ou synchronisé porte la date de la copie, pas celle de l'inventaire.
    Formats acceptés : « stock01 06 2026.xlsx », « stock OPS 03 09 2026.xlsx »,
    « stock et prix 22 06 2026.xlsx ».
    """
    cands = []
    for dossier in ("STOCKS", "STATS"):
        d = BASE / dossier
        if not d.is_dir():
            continue
        for f in d.glob("*.xlsx"):
            if f.name.startswith("~$") or "stock" not in f.name.lower():
                continue
            m = re.search(r"(\d{2})[ _-](\d{2})[ _-](\d{4})", f.name)
            if not m:
                continue
            j, mo, a = m.groups()
            try:
                dt = datetime.date(int(a), int(mo), int(j))
            except ValueError:
                continue
            cands.append((dt, f))
    if not cands:
        raise SystemExit("[stock] aucun export trouvé dans STOCKS/ ni STATS/")
    cands.sort(key=lambda x: x[0])
    dt, f = cands[-1]
    autres = [c[1].name for c in cands[:-1]]
    print(f"[stock] {len(cands)} export(s) trouvé(s) · retenu : {f.name} (inventaire du {dt})")
    if autres:
        print(f"[stock] plus anciens, ignorés : {', '.join(autres)}")
    age = (datetime.date.today() - dt).days
    if age > 35:
        print(f"[stock] ⚠️ cet inventaire a {age} jours — les couvertures seront optimistes")
    return f, dt.isoformat()


SRC, DATE = dernier_export()

def clean_str(v):
    if v is None:
        return ""
    return str(v).strip()

def clean_int(v):
    if v is None:
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0

def main():
    print(f"[stock] Lecture : {SRC}")
    wb = openpyxl.load_workbook(str(SRC), read_only=True, data_only=True)
    ws = wb.active

    # --- Headers ---
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    print(f"[stock] Colonnes détectées : {headers}")

    # Index colonnes
    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx = {
        "artcode":       col("artcode"),
        "artdesignation":col("artdesignation"),
        "artsousfamille":col("artsousfamille"),
        "artnature":     col("artnature"),
        "artcategorie":  col("artcategorie"),
        "artcollection": col("artcollection"),
        "artcodebarre":  col("artcodebarre"),
        "afmcode":       col("afmcode"),
        "stockcde":      col("stockcde"),
        "stockdispo":    col("stockdispo"),
        "artclasse":     col("artclasse"),
        "artmol":        col("artmol"),
    }
    print(f"[stock] Index colonnes : {idx}")

    stock = {}      # artcode -> dict
    ean_index = {}  # ean -> artcode

    total_rows = 0
    skipped_zero = 0
    skipped_no_code = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1

        artcode = clean_str(row[idx["artcode"]]) if idx["artcode"] is not None else ""
        if not artcode:
            skipped_no_code += 1
            continue

        stockdispo = clean_int(row[idx["stockdispo"]]) if idx["stockdispo"] is not None else 0
        if stockdispo <= 0:
            skipped_zero += 1
            continue

        ean = clean_str(row[idx["artcodebarre"]]) if idx["artcodebarre"] is not None else ""
        # Normalise EAN : garde seulement chiffres
        ean_clean = "".join(c for c in ean if c.isdigit())

        entry = {
            "nom":        clean_str(row[idx["artdesignation"]]) if idx["artdesignation"] is not None else "",
            "ean":        ean_clean,
            "marque":     clean_str(row[idx["afmcode"]]) if idx["afmcode"] is not None else "",
            "dispo":      stockdispo,
            "commande":   clean_int(row[idx["stockcde"]]) if idx["stockcde"] is not None else 0,
            "mol":        clean_str(row[idx["artmol"]]) if idx["artmol"] is not None else "",
            "sousfamille":clean_str(row[idx["artsousfamille"]]) if idx["artsousfamille"] is not None else "",
            "nature":     clean_str(row[idx["artnature"]]) if idx["artnature"] is not None else "",
            "categorie":  clean_str(row[idx["artcategorie"]]) if idx["artcategorie"] is not None else "",
            "collection": clean_str(row[idx["artcollection"]]) if idx["artcollection"] is not None else "",
            "classe":     clean_str(row[idx["artclasse"]]) if idx["artclasse"] is not None else "",
        }

        stock[artcode] = entry

        if ean_clean and len(ean_clean) >= 8:
            ean_index[ean_clean] = artcode

    wb.close()

    # --- Stats ---
    total_refs = len(stock)
    total_units = sum(e["dispo"] for e in stock.values())
    top5 = sorted(stock.items(), key=lambda x: x[1]["dispo"], reverse=True)[:5]

    print(f"\n[stock] Résultat :")
    print(f"  Lignes lues         : {total_rows}")
    print(f"  Filtrées (dispo=0)  : {skipped_zero}")
    print(f"  Filtrées (no code)  : {skipped_no_code}")
    print(f"  Références en stock : {total_refs}")
    print(f"  Total unités        : {total_units:,}")
    print(f"  EAN indexés         : {len(ean_index)}")
    print(f"\n  Top 5 par dispo :")
    for code, e in top5:
        print(f"    {code} | {e['nom'][:50]} | dispo={e['dispo']:,}")

    # --- Génération JS ---
    # Serialize stock entries en JS object literals
    lines = []
    for artcode, e in stock.items():
        # Échappement minimal JSON-safe
        def js_str(s):
            return json.dumps(s, ensure_ascii=False)

        line = (
            f'  {js_str(artcode)}: {{'
            f' nom: {js_str(e["nom"])},'
            f' ean: {js_str(e["ean"])},'
            f' marque: {js_str(e["marque"])},'
            f' dispo: {e["dispo"]},'
            f' commande: {e["commande"]},'
            f' mol: {js_str(e["mol"])},'
            f' sousfamille: {js_str(e["sousfamille"])},'
            f' nature: {js_str(e["nature"])},'
            f' categorie: {js_str(e["categorie"])},'
            f' collection: {js_str(e["collection"])},'
            f' classe: {js_str(e["classe"])}'
            f' }}'
        )
        lines.append(line)

    # EAN index lines
    ean_lines = []
    for ean, artcode in ean_index.items():
        ean_lines.append(f'  {json.dumps(ean, ensure_ascii=False)}: {json.dumps(artcode, ensure_ascii=False)}')

    js_content = """// Inventaire stock Intégral Pharma au ___DATE___
// Source : ___SRC___
// Généré par generate_stock.py
// Références en stock : ___REFS___ | Total unités : ___UNITS___
window.STOCK = {
___LINES___
};

// Index secondaire EAN → artcode (pour cross-référence catalogue)
window.STOCK_EAN = {
___EAN_LINES___
};

// Date du stock
window.STOCK_DATE = "___DATE___";
"""
    js_content = (js_content
        .replace("___REFS___", str(total_refs))
        .replace("___UNITS___", f"{total_units:,}")
        .replace("___LINES___", ",\n".join(lines))
        .replace("___EAN_LINES___", ",\n".join(ean_lines))
        .replace("___DATE___", DATE)
        .replace("___SRC___", f"{SRC.parent.name}/{SRC.name}"))

    OUT.write_text(js_content, encoding="utf-8")
    size_kb = OUT.stat().st_size / 1024
    print(f"\n[stock] Fichier généré : {OUT}")
    print(f"[stock] Taille         : {size_kb:.1f} KB")
    print(f"\n[stock] DONE")

    # Retour structuré pour le rapport
    return {
        "refs": total_refs,
        "units": total_units,
        "eans": len(ean_index),
        "size_kb": size_kb,
        "top5": [(code, e["nom"], e["dispo"]) for code, e in top5],
    }

if __name__ == "__main__":
    main()
