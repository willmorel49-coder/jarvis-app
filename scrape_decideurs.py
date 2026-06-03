#!/usr/bin/env python3
"""
Scraper décideurs groupements pharmaceutiques — V2
Sources : LaboData → Sirene → Societe.com → Sites officiels (deep)
"""

import json
import re
import time
import unicodedata
import logging
from datetime import date
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LETTERS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M',
           'N', 'O', 'P', 'R', 'S', 'U', 'V', 'W']

LABODATA_BASE = 'https://www.labodata.com'
SIRENE_API    = 'https://recherche-entreprises.api.gouv.fr/search'
SOCIETE_BASE  = 'https://www.societe.com'

UA            = 'PharmaResearch/1.0'
DELAY_LABODATA = 1.0
DELAY_SITE     = 2.0
DELAY_SOCIETE  = 1.5
SIRENE_DELAY   = 0.15
TIMEOUT        = 15

OUTPUT_DIR       = Path('output')
CACHE_LABODATA   = Path('cache_labodata.json')
CACHE_SIRENE     = Path('cache_sirene.json')
CACHE_SOCIETE    = Path('cache_societe.json')
CACHE_SITES      = Path('cache_sites.json')

FONCTIONS_KEYWORDS = [
    'président', 'presidente', 'directeur général', 'directrice générale',
    'directeur generale', 'directrice generale', 'dg', 'gérant', 'gerant',
    'directeur commercial', 'directrice commerciale',
    'directeur adhésions', 'directrice adhésions',
    'responsable adhésions', 'responsable commercial',
    'fondateur', 'co-fondateur', 'cofondateur', 'ceo',
    'directeur général délégué', 'directeur regional',
]

TEAM_SLUGS = [
    '/equipe', '/notre-equipe', '/team', '/direction', '/gouvernance',
    '/bureau', '/conseil', '/qui-sommes-nous', '/a-propos', '/about',
    '/le-groupement', '/presentation', '/contact', '/nous-contacter',
]

TEAM_PAGE_KEYWORDS = [
    'équipe', 'equipe', 'team', 'qui sommes-nous', 'qui-sommes-nous',
    'gouvernance', 'direction', 'notre équipe', 'à propos', 'a-propos',
    'contact', 'bureau', 'conseil', 'dirigeant',
]

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger(__name__)

# ── Helpers ───────────────────────────────────────────────────

def get(url, delay=0, session=None, verify=True):
    if delay:
        time.sleep(delay)
    s = session or requests.Session()
    try:
        r = s.get(url, headers={'User-Agent': UA}, timeout=TIMEOUT, verify=verify)
        r.raise_for_status()
        return r
    except Exception as e:
        log.debug(f'GET {url} → {e}')
        return None


def load_cache(path):
    if path.exists():
        return json.loads(path.read_text('utf-8'))
    return {}


def save_cache(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), 'utf-8')


def load_or_empty(path, label):
    if path.exists():
        data = load_cache(path)
        log.info(f'Cache {label} repris ({len(data)} entrées)')
        return data
    return {}


def normalize(s):
    return re.sub(r'\s+', ' ', (s or '').lower().strip())


def slugify(s):
    s = unicodedata.normalize('NFKD', (s or '')).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]+', '-', s.lower()).strip('-')


def extract_emails(text):
    return list(set(re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', text)))


def extract_phones(text):
    return list(set(re.findall(r'(?:(?:\+33|0033)\s?|0)[1-9](?:[\s.\-]?\d{2}){4}', text)))


def check_robots(base_url, path='/', session=None):
    """Retourne False si Disallow: {path} est dans le bloc User-agent: *."""
    try:
        r = get(f'{base_url}/robots.txt', session=session)
        if not r:
            return True
        current_agents = []
        for line in r.text.splitlines():
            line = line.strip()
            if line.lower().startswith('user-agent:'):
                agent = line.split(':', 1)[1].strip()
                current_agents.append(agent)
            elif line.lower().startswith('disallow:'):
                disallow_path = line.split(':', 1)[1].strip()
                if '*' in current_agents and disallow_path and path.startswith(disallow_path):
                    return False
            elif line == '':
                current_agents = []
    except Exception:
        pass
    return True

# ── Étape 1 : LaboData ────────────────────────────────────────

def scrape_labodata(letters, cache):
    session = requests.Session()
    if not check_robots(LABODATA_BASE, '/', session):
        log.error('LaboData robots.txt interdit le scraping.')
        return cache

    results = dict(cache)
    for letter in letters:
        url = f'{LABODATA_BASE}/annuaire-des-groupements-de-pharmacies/{letter}'
        log.info(f'LaboData lettre {letter} → {url}')
        r = get(url, delay=DELAY_LABODATA, session=session)
        if not r:
            continue

        soup = BeautifulSoup(r.text, 'lxml')
        cards = soup.find_all('div', class_=lambda c: c and 'card' in c and 'mb-4' in c)
        log.info(f'  → {len(cards)} groupements lettre {letter}')

        for card in cards:
            card_id = card.get('id', '')
            key = f'/{letter}/{card_id}'
            if key in results:
                continue

            title_el = card.find('h2') or card.find('strong')
            nom = title_el.get_text(strip=True) if title_el else card_id

            site_web = None
            for a in card.find_all('a', href=True):
                href = a['href']
                if href.startswith('http') and LABODATA_BASE not in href:
                    site_web = href
                    break

            mail_a = card.find('a', href=lambda h: h and h.startswith('mailto:'))
            email  = mail_a['href'].replace('mailto:', '') if mail_a else None

            footer = card.find('div', class_='card-footer')
            tel    = extract_phones(footer.get_text())[0] if footer else None

            results[key] = {
                'url_fiche': url + '#' + card_id,
                'nom': nom,
                'site_web': site_web,
                'email_general': email,
                'tel_general': tel,
            }
            log.info(f'    {nom} | {site_web or "-"} | {email or "-"}')

    save_cache(CACHE_LABODATA, results)
    return results

# ── Étape 2 : Sirene ─────────────────────────────────────────

def sirene_score(nom_query, nom_result):
    nq, nr = normalize(nom_query), normalize(nom_result)
    if nq == nr:
        return 100
    if nq in nr or nr in nq:
        return 50
    return 0


def search_sirene(nom, cache):
    if nom in cache:
        return cache[nom]
    time.sleep(SIRENE_DELAY)
    try:
        r = requests.get(SIRENE_API, params={'q': nom, 'per_page': 5},
                         headers={'User-Agent': UA}, timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        log.warning(f'Sirene "{nom}" → {e}')
        cache[nom] = None
        return None

    best, best_score = None, -1
    for res in data.get('results', []):
        score = sirene_score(nom, res.get('nom_complet', ''))
        if res.get('etat_administratif') == 'A':
            score += 5
        if score > best_score:
            best_score, best = score, res

    if best is None or best_score < 10:
        cache[nom] = None
        return None

    siege = best.get('siege', {})
    dirigeants = []
    for d in best.get('dirigeants', []):
        prenom = d.get('prenoms', '')
        nom_d  = d.get('nom', '') or d.get('denomination', '')
        full   = f"{prenom} {nom_d}".strip()
        if full:
            dirigeants.append({'nom': full, 'qualite': d.get('qualite', '')})

    result = {
        'siren':           best.get('siren'),
        'raison_sociale':  best.get('nom_complet'),
        'nature_juridique': best.get('nature_juridique'),
        'code_ape':        best.get('activite_principale'),
        'statut':          'Actif' if best.get('etat_administratif') == 'A' else 'Cessé',
        'adresse_siege':   f"{siege.get('code_postal','')} {siege.get('libelle_commune','')}".strip(),
        'dirigeants':      dirigeants,
    }
    cache[nom] = result
    return result

# ── Étape 3 : Societe.com (via SIREN) ────────────────────────

def scrape_societe(nom, siren, raison_sociale, cache):
    if nom in cache:
        return cache[nom]
    if not siren:
        cache[nom] = None
        return None

    slug = slugify(raison_sociale or nom)
    url  = f'{SOCIETE_BASE}/societe/{slug}-{siren}.html'

    if not check_robots(SOCIETE_BASE, f'/societe/'):
        cache[nom] = None
        return None

    time.sleep(DELAY_SOCIETE)
    r = get(url)
    if not r:
        # Fallback : essai avec slugify du nom LaboData
        url2 = f'{SOCIETE_BASE}/societe/{slugify(nom)}-{siren}.html'
        r = get(url2)
        if not r:
            cache[nom] = None
            return None
        url = url2

    soup = BeautifulSoup(r.text, 'lxml')

    # Chaque dirigeant est dans un <article> avec span.ui-label + <p> fonction
    dirigeants = []
    for article in soup.find_all('article'):
        label = article.find(class_='ui-label')
        if not label:
            continue
        nom_dirigeant = label.get_text(strip=True)
        # Paragraphes : naissance, fonction
        paras = article.find_all('p')
        fonction, depuis = '', ''
        for p in paras:
            txt = p.get_text(' ', strip=True)
            # "Président Depuis le 04 avril 2023"
            if any(f in txt.lower() for f in FONCTIONS_KEYWORDS + ['directeur', 'gérant', 'président']):
                sub = p.find(class_='ui-sub')
                if sub:
                    depuis = sub.get_text(strip=True)
                    sub.extract()
                fonction = p.get_text(strip=True)
        if nom_dirigeant:
            dirigeants.append({'nom': nom_dirigeant, 'qualite': fonction, 'depuis': depuis})

    if not dirigeants:
        # Fallback : regex sur texte brut
        text = soup.get_text(' ')
        for match in re.finditer(
            r'([A-Z][A-Z\s\-\.\']{3,40})\s+Né[e]? en \d{4}[^\.]*\.\s*([^\.\n]{5,60})',
            text
        ):
            dirigeants.append({'nom': match.group(1).strip(), 'qualite': match.group(2).strip(), 'depuis': ''})

    result = {'url': url, 'dirigeants': dirigeants[:6]}
    log.info(f'    Societe.com {nom} → {len(dirigeants)} dirigeant(s)')
    cache[nom] = result
    return result

# ── Étape 4 : Scrape approfondi sites officiels ───────────────

def get_sitemap_team_pages(base_url, session):
    """Parse sitemap.xml pour trouver pages équipe/direction."""
    for sitemap_path in ['/sitemap.xml', '/sitemap_index.xml']:
        r = get(base_url + sitemap_path, session=session)
        if not r or '<urlset' not in r.text and '<sitemapindex' not in r.text:
            continue
        soup = BeautifulSoup(r.text, 'lxml-xml')
        urls = [loc.get_text() for loc in soup.find_all('loc')]
        matches = [u for u in urls if any(kw in u.lower() for kw in TEAM_PAGE_KEYWORDS)]
        if matches:
            return matches[:5]
    return []


def extract_decideurs_from_page(url, session):
    r = get(url, delay=DELAY_SITE, session=session)
    if not r:
        return [], []

    soup = BeautifulSoup(r.text, 'lxml')
    text = soup.get_text(' ')
    emails = extract_emails(text)
    phones = extract_phones(text)
    decideurs = []

    # 1. Blocs structurés (cards équipe)
    for cls_kw in ['team', 'member', 'person', 'equipe', 'dirigeant', 'staff', 'direction', 'bureau']:
        for tag in soup.find_all(True, class_=re.compile(cls_kw, re.I)):
            tag_text = tag.get_text(' ', strip=True)
            for fk in FONCTIONS_KEYWORDS:
                if fk in tag_text.lower():
                    lines = [l.strip() for l in tag_text.split('\n') if l.strip()]
                    nom_candidate = lines[0][:80] if lines else None
                    if nom_candidate and len(nom_candidate) > 3:
                        em = extract_emails(tag_text)
                        decideurs.append({'nom': nom_candidate, 'fonction': fk.title(),
                                          'email': em[0] if em else None})
                    break

    # 2. Schema.org Person
    for script in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(script.string or '{}')
            if isinstance(data, list):
                data = data[0] if data else {}
            persons = []
            if data.get('@type') == 'Person':
                persons = [data]
            elif 'employee' in data or 'member' in data:
                persons = data.get('employee', data.get('member', []))
            for p in persons:
                if isinstance(p, dict) and p.get('name'):
                    decideurs.append({'nom': p['name'], 'fonction': p.get('jobTitle', ''),
                                      'email': p.get('email', '')})
        except Exception:
            pass

    # 3. Regex contextuel sur tout le texte (fallback)
    if not decideurs:
        for fk in FONCTIONS_KEYWORDS:
            pattern = re.compile(
                rf'([A-ZÀ-ÿ][a-zà-ÿ\-]+(?:\s[A-ZÀ-ÿ][A-Za-zà-ÿ\-]+){{1,3}})\s*'
                rf'[,\n\-–:]*\s*{re.escape(fk)}',
                re.IGNORECASE
            )
            for m in pattern.finditer(text):
                nom_candidate = m.group(1).strip()
                if len(nom_candidate) > 4:
                    decideurs.append({'nom': nom_candidate, 'fonction': fk.title(), 'email': None})

    # Dédoublonnage
    seen, unique = set(), []
    for d in decideurs:
        key = (d['nom'].lower(), d['fonction'].lower())
        if key not in seen:
            seen.add(key)
            unique.append(d)

    return unique[:6], emails[:4]


def scrape_site_deep(nom, site_web, cache):
    if nom in cache:
        return cache[nom]

    session = requests.Session()

    if not check_robots(site_web, '/', session):
        cache[nom] = {'team_url': None, 'decideurs': [], 'emails': []}
        return cache[nom]

    all_decideurs, all_emails, team_url = [], [], None

    # 1. Pages directes probables (sans navigation)
    base = site_web.rstrip('/')
    direct_pages = [base + slug for slug in TEAM_SLUGS]
    # Extension linguistiques
    direct_pages += [base + slug + '/' for slug in TEAM_SLUGS]

    visited = set()
    for page_url in direct_pages:
        if page_url in visited:
            continue
        visited.add(page_url)
        r = get(page_url, delay=0.5, session=session)
        if not r:
            continue
        decs, emails = extract_decideurs_from_page(page_url, session)
        if decs or emails:
            if not team_url:
                team_url = page_url
            all_decideurs.extend(decs)
            all_emails.extend(emails)

    # 2. Liens dans la navigation de la page d'accueil
    if not all_decideurs:
        r_home = get(site_web, session=session)
        if r_home:
            soup_home = BeautifulSoup(r_home.text, 'lxml')
            nav_pages = []
            for a in soup_home.find_all('a', href=True):
                text = a.get_text(strip=True).lower()
                href = a['href']
                if any(kw in text or kw in href.lower() for kw in TEAM_PAGE_KEYWORDS):
                    full = urljoin(site_web, href)
                    if urlparse(full).netloc == urlparse(site_web).netloc and full not in visited:
                        nav_pages.append(full)

            for page_url in nav_pages[:4]:
                visited.add(page_url)
                decs, emails = extract_decideurs_from_page(page_url, session)
                if decs or emails:
                    if not team_url:
                        team_url = page_url
                    all_decideurs.extend(decs)
                    all_emails.extend(emails)

    # 3. Sitemap.xml
    if not all_decideurs:
        sitemap_pages = get_sitemap_team_pages(site_web, session)
        for page_url in sitemap_pages:
            if page_url in visited:
                continue
            visited.add(page_url)
            decs, emails = extract_decideurs_from_page(page_url, session)
            if decs or emails:
                if not team_url:
                    team_url = page_url
                all_decideurs.extend(decs)
                all_emails.extend(emails)

    # Dédoublonnage final
    seen, unique = set(), []
    for d in all_decideurs:
        key = (d['nom'].lower(), d['fonction'].lower())
        if key not in seen:
            seen.add(key)
            unique.append(d)

    result = {
        'team_url':  team_url,
        'decideurs': unique[:6],
        'emails':    list(set(all_emails))[:4],
    }
    cache[nom] = result
    return result

# ── Export Excel ──────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', fgColor='1E3A8A')
HEADER_FONT = Font(color='FFFFFF', bold=True, name='Arial', size=10)
BODY_FONT   = Font(name='Arial', size=10)
ALT_FILL    = PatternFill('solid', fgColor='F0F4FF')


def write_xlsx(rows, path, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    headers = [
        'Nom_Groupement', 'URL_Fiche_LaboData', 'Site_Web_Officiel',
        'Email_General', 'Tel_General',
        'Raison_Sociale', 'SIREN', 'Forme_Juridique', 'Code_APE', 'Statut', 'Adresse_Siege',
        # Dirigeants légaux (Sirene)
        'Dirigeant_Legal_1', 'Fonction_Legale_1',
        'Dirigeant_Legal_2', 'Fonction_Legale_2',
        'Dirigeant_Legal_3', 'Fonction_Legale_3',
        # Dirigeants enrichis (Societe.com — plus détaillés)
        'Decideur_Societe_1', 'Fonction_Societe_1', 'Depuis_1',
        'Decideur_Societe_2', 'Fonction_Societe_2', 'Depuis_2',
        'Decideur_Societe_3', 'Fonction_Societe_3', 'Depuis_3',
        # Décideurs site officiel
        'Decideur_Site_1', 'Fonction_Site_1', 'Email_Decideur_1',
        'Decideur_Site_2', 'Fonction_Site_2', 'Email_Decideur_2',
        # Méta
        'Source_Sirene', 'Source_Societe', 'Source_Site',
        'Score_Fiabilite', 'Date_Extraction',
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 18
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(1, col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A2'

    for i, row in enumerate(rows, 2):
        ws.append(row)
        if i % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(i, col_idx).fill = ALT_FILL
        for col_idx in range(1, len(headers) + 1):
            ws.cell(i, col_idx).font = BODY_FONT

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    # Tri alphabétique
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows.sort(key=lambda r: (r[0] or '').lower())
    for r_idx, row_data in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(r_idx, c_idx).value = val

    wb.save(path)
    log.info(f'Fichier écrit : {path} ({len(rows)} lignes)')


def build_rows(labodata, sirene_cache, societe_cache, sites_cache):
    today = date.today().isoformat()
    rows, manual_rows = [], []

    for slug, info in labodata.items():
        nom = info.get('nom', slug)
        si  = sirene_cache.get(nom)
        soc = societe_cache.get(nom)
        sc  = sites_cache.get(nom)

        # Dirigeants Sirene (légaux, qualités courtes)
        dirs_si  = si.get('dirigeants', []) if si else []
        dleg = [(dirs_si[i]['nom'] if i < len(dirs_si) else None,
                 dirs_si[i]['qualite'] if i < len(dirs_si) else None)
                for i in range(3)]

        # Dirigeants Societe.com (plus détaillés + date nomination)
        dirs_soc = soc.get('dirigeants', []) if soc else []
        dsoc = [(dirs_soc[i]['nom'] if i < len(dirs_soc) else None,
                 dirs_soc[i]['qualite'] if i < len(dirs_soc) else None,
                 dirs_soc[i].get('depuis', '') if i < len(dirs_soc) else None)
                for i in range(3)]

        # Décideurs site officiel
        decs  = sc.get('decideurs', []) if sc else []
        dsite = [(decs[i]['nom'] if i < len(decs) else None,
                  decs[i]['fonction'] if i < len(decs) else None,
                  decs[i].get('email') if i < len(decs) else None)
                 for i in range(2)]

        has_sirene  = bool(si)
        has_societe = bool(soc and soc.get('dirigeants'))
        has_site    = bool(sc and (sc.get('decideurs') or sc.get('emails')))

        # Score fiabilité : 1 = toutes sources, 2 = Sirene+Societe, 3 = une source, 4 = rien
        nb_sources = sum([has_sirene, has_societe, has_site])
        score = {3: 1, 2: 2, 1: 3, 0: 4}[nb_sources]

        row = [
            nom,
            info.get('url_fiche'),
            info.get('site_web'),
            info.get('email_general'),
            info.get('tel_general'),
            si.get('raison_sociale') if si else None,
            si.get('siren') if si else None,
            si.get('nature_juridique') if si else None,
            si.get('code_ape') if si else None,
            si.get('statut') if si else None,
            si.get('adresse_siege') if si else None,
            # Légaux
            dleg[0][0], dleg[0][1],
            dleg[1][0], dleg[1][1],
            dleg[2][0], dleg[2][1],
            # Societe.com
            dsoc[0][0], dsoc[0][1], dsoc[0][2],
            dsoc[1][0], dsoc[1][1], dsoc[1][2],
            dsoc[2][0], dsoc[2][1], dsoc[2][2],
            # Site officiel
            dsite[0][0], dsite[0][1], dsite[0][2],
            dsite[1][0], dsite[1][1], dsite[1][2],
            # Méta
            'oui' if has_sirene else 'non',
            soc.get('url') if soc else None,
            sc.get('team_url') if sc else None,
            score,
            today,
        ]
        rows.append(row)

        # À compléter manuellement : score 3-4 ou aucun décideur identifié
        any_decideur = any(d[0] for d in dleg + dsoc + dsite)
        if score >= 3 or not any_decideur:
            manual_rows.append(row)

    return rows, manual_rows

# ── Main ──────────────────────────────────────────────────────

def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    labodata_cache = load_or_empty(CACHE_LABODATA, 'LaboData')
    sirene_cache   = load_or_empty(CACHE_SIRENE,   'Sirene')
    societe_cache  = load_or_empty(CACHE_SOCIETE,  'Societe.com')
    sites_cache    = load_or_empty(CACHE_SITES,    'Sites')

    # Étape 1 — LaboData
    log.info('=== Étape 1 : LaboData ===')
    labodata = scrape_labodata(LETTERS, labodata_cache)

    # Étape 2 — Sirene
    log.info('=== Étape 2 : Sirene ===')
    for slug, info in labodata.items():
        nom = info.get('nom', '')
        if not nom or nom in sirene_cache:
            continue
        log.info(f'  Sirene : {nom}')
        search_sirene(nom, sirene_cache)
    save_cache(CACHE_SIRENE, sirene_cache)

    # Étape 3 — Societe.com
    log.info('=== Étape 3 : Societe.com ===')
    for slug, info in labodata.items():
        nom = info.get('nom', '')
        if not nom or nom in societe_cache:
            continue
        si = sirene_cache.get(nom)
        if not si:
            societe_cache[nom] = None
            continue
        log.info(f'  Societe.com : {nom} (SIREN {si["siren"]})')
        scrape_societe(nom, si.get('siren'), si.get('raison_sociale'), societe_cache)
    save_cache(CACHE_SOCIETE, societe_cache)

    # Étape 4 — Sites officiels (deep)
    log.info('=== Étape 4 : Sites officiels (deep scraping) ===')
    for slug, info in labodata.items():
        nom  = info.get('nom', '')
        site = info.get('site_web')
        if not site or not nom or nom in sites_cache:
            continue
        log.info(f'  Site : {nom} → {site}')
        scrape_site_deep(nom, site, sites_cache)
    save_cache(CACHE_SITES, sites_cache)

    # Export
    log.info('=== Export Excel ===')
    rows, manual_rows = build_rows(labodata, sirene_cache, societe_cache, sites_cache)
    write_xlsx(rows, OUTPUT_DIR / 'groupements_decideurs.xlsx', 'Décideurs')
    if manual_rows:
        write_xlsx(manual_rows, OUTPUT_DIR / 'a_completer_manuellement.xlsx', 'Manuel')

    # Stats
    n_total     = len(labodata)
    n_sirene    = sum(1 for v in sirene_cache.values() if v)
    n_dirigeants_si  = sum(1 for v in sirene_cache.values() if v and v.get('dirigeants'))
    n_societe   = sum(1 for v in societe_cache.values() if v and v.get('dirigeants'))
    n_site      = sum(1 for v in sites_cache.values() if v and v.get('team_url'))
    n_decideurs = sum(1 for v in sites_cache.values() if v and v.get('decideurs'))

    print('\n══ Statistiques ══════════════════════════════')
    print(f'  {n_total:>4} groupements LaboData')
    print(f'  {n_sirene:>4} trouvés sur Sirene ({n_dirigeants_si} avec dirigeant légal)')
    print(f'  {n_societe:>4} enrichis via Societe.com')
    print(f'  {n_site:>4} avec page équipe scrapée ({n_decideurs} avec décideur site)')
    print(f'  {len(manual_rows):>4} à compléter manuellement')
    print(f'  Fichiers : output/groupements_decideurs.xlsx')
    if manual_rows:
        print(f'             output/a_completer_manuellement.xlsx ({len(manual_rows)} lignes)')
    print()


if __name__ == '__main__':
    main()
