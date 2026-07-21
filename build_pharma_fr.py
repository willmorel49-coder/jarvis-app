# -*- coding: utf-8 -*-
"""Copilote — carte nationale des pharmacies par UGA.

Lit `Base France Décembre 2024.xlsx` (base nationale ~23 000 officines),
EXCLUT la Corse (dép. 20 / 2A / 2B), géocode par commune (BAN gratuit, cache
STATS/geocode_cache.json), éclate les points d'une même commune (jitter
déterministe) et écrit crm/v2/pharma-fr-data.js (window.PHARMA_FR, compact).

100% gratuit (BAN api-adresse.data.gouv.fr). Python 3.9, openpyxl.
"""
import openpyxl
import json
import os
import io
import re
import csv
import math
import time
import unicodedata
import urllib.request

ROOT = os.path.dirname(__file__)
SRC = os.path.join(ROOT, 'Base France Décembre 2024.xlsx')
# Mapping national pharmacie -> groupement (100 groupements, ~17 500 officines),
# hors dossier APP. Sert à remplir le groupement quand la Base France est vide.
GRP_MAP = os.path.join(ROOT, '..', 'GROUPEMENTS', 'data', 'output', 'pharmacies_par_groupement.xlsx')
# Registre national officiel FINESS (etalab). Sert à compléter les officines absentes
# de la Base France : Corse + Outre-mer (que la Base France n'inclut pas).
FINESS = os.path.join(ROOT, '..', 'GROUPEMENTS', 'data', 'finess', 'finess1.csv')
STATS = os.path.join(ROOT, 'STATS')
CACHE = os.path.join(ROOT, 'STATS', 'geocode_cache.json')
ADDR_CACHE = os.path.join(ROOT, 'STATS', 'geocode_addr_cache.json')  # adresse exacte -> [lat,lng]
OUT = os.path.join(ROOT, 'crm', 'v2', 'pharma-fr-data.js')
BAN = 'https://data.geopf.fr/geocodage/search/csv/'  # Géoplateforme IGN (remplace la BAN décommissionnée, mêmes colonnes)


def norm(s):
    return ' '.join(str(s or '').strip().upper().split())


_STOP = ('PHARMACIE', 'PHARMACIES', 'PHARMA', 'PHIE', 'PHIES', 'GRANDE', 'DE', 'DU',
         'DES', 'LA', 'LE', 'LES', 'L', 'D', 'SARL', 'SELARL', 'SELAS', 'SNC', 'EURL')


def name_key(s):
    """Nom normalisé pour matcher officine ↔ mapping groupement (accents/mots vides retirés)."""
    s = ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9 ]', ' ', s.upper())
    return ' '.join(t for t in s.split() if t and t not in _STOP)


def cp5(s):
    d = re.sub(r'[^0-9]', '', str(s or ''))
    return d.zfill(5)[:5] if d else ''


# Établissements qui ne sont PAS des officines et polluent la Base France.
_PH_WORDS = ('PHARMACIE', 'PHARMACIES', 'PHIE', 'PHARMA', 'OFFICINE')
_NON_PH_KW = ('DENTAL', 'DENTAIRE', 'MATERIEL MEDICAL', 'OPTICIEN', 'LABORATOIRE',
              'VETERINAIRE', 'AUDIOPROTH', 'ORTHOPEDIE', 'EHPAD', 'CLINIQUE',
              'HOPITAL', 'MUTUELLE', 'GROSSISTE')
_NON_PH_IDS = {'2002843'}   # BOTICINAL SERVICES : entité services d'un groupement, pas une officine


def is_non_pharmacy(name, _id):
    if str(_id or '').strip() in _NON_PH_IDS:
        return True
    n = ''.join(c for c in unicodedata.normalize('NFD', str(name or '')) if unicodedata.category(c) != 'Mn').upper()
    if any(w in n for w in _PH_WORDS):
        return False                        # a « pharmacie/officine » dans le nom -> vraie officine
    return any(b in n for b in _NON_PH_KW)  # sinon : intrus seulement si mot-clé non-officine


def grp_canon(s):
    """Clé de fusion d'un nom de groupement (casse/accents/ponctuation ignorés)."""
    s = ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^A-Z0-9]', '', s.upper())


def load_groupement_map():
    """Renvoie (lookup, canon) :
    - lookup : (cp, nom normalisé) -> groupement (match unique seulement, clés ambigües ignorées) ;
    - canon  : clé canon -> libellé propre (titres du mapping), pour fusionner
      « APOTHERA » (Base France) et « Apothera » (mapping) en un seul groupement."""
    if not os.path.exists(GRP_MAP):
        print('  [grp] mapping absent :', GRP_MAP)
        return {}, {}
    try:
        wb = openpyxl.load_workbook(GRP_MAP, read_only=True, data_only=True)
    except Exception as e:
        print('  [grp] erreur lecture mapping :', e)
        return {}, {}
    acc = {}
    canon = {}
    for ws in wb.worksheets:
        if ws.title == 'Sommaire':
            continue
        canon.setdefault(grp_canon(ws.title), ws.title)
        it = ws.iter_rows(values_only=True)
        next(it, None)
        for r in it:
            nom = r[0] if len(r) > 0 else ''
            cp = r[3] if len(r) > 3 else ''
            nk, cc = name_key(nom), cp5(cp)
            if nk and cc:
                acc.setdefault((cc, nk), set()).add(ws.title)
    wb.close()
    out = {k: next(iter(v)) for k, v in acc.items() if len(v) == 1}
    print('  [grp] mapping : %d clés uniques (%d ambigües ignorées), %d groupements'
          % (len(out), sum(1 for v in acc.values() if len(v) > 1), len(canon)))
    return out, canon


def load_finess_suppl(bf_keys, grpmap, grpcanon):
    """Officines FINESS (cat 620) absentes de la Base France, restreint à la CORSE
    (France métropolitaine). L'Outre-mer est volontairement exclu. Renvoie des tuples
    pharma (name, tit, ville, cp, uga, grp, seg, tel, mail, comm, ca, id) + communes."""
    if not os.path.exists(FINESS):
        print('  [finess] fichier absent :', FINESS)
        return [], {}
    csv.field_size_limit(10 ** 7)
    out, need = [], {}
    seen = set()
    with open(FINESS, encoding='utf-8', errors='ignore') as f:
        for row in csv.reader(f, delimiter=';'):
            if len(row) < 19 or row[0] != 'structureet' or row[18].strip() != '620':
                continue
            m = re.match(r'\s*(\d{5})\s+(.*)', row[15] or '')
            if not m:
                continue
            cp = m.group(1)
            if cp[:2] not in ('20', '2A', '2B'):
                continue  # France métropolitaine = métropole + Corse seulement (pas l'Outre-mer)
            rs = (row[3] or row[4] or '').strip()
            key = (cp, name_key(rs))
            if not rs or key in bf_keys or key in seen:
                continue  # déjà dans la Base France ou doublon FINESS
            seen.add(key)
            ville = norm(m.group(2))
            if not ville:
                continue
            grp = grpmap.get((cp, name_key(rs)), '')
            if grp:
                grp = grpcanon.get(grp_canon(grp), grp)
            grp = grp or '—'
            tel = str(row[16] or '').strip() if len(row) > 16 else ''
            out.append((rs[:40], '', ville, cp, '', grp, 'Prospect', tel, '', '', 0, str(row[1] or '').strip()))
            need[(ville, cp)] = None
    print('  [finess] complément Corse : %d officines ajoutées' % len(out))
    return out, need


def is_corse(cp):
    cp = str(cp or '').strip()
    return cp[:2] in ('20', '2A', '2B')


def load_cache():
    try:
        return json.load(open(CACHE, encoding='utf-8'))
    except Exception:
        return {}


def wml_commercials():
    """ID officine (CRM) -> commercial + CA. Nos clients réseau, comme la carte secteur."""
    path = os.path.join(ROOT, 'crm', 'v2', 'wml-officines-data.js')
    comm, ca = {}, {}
    try:
        txt = open(path, encoding='utf-8').read()
    except Exception:
        return comm, ca
    for obj in re.findall(r'\{[^{}]*\}', txt):
        mid = re.search(r'"id":"?([\w-]+)"?', obj)
        if not mid:
            continue
        mc = re.search(r'"comms":\[\s*"([^"]+)"', obj)
        mca = re.search(r'"ca":(\d+)', obj)
        if mc:
            comm[mid.group(1)] = mc.group(1)
        if mca:
            ca[mid.group(1)] = int(mca.group(1))
    return comm, ca


def load_addresses():
    """TIRCODE (CIP) -> 'rue, cp ville' depuis les fichiers STATS *_geolocalisation_*.xlsx."""
    import glob
    out = {}
    for path in glob.glob(os.path.join(STATS, '*_geolocalisation_*.xlsx')):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        hdr = [str(c) for c in next(it)]
        ix = {h: i for i, h in enumerate(hdr)}
        def g(r, k):
            i = ix.get(k)
            return str(r[i]).strip() if (i is not None and i < len(r) and r[i] is not None) else ''
        for r in it:
            code = g(r, 'TIRCODE')
            if not code:
                continue
            rue = g(r, 'ADRL3') or g(r, 'ADRL2') or g(r, 'ADRL1')
            cp = g(r, 'ADRCODEPOSTAL').zfill(5) if g(r, 'ADRCODEPOSTAL') else ''
            ville = g(r, 'ADRVILLE')
            if rue and (cp or ville):
                out[code] = (rue + ', ' + cp + ' ' + ville).strip(' ,')
        wb.close()
    return out


def ban_addr_bulk(rows):
    """rows = list of (id, adresse). Renvoie {id:(lat,lng)} via BAN CSV (adresse exacte)."""
    out = {}
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['id', 'adresse'])
    for i, a in rows:
        w.writerow([i, a])
    body = buf.getvalue().encode('utf-8')
    boundary = '----banjarvisA'
    parts = []
    for name, val in [('columns', 'adresse')]:
        parts += ['--' + boundary, 'Content-Disposition: form-data; name="%s"' % name, '', val]
    parts += ['--' + boundary, 'Content-Disposition: form-data; name="data"; filename="a.csv"', 'Content-Type: text/csv', '']
    payload = ('\r\n'.join(parts) + '\r\n').encode('utf-8') + body + ('\r\n--' + boundary + '--\r\n').encode('utf-8')
    req = urllib.request.Request(BAN, data=payload, headers={'Content-Type': 'multipart/form-data; boundary=' + boundary})
    with urllib.request.urlopen(req, timeout=180) as r:
        txt = r.read().decode('utf-8', 'ignore')
    for row in csv.DictReader(io.StringIO(txt)):
        try:
            lat = float(row['latitude']); lng = float(row['longitude']); sc = float(row.get('result_score') or 0)
        except (ValueError, TypeError, KeyError):
            continue
        if sc >= 0.4:
            out[row['id']] = (round(lat, 5), round(lng, 5))
    return out


def exact_coords():
    """{id:(lat,lng)} — adresse exacte géocodée (cache STATS/geocode_addr_cache.json)."""
    addr = load_addresses()
    try:
        cache = json.load(open(ADDR_CACHE, encoding='utf-8'))
    except Exception:
        cache = {}
    todo, coords = [], {}
    for cid, a in addr.items():
        if a in cache and cache[a]:
            coords[cid] = tuple(cache[a])
        else:
            todo.append((cid, a))
    print('Adresses exactes : %d connues | %d en cache | %d à géocoder' % (len(addr), len(coords), len(todo)))
    CH = 800
    for i in range(0, len(todo), CH):
        chunk = todo[i:i + CH]
        try:
            got = ban_addr_bulk(chunk)
        except Exception as e:
            print('  [addr] err', e); break
        for cid, a in chunk:
            if cid in got:
                coords[cid] = got[cid]
                cache[a] = list(got[cid])
        print('  adresses géocodées %d/%d' % (min(i + CH, len(todo)), len(todo)))
        time.sleep(0.4)
    json.dump(cache, open(ADDR_CACHE, 'w', encoding='utf-8'), ensure_ascii=False)
    return coords


def ban_bulk(rows):
    """rows = list of (ville, cp). Renvoie {(ville,cp):(lat,lng)} via BAN CSV."""
    out = {}
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['ville', 'cp'])
    for v, c in rows:
        w.writerow([v, c])
    body = buf.getvalue().encode('utf-8')
    boundary = '----banjarvis'
    parts = []
    for name, val in [('columns', 'ville'), ('postcode', 'cp')]:
        parts.append('--' + boundary)
        parts.append('Content-Disposition: form-data; name="%s"' % name)
        parts.append('')
        parts.append(val)
    parts.append('--' + boundary)
    parts.append('Content-Disposition: form-data; name="data"; filename="a.csv"')
    parts.append('Content-Type: text/csv')
    parts.append('')
    payload = ('\r\n'.join(parts) + '\r\n').encode('utf-8') + body + ('\r\n--' + boundary + '--\r\n').encode('utf-8')
    req = urllib.request.Request(BAN, data=payload, headers={'Content-Type': 'multipart/form-data; boundary=' + boundary})
    with urllib.request.urlopen(req, timeout=120) as r:
        txt = r.read().decode('utf-8', 'ignore')
    rd = csv.DictReader(io.StringIO(txt))
    for row in rd:
        try:
            lat = float(row['latitude']); lng = float(row['longitude'])
            sc = float(row.get('result_score') or 0)
        except (ValueError, TypeError, KeyError):
            continue
        if sc >= 0.3:
            out[(row['ville'], row['cp'])] = (round(lat, 5), round(lng, 5))
    return out


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c) for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    ci = {k: ix.get(k) for k in ('ID', 'Etablissement', 'Titulaire', 'CP', 'Ville', 'UGA',
                                 'SEGMENTATION', 'Groupement', 'Téléphone', 'Email', 'Statut')}

    comm_by_id, ca_by_id = wml_commercials()   # ID officine -> commercial + CA (nos clients réseau)
    EX = exact_coords()                          # ID officine -> (lat,lng) adresse exacte
    GRPMAP, GRPCANON = load_groupement_map()     # (cp,nom)->groupement + fusion des libellés
    SEG_MAP = {'clients a': 'Client A', 'clients b': 'Client B', 'clients c': 'Client C', 'prospects': 'Prospect'}

    nGrpFill = 0        # groupements ajoutés depuis le mapping (Base France vide)
    nVetSkip = 0        # grossistes vétérinaires écartés
    nNonPh = 0          # établissements non-officine écartés (dentaire, labo, EHPAD…)
    pharmas = []          # (name, tit, ville, cp, uga, grp, seg, tel, mail, comm, ca, id)
    need = {}             # (ville,cp) -> None
    bf_keys = set()       # (cp, nom normalisé) présents dans la Base France
    for r in it:
        if str(r[ci['Statut']] or '').strip().lower() == 'supprimée':
            continue
        _id = str(r[ci['ID']] or '').strip()
        if _id.upper().startswith('CAV'):   # grossistes véto (ALCYON, CENTRAVET) : pas des officines
            nVetSkip += 1
            continue
        cp = str(r[ci['CP']] or '').strip()
        if not cp or is_corse(cp):
            continue
        cp = cp.zfill(5)
        ville = norm(r[ci['Ville']])
        if not ville:
            continue
        tit = str(r[ci['Titulaire']] or '').strip()
        name = str(r[ci['Etablissement']] or tit or '').strip()
        if is_non_pharmacy(name, _id):      # dentaire, matériel médical, labo, EHPAD… : pas une officine
            nNonPh += 1
            continue
        uga = str(r[ci['UGA']] or '').strip()
        grp = str(r[ci['Groupement']] or '').strip()
        if not grp or grp == '—':           # Base France sans groupement -> mapping national
            g2 = GRPMAP.get((cp, name_key(name or tit)))
            if g2:
                grp = g2
                nGrpFill += 1
        if grp and grp != '—':              # fusion des libellés (APOTHERA == Apothera)
            grp = GRPCANON.get(grp_canon(grp), grp)
        grp = grp or '—'
        seg = SEG_MAP.get(str(r[ci['SEGMENTATION']] or '').strip().lower(), 'Non défini')
        tel = str(r[ci['Téléphone']] or '').strip()
        mail = str(r[ci['Email']] or '').strip()
        comm = comm_by_id.get(_id, '')
        ca = ca_by_id.get(_id, 0)
        pharmas.append((name, tit, ville, cp, uga, grp, seg, tel, mail, comm, ca, _id))
        need[(ville, cp)] = None
        bf_keys.add((cp, name_key(name or tit)))
    wb.close()
    print('Pharmacies Base France (métropole) :', len(pharmas), '| communes uniques :', len(need))

    # complément métropolitain : Corse depuis FINESS (absente de la Base France)
    suppl, suppl_need = load_finess_suppl(bf_keys, GRPMAP, GRPCANON)
    pharmas.extend(suppl)
    for k in suppl_need:
        need.setdefault(k, None)
    print('Total officines :', len(pharmas), '| communes uniques :', len(need))

    # géocodage : cache + BAN pour les manquantes
    cache = load_cache()
    coords = {}
    todo = []
    for (ville, cp) in need:
        key = cp + ' ' + ville
        if key in cache and cache[key]:
            coords[(ville, cp)] = tuple(cache[key])
        else:
            todo.append((ville, cp))
    print('En cache :', len(coords), '| à géocoder :', len(todo))
    CH = 1200
    for i in range(0, len(todo), CH):
        chunk = todo[i:i + CH]
        got = ban_bulk(chunk)
        for k, v in got.items():
            coords[k] = v
            cache[k[1] + ' ' + k[0]] = list(v)
        print('  géocodé %d/%d (+%d)' % (min(i + CH, len(todo)), len(todo), len(got)))
        time.sleep(0.4)
    json.dump(cache, open(CACHE, 'w', encoding='utf-8'), ensure_ascii=False)

    # éclatement des points d'une même commune (spirale déterministe)
    seen = {}
    ugas, grps, segs, comms = {}, {}, {}, {'': 0}   # comm index 0 = pas notre client
    P = []
    dropped = 0
    nExact = 0
    nBadGeo = 0
    for (name, tit, ville, cp, uga, grp, seg, tel, mail, comm, ca, _id) in pharmas:
        ex = EX.get(_id)
        if ex:
            lat, lng = ex; nExact += 1   # adresse exacte : position réelle, pas de jitter
        else:
            c = coords.get((ville, cp))
            if not c:
                dropped += 1
                continue
            k = (ville, cp)
            idx = seen.get(k, 0); seen[k] = idx + 1
            lat, lng = c
            if idx:
                ring = int((math.sqrt(idx))) + 1
                ang = idx * 2.399963  # angle d'or
                rad = 0.0016 * ring
                lat = round(lat + rad * math.cos(ang), 5)
                lng = round(lng + rad * math.sin(ang) / max(0.3, math.cos(math.radians(lat))), 5)
        # garde-fou : un CP métropolitain doit tomber dans la France métro (bbox).
        # Sinon = géocodage foireux (ex. St-Barthélemy-d'Anjou 49 envoyé aux Caraïbes) -> on écarte.
        if cp[:2] not in ('97', '98') and not (41.0 <= lat <= 51.6 and -5.5 <= lng <= 9.8):
            nBadGeo += 1
            continue
        ui = ugas.setdefault(uga, len(ugas))
        gi = grps.setdefault(grp, len(grps))
        si = segs.setdefault(seg, len(segs))
        ki = comms.setdefault(comm, len(comms))
        P.append([lat, lng, ui, gi, si, ki, name[:40], ville[:22], cp, tel[:18], tit[:34], mail[:44], ca or 0, _id])

    inv = lambda d: [k for k, _ in sorted(d.items(), key=lambda x: x[1])]
    nClients = sum(1 for p in P if segs and inv(segs)[p[4]].startswith('Client'))
    data = {
        'meta': {'n': len(P), 'communes': len(need), 'clients': nClients,
                 'source': 'Base France 12/2024 + FINESS (Corse) · France métropolitaine'},
        'uga': inv(ugas), 'grp': inv(grps), 'seg': inv(segs), 'comm': inv(comms), 'p': P,
    }
    with open(OUT, 'w', encoding='utf-8') as fh:
        fh.write('// Copilote — carte nationale pharmacies (hors Corse) : UGA, groupement,\n')
        fh.write('// segmentation client/prospect, commercial réseau. build_pharma_fr.py.\n')
        fh.write('// Chaque point: [lat,lng,ugaIdx,grpIdx,segIdx,commIdx,nom,ville,cp,tel,titulaire,email,ca,id]\n')
        fh.write('window.PHARMA_FR=' + json.dumps(data, ensure_ascii=False, separators=(',', ':')) + ';\n')
    nGrpCovered = sum(1 for p in P if inv(grps)[p[3]] != '—')
    print('OK ->', OUT, '(%.1f Mo, %d pts, %d clients, %d adresses exactes, %d UGA, %d comm, %d dropped)'
          % (os.path.getsize(OUT) / 1048576.0, len(P), nClients, nExact, len(ugas), len(comms) - 1, dropped))
    print('   groupements : %d officines rattachées (%.0f%%), dont %d ajoutées via le mapping national'
          % (nGrpCovered, 100.0 * nGrpCovered / max(1, len(P)), nGrpFill))
    print('   écartés : %d grossistes véto (CAV) · %d non-officines (dentaire/labo/EHPAD…) · %d géocodages hors métropole' % (nVetSkip, nNonPh, nBadGeo))


if __name__ == '__main__':
    main()
