#!/usr/bin/env python3
"""
Génère le fichier Excel final groupements_decideurs_V4.xlsx
en fusionnant: Sirene, Societe.com, Sites, Agents web, Panorama QdP
Nouvelles colonnes V4: tel_dirigeant, email_dirigeant, nb_adherents, nb_labos, cotisation, alliance
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = Path('/Users/williammorel/JARVIS/APP')
OUTPUT_DIR = BASE / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Chargement des caches ──────────────────────────────────────

def load_json(path):
    p = Path(path)
    if p.exists():
        with open(p, encoding='utf-8') as f:
            return json.load(f)
    return {}

labo     = load_json(BASE / 'cache_labodata.json')
sirene   = load_json(BASE / 'cache_sirene.json')
societe  = load_json(BASE / 'cache_societe.json')
sites    = load_json(BASE / 'cache_sites.json')
agents   = load_json(BASE / 'cache_agents.json')
panorama = load_json(BASE / 'cache_panorama.json')

# Mapping panorama_key → labodata_nom (pour les clés qui diffèrent)
PANORAMA_TO_LABO = {
    "Cap'Unipharm":          "Cap’Unipharm",   # apostrophe courbe
    "Pharm-free":            "Pharm&Free",
    "Pharm-upp":             "Pharm-Upp",
    "Pharmacorp":            "PharmaCorp",
    "Pharmavance":           "PharmAvance",
    "Univers Pharmacie":     "Groupe Univers Pharmacie",
}
# Inverse pour lookup panorama depuis labodata nom
LABO_TO_PANORAMA = {v: k for k, v in PANORAMA_TO_LABO.items()}

def get_panorama(nom):
    pan_key = LABO_TO_PANORAMA.get(nom, nom)
    return panorama.get(pan_key) or {}

# Groupements dans panorama/agents mais absents de labodata
NOUVEAUX = [
    # Depuis images panorama QdP
    "D Docteurs en Pharmacie et PHD",
    "Officinal by Boticinal",
    "Pharm'Auvergne",
    "Pharmidee",
    "UniQ",
    "Unipharm Grand Ouest",
    "Unipharm Normandie IDF",
    # Depuis APSAGIR / Federgy / EVECIAL (recherche APE + fédérations)
    "Les Officinales",
    "DéPhie",
    "Le Club du Comptoir",
    "Trentactiv",
    "Orphie",
    "Tag Pharma",
    "Pharma Star",
    "Les Pharmaciens de Méditerranée",
]

# ── Helpers ────────────────────────────────────────────────────

def is_real_dir(d):
    n = d.get('nom', '') or d.get('name', '') or ''
    bad = ['obtenir', 'juridique', 'activité', 'code', 'ape', 'siège',
           'siren', 'effectif', 'chiffre', 'forme', 'rapport']
    if len(n) < 3 or len(n) > 80:
        return False
    return not any(b in n.lower() for b in bad)

def fmt_dir(d):
    nom = d.get('nom', '') or d.get('name', '') or ''
    fn  = d.get('qualite', '') or d.get('fonction', '') or d.get('jobTitle', '') or ''
    return f"{nom} ({fn})" if fn else nom

def first_dir(dirs):
    real = [d for d in dirs if is_real_dir(d)]
    return real[0] if real else None

def all_real(dirs):
    return [d for d in dirs if is_real_dir(d)]

def pan_dirigeants(pan):
    """Convertit les champs président/dg/fondateur du panorama en liste de dicts dirigeants."""
    result = []
    if pan.get('president'):
        fn = pan.get('president_titre', 'Président(e)')
        result.append({'nom': pan['president'], 'fonction': fn})
    if pan.get('fondateur') and pan.get('fondateur') != pan.get('president'):
        result.append({'nom': pan['fondateur'], 'fonction': 'Fondateur(trice)'})
    if pan.get('dg') and pan.get('dg') != pan.get('president'):
        fn = pan.get('dg_titre', 'Directeur(trice) général(e)')
        result.append({'nom': pan['dg'], 'fonction': fn})
    if pan.get('dirigeante') and pan.get('dirigeante') not in [pan.get('president'), pan.get('dg')]:
        result.append({'nom': pan['dirigeante'], 'fonction': pan.get('president_titre', 'Dirigeant(e)')})
    if pan.get('direction_commerciale'):
        result.append({'nom': pan['direction_commerciale'], 'fonction': 'Direction commerciale'})
    if pan.get('direction_marketing'):
        result.append({'nom': pan['direction_marketing'], 'fonction': 'Direction marketing'})
    if pan.get('directeur_reseau'):
        result.append({'nom': pan['directeur_reseau'], 'fonction': 'Directeur réseau'})
    if pan.get('codirigants'):
        for nom in pan['codirigants']:
            result.append({'nom': nom, 'fonction': 'Co-dirigeant'})
    if pan.get('dg2'):
        result.append({'nom': pan['dg2'], 'fonction': 'DG'})
    if pan.get('dirigeant'):
        result.append({'nom': pan['dirigeant'], 'fonction': 'Dirigeant'})
    return result

# ── Build row helper ───────────────────────────────────────────

def build_row(nom, site='', email='', tel=''):
    pan = get_panorama(nom)

    # Sirene
    s = sirene.get(nom) or {}
    if not isinstance(s, dict): s = {}
    siren_num  = s.get('siren', '')
    raison_soc = s.get('raison_sociale', '')
    libelle_ape = s.get('libelle_ape', '') or s.get('code_ape', '')
    statut     = s.get('statut', '')
    adresse    = s.get('adresse_siege', '')
    dirs_sirene = all_real(s.get('dirigeants', []))

    # Societe.com
    soc = societe.get(nom) or {}
    if not isinstance(soc, dict): soc = {}
    dirs_soc = all_real(soc.get('dirigeants', []))
    url_soc = soc.get('url', '')

    # Sites
    si = sites.get(nom) or {}
    if not isinstance(si, dict): si = {}
    dirs_site = all_real(si.get('decideurs', []))

    # Agents
    ag = agents.get(nom) or {}
    if not isinstance(ag, dict): ag = {}
    dirs_ag_raw = ag.get('decideurs', [])
    dirs_ag = [d for d in dirs_ag_raw if is_real_dir(d)]

    # Panorama dirigeants (priorité haute — données vérifiées)
    dirs_pan = pan_dirigeants(pan)

    # Fusion - priorité: Panorama > Sirene > Agents > Societe > Sites
    all_dirs = dirs_pan + dirs_sirene + dirs_ag + dirs_soc + dirs_site

    seen = set()
    merged = []
    for d in all_dirs:
        n = (d.get('nom', '') or '').strip().upper()
        if n and n not in seen and is_real_dir(d):
            seen.add(n)
            merged.append(d)

    p1 = merged[0] if merged else None
    nom_p1 = (p1.get('nom', '') if p1 else '') or ''
    fn_p1  = (p1.get('qualite', '') or p1.get('fonction', '') if p1 else '') or ''

    p2 = merged[1] if len(merged) > 1 else None
    nom_p2 = (p2.get('nom', '') if p2 else '') or ''
    fn_p2  = (p2.get('qualite', '') or p2.get('fonction', '') if p2 else '') or ''

    all_str = ' | '.join(fmt_dir(d) for d in merged[:6]) if merged else ''

    score = 1
    if dirs_pan:    score = max(score, 5)
    if dirs_sirene: score = max(score, 4)
    elif dirs_ag:   score = max(score, 3)
    elif dirs_soc:  score = max(score, 2)
    elif dirs_site: score = max(score, 2)

    # Enrichissements panorama
    tel_dir    = pan.get('tel_dirigeant', '')
    email_dir  = pan.get('email_dirigeant', '')
    nb_adh     = pan.get('nb_adherents', '')
    nb_labos   = pan.get('nb_labos', '')
    cotisation = pan.get('cotisation', '')
    alliance   = pan.get('alliance', '')

    return {
        'nom': nom,
        'site_web': site,
        'email': email,
        'tel': tel,
        'siren': siren_num,
        'raison_sociale': raison_soc,
        'statut': statut,
        'adresse': adresse,
        'libelle_ape': libelle_ape,
        'nom_p1': nom_p1,
        'fn_p1': fn_p1,
        'nom_p2': nom_p2,
        'fn_p2': fn_p2,
        'all_dirigeants': all_str,
        'url_societe': url_soc,
        'score': score,
        'nb_dirs': len(merged),
        # Nouvelles colonnes V4
        'tel_dirigeant': tel_dir or '',
        'email_dirigeant': email_dir or '',
        'nb_adherents': nb_adh if nb_adh else '',
        'nb_labos': nb_labos if nb_labos else '',
        'cotisation': cotisation or '',
        'alliance': alliance or '',
    }

# ── Build rows depuis labodata ─────────────────────────────────

rows = []
for path, data in sorted(labo.items(), key=lambda x: x[1].get('nom', '') if isinstance(x[1], dict) else ''):
    if not isinstance(data, dict):
        continue
    nom  = data.get('nom', '')
    site = data.get('site_web', '') or ''
    email = data.get('email_general', '') or ''
    tel   = data.get('tel_general', '') or ''
    rows.append(build_row(nom, site, email, tel))

# ── Ajout des groupements nouveaux (panorama uniquement) ────────

for pk in NOUVEAUX:
    pan = panorama.get(pk) or {}
    # Essayer de trouver site/email/tel dans panorama
    site_web   = ''
    email_gen  = pan.get('email_dirigeant', '') or ''
    tel_gen    = pan.get('tel_dirigeant', '') or ''
    # Nom réel
    nom_reel = pan.get('nom_groupement', pk)
    rows.append(build_row(pk, site_web, '', ''))

# Tri alphabétique
rows.sort(key=lambda r: r['nom'].lower())

# ── Colonnes Excel ─────────────────────────────────────────────

COLS = [
    ('Groupement',            32),
    ('Site web',              26),
    ('Email groupement',      28),
    ('Tél groupement',        14),
    ('Alliance',              18),
    ('Nb adhérents',          10),
    ('Nb labos',              9),
    ('Cotisation',            22),
    ('Nom Décideur 1',        26),
    ('Fonction 1',            24),
    ('Nom Décideur 2',        26),
    ('Fonction 2',            24),
    ('Tél dirigeant',         16),
    ('Email dirigeant',       28),
    ('Tous dirigeants',       55),
    ('SIREN',                 12),
    ('Raison sociale',        28),
    ('Statut',                10),
    ('Adresse siège',         28),
    ('APE/Activité',          20),
    ('Score',                  7),
    ('Nb dirigeants',          7),
    ('Societe.com',            26),
]

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
ALT_FILL    = PatternFill('solid', fgColor='E8EEF7')
LINK_FONT   = Font(color='0563C1', underline='single', name='Calibri', size=10)
BODY_FONT   = Font(name='Calibri', size=10)
GREEN_FONT  = Font(color='1A7C3C', name='Calibri', size=10, bold=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Groupements Décideurs V4'
ws.freeze_panes = 'A2'

# Header
for ci, (label, width) in enumerate(COLS, 1):
    c = ws.cell(row=1, column=ci, value=label)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(ci)].width = width

ws.row_dimensions[1].height = 28

# Lignes
for ri, r in enumerate(rows, 2):
    is_alt = ri % 2 == 0
    fill = ALT_FILL if is_alt else None

    vals = [
        r['nom'],
        r['site_web'],
        r['email'],
        r['tel'],
        r['alliance'],
        r['nb_adherents'],
        r['nb_labos'],
        r['cotisation'],
        r['nom_p1'],
        r['fn_p1'],
        r['nom_p2'],
        r['fn_p2'],
        r['tel_dirigeant'],
        r['email_dirigeant'],
        r['all_dirigeants'],
        r['siren'],
        r['raison_sociale'],
        r['statut'],
        r['adresse'],
        r['libelle_ape'],
        r['score'],
        r['nb_dirs'],
        r['url_societe'],
    ]

    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=val if val != '' else None)
        c.font = BODY_FONT
        c.alignment = Alignment(vertical='center', wrap_text=False)
        if fill:
            c.fill = fill

    # Couleur alliance
    if r['alliance']:
        ac = ws.cell(row=ri, column=5)
        ac.font = Font(color='0057FF', name='Calibri', size=10, bold=True)

    # Hyperlinks
    if r['site_web']:
        c = ws.cell(row=ri, column=2)
        c.hyperlink = r['site_web']
        c.font = LINK_FONT
    if r['url_societe']:
        c = ws.cell(row=ri, column=23)
        c.hyperlink = r['url_societe']
        c.font = LINK_FONT
    if r['email_dirigeant']:
        c = ws.cell(row=ri, column=14)
        mailto = r['email_dirigeant']
        if '@' in mailto:
            c.hyperlink = f'mailto:{mailto}'
        c.font = GREEN_FONT

    ws.row_dimensions[ri].height = 16

# ── Feuille 2 : Sans décideurs ─────────────────────────────────

ws2 = wb.create_sheet('À compléter manuellement')
ws2.freeze_panes = 'A2'
hdrs2 = [('Groupement', 32), ('Site web', 26), ('Email', 28), ('SIREN', 12), ('Note', 40)]
for ci, (label, width) in enumerate(hdrs2, 1):
    c = ws2.cell(row=1, column=ci, value=label)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.column_dimensions[get_column_letter(ci)].width = width

r2i = 2
for r in rows:
    if r['nb_dirs'] == 0:
        fill = ALT_FILL if r2i % 2 == 0 else None
        ws2.cell(row=r2i, column=1, value=r['nom']).font = BODY_FONT
        c2 = ws2.cell(row=r2i, column=2, value=r['site_web'])
        c2.font = LINK_FONT if r['site_web'] else BODY_FONT
        if r['site_web']:
            c2.hyperlink = r['site_web']
        ws2.cell(row=r2i, column=3, value=r['email']).font = BODY_FONT
        ws2.cell(row=r2i, column=4, value=r['siren']).font = BODY_FONT
        ws2.cell(row=r2i, column=5, value='Aucun décideur trouvé').font = BODY_FONT
        for ci2 in range(1, 6):
            ws2.cell(row=r2i, column=ci2).alignment = Alignment(vertical='center')
            if fill:
                ws2.cell(row=r2i, column=ci2).fill = fill
        r2i += 1

# ── Stats et sauvegarde ────────────────────────────────────────

out_path = OUTPUT_DIR / 'groupements_decideurs_V4.xlsx'
wb.save(out_path)

total    = len(rows)
avec_dir = sum(1 for r in rows if r['nb_dirs'] > 0)
avec_tel = sum(1 for r in rows if r['tel_dirigeant'])
avec_pan = sum(1 for r in rows if r['nb_adherents'] != '')
avec_all = sum(1 for r in rows if r['alliance'])
sans_dir = total - avec_dir

print(f'Excel V4 généré : {out_path}')
print(f'{total} groupements | {avec_dir} avec décideurs | {sans_dir} sans décideurs')
print(f'{avec_tel} avec tél dirigeant | {avec_pan} avec nb adhérents | {avec_all} avec alliance')
