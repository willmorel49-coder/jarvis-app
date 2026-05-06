#!/usr/bin/env python3
"""
Benchmark data V2 — Intégral Pharma CRM
Sources:
  - TOP rotations France 2026.xlsx    → ip_qty, ip_ca, ranking, categorie
  - TOP IP DÉCROISSANT.xlsx           → CIP13 direct, prix IP, is_froid
  - WML_01/02/03_2026.xlsx            → name→CIP13 bridge (fallback)
  - 2025-01-a-06 / 2025-07-a-12 / 2026-01 Ameli CIP files → 13 months data
Output: crm/benchmark-data.js
"""
import warnings; warnings.filterwarnings('ignore')
import openpyxl, re, os, sys
from datetime import datetime
from pathlib import Path

BASE = str(Path(__file__).parent)
OUT  = os.path.join(BASE, 'crm', 'benchmark-data.js')
N_PHARMA = 19000

def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').strip().upper())

def safe_float(v, default=0.0):
    try: return float(str(v).replace(',','.').replace(' ','')) if v is not None else default
    except: return default

def safe_int(v, default=0):
    try: return int(float(str(v).replace(',','.').replace(' ',''))) if v is not None else default
    except: return default

# ── 1. TOP ROTATIONS → base produit ──────────────────────────────────────────
print("Loading TOP rotations France 2026.xlsx...")
sheet_cats = {'TOP Froid':'froid','TOP Med010':'nr','TOP inf 4,8':'pp','TOP median':'mi','TOP sup 480':'ch'}
records = {}  # norm_designation → record dict

_top_file = os.path.join(BASE, 'TOP rotations France 2026.xlsx')
if not os.path.exists(_top_file):
    print(f'ERREUR : fichier introuvable → {_top_file}')
    sys.exit(1)
wb = openpyxl.load_workbook(_top_file, read_only=True, data_only=True)
for sname, cat in sheet_cats.items():
    if sname not in wb.sheetnames:
        print(f"  WARNING: sheet {sname} not found"); continue
    ws = wb[sname]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        desig = str(row[0]).strip()
        if not desig or desig == 'None': continue
        ip_ca  = safe_float(row[1] if len(row)>1 else None)
        ip_qty = safe_int(row[2] if len(row)>2 else None)
        ip_rank_qty = safe_int(row[3] if len(row)>3 else None)
        ip_rank_ca  = safe_int(row[4] if len(row)>4 else None)
        key = norm(desig)
        records[key] = {
            'designation': desig, 'categorie': cat,
            'ip_qty': ip_qty, 'ip_ca': ip_ca,
            'ip_rank_qty': ip_rank_qty, 'ip_rank_ca': ip_rank_ca,
            'cip13': '', 'prix_ht': 0.0, 'prix_ip': 0.0,
            'remise_pct': 0.0, 'offre_ip': 0.0, 'is_froid': cat=='froid',
        }
        count += 1
    print(f"  {sname}: {count} produits")
wb.close()
print(f"  Total base produits: {len(records)}")

# ── 2. TOP IP DÉCROISSANT → CIP13 + prix ─────────────────────────────────────
print("\nLoading TOP IP DÉCROISSANT.xlsx...")
cat_map = {'Petit prix':'pp','Intermédiaire':'mi','NR':'nr','Non remboursé':'nr'}
sheet_cat_ip = {'TOP PETITS PRIX':'pp','TOP INTERMEDIAIRE':'mi','TOP NR':'nr','BIOSIMILAIRES':'biosim'}
cip_to_prix = {}  # cip13 → prix info

_ip_file = os.path.join(BASE, 'TOP IP DÉCROISSANT.xlsx')
matched_from_ip = 0
if not os.path.exists(_ip_file):
    print(f'AVERTISSEMENT : fichier introuvable → {_ip_file} (les prix IP seront absents)')
    wb2 = None
else:
    wb2 = openpyxl.load_workbook(_ip_file, read_only=True, data_only=True)
for sname, cat_default in (sheet_cat_ip.items() if wb2 is not None else {}.items()):
    if sname not in wb2.sheetnames:
        print(f"  WARNING: sheet {sname} not found"); continue
    ws2 = wb2[sname]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        cip_raw = row[0]
        try:
            cip = str(int(float(str(cip_raw).replace(' ',''))))
        except: continue
        if len(cip) < 10: continue
        desig     = str(row[1]).strip() if len(row)>1 and row[1] else ''
        prix_ht   = safe_float(row[4] if len(row)>4 else None)
        prix_ip   = safe_float(row[5] if len(row)>5 else None)
        remise    = safe_float(row[6] if len(row)>6 else None)
        offre_ip  = safe_float(row[8] if len(row)>8 else None)
        froid_val = row[9] if len(row)>9 else None
        is_froid  = froid_val is not None and str(froid_val).strip() not in ('', 'None')

        cip_to_prix[cip] = {'prix_ht':prix_ht,'prix_ip':prix_ip,'remise_pct':remise,'offre_ip':offre_ip,'is_froid':is_froid,'cat':cat_default}

        # Match dans records par designation
        key = norm(desig)
        if key in records:
            records[key]['cip13'] = cip
            records[key]['prix_ht'] = prix_ht
            records[key]['prix_ip'] = prix_ip
            records[key]['remise_pct'] = remise
            records[key]['offre_ip'] = offre_ip
            records[key]['is_froid'] = is_froid or records[key]['is_froid']
            matched_from_ip += 1

if wb2 is not None:
    wb2.close()
print(f"  Produits matchés par CIP direct: {matched_from_ip}")

# ── 3. WML bridge name→CIP13 + prix réels IP ─────────────────────────────────
print("\nBuilding WML name→CIP13 bridge + prix IP réels...")
name_to_cip = {}
cip_wml_prix = {}   # cip13 → liste de puNet observés → pour faire la médiane

for wf in ['WML_01_2026.xlsx','WML_02_2026.xlsx','WML_03_2026.xlsx']:
    path = os.path.join(BASE, wf)
    if not os.path.exists(path): continue
    try:
        wb3 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws3 = wb3.active
        headers = [c.value for c in next(ws3.iter_rows(max_row=1))]
        try:
            di   = headers.index('PLVDESIGNATION')
            ci   = headers.index('ARTCODEBARRE')
            pni  = headers.index('PLVPUNET')
            pbi  = headers.index('PLVPUBRUT')
        except ValueError: wb3.close(); continue
        for row in ws3.iter_rows(min_row=2, values_only=True):
            desig   = row[di]
            cip_raw = row[ci]
            punet   = row[pni]
            pubrut  = row[pbi]
            if desig is None or cip_raw is None: continue
            try:
                cip = str(int(float(str(cip_raw).replace(',','.'))))
                if len(cip) < 10: continue
                name_to_cip[norm(desig)] = cip
                # Collecter les prix nets réels (positifs uniquement = ventes, pas retours)
                pn = safe_float(punet)
                pb = safe_float(pubrut)
                if pn > 0 and pb > 0:
                    if cip not in cip_wml_prix:
                        cip_wml_prix[cip] = {'nets': [], 'bruts': []}
                    cip_wml_prix[cip]['nets'].append(pn)
                    cip_wml_prix[cip]['bruts'].append(pb)
            except: pass
        wb3.close()
    except Exception as e:
        print(f"  ERROR {wf}: {e}")

# Calculer prix médian par CIP13 depuis WML
def mediane(lst):
    s = sorted(lst)
    n = len(s)
    return s[n//2] if n % 2 else (s[n//2-1]+s[n//2])/2

cip_wml_prix_med = {}
for cip, v in cip_wml_prix.items():
    if v['nets']:
        pn_med = mediane(v['nets'])
        pb_med = mediane(v['bruts'])
        remise = round((pb_med - pn_med) / pb_med * 100, 2) if pb_med > 0 else 0.0
        cip_wml_prix_med[cip] = {
            'prix_ht': round(pb_med, 4),
            'prix_ip': round(pn_med, 4),
            'remise_pct': remise,
            'offre_ip': 0.0,
        }

print(f"  Prix WML disponibles pour {len(cip_wml_prix_med)} CIP13")

# Apply WML bridge to unmatched records
wml_matched = 0
wml_prix_added = 0
for key, rec in records.items():
    # Étape A : bridge CIP13 si manquant
    if not rec['cip13'] and key in name_to_cip:
        rec['cip13'] = name_to_cip[key]
        wml_matched += 1
    # Étape B : enrichir prix depuis TOP IP DÉCROISSANT si CIP connu
    if rec['cip13'] and rec['prix_ip'] == 0 and rec['cip13'] in cip_to_prix:
        p = cip_to_prix[rec['cip13']]
        rec.update({k:p[k] for k in ('prix_ht','prix_ip','remise_pct','offre_ip','is_froid')})
        wml_prix_added += 1
    # Étape C : enrichir prix depuis WML si toujours manquant
    if rec['cip13'] and rec['prix_ip'] == 0 and rec['cip13'] in cip_wml_prix_med:
        p = cip_wml_prix_med[rec['cip13']]
        rec.update({k:p[k] for k in ('prix_ht','prix_ip','remise_pct','offre_ip')})
        wml_prix_added += 1

print(f"  WML bridge CIP13 nouveaux: {wml_matched}")
print(f"  Prix enrichis (TOP IP + WML): {wml_prix_added}")
total_with_cip  = sum(1 for r in records.values() if r['cip13'])
total_with_prix = sum(1 for r in records.values() if r['prix_ip'] > 0)
print(f"  Total avec CIP13: {total_with_cip} / {len(records)}")
print(f"  Total avec prix_ip: {total_with_prix} / {len(records)}")

# ── 3.5. STOCK → artnature par CIP13 (source de vérité catégorie) ────────────
print("\nLoading stock file for artnature enrichment...")
import glob as _glob
stock_files = sorted(_glob.glob(os.path.join(BASE, 'stock*.xlsx')))
cip_to_nature = {}  # cip13 str → artnature normalized
if not stock_files:
    print("  AVERTISSEMENT : aucun fichier stock*.xlsx trouvé — artnature sera vide")
else:
    stock_path = stock_files[-1]  # le plus récent alphabétiquement
    print(f"  Stock utilisé : {os.path.basename(stock_path)}")
    try:
        wb_s = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
        ws_s = wb_s.active
        rows_s = list(ws_s.iter_rows(values_only=True))
        hdr_s = [str(c).lower().strip() if c else '' for c in rows_s[0]]
        col_s = {n: i for i, n in enumerate(hdr_s) if n}
        i_cip = col_s.get('artcodebarre')
        i_nat = col_s.get('artnature')
        i_afm = col_s.get('afmcode')
        if i_cip is not None and i_nat is not None:
            for row in rows_s[1:]:
                cip_raw = row[i_cip]
                nat_raw = row[i_nat]
                if not cip_raw or not nat_raw: continue
                nat_str = str(nat_raw).strip()
                if nat_str in ('#N/A','None','nan',''): continue
                try:
                    cip = str(int(float(str(cip_raw).replace(' ',''))))
                except: continue
                if len(cip) < 10: continue
                # normalize: 'Generique Partenaire' → 'generique_partenaire'
                cip_to_nature[cip] = nat_str.lower().replace(' ','_')
        wb_s.close()
        print(f"  CIP avec artnature : {len(cip_to_nature)}")
        from collections import Counter as _Ctr
        print(f"  Distribution : {dict(_Ctr(cip_to_nature.values()).most_common())}")
    except Exception as e:
        print(f"  ERREUR lecture stock : {e}")

# Injecter artnature dans records
nature_matched = 0
for rec in records.values():
    cip = rec.get('cip13','')
    nat = cip_to_nature.get(cip, '')
    rec['artnature'] = nat
    if nat: nature_matched += 1
print(f"  Produits enrichis artnature : {nature_matched} / {len(records)}")

# ── 4. AMELI MONTHLY DATA (13 months: Jan25→Jan26) ───────────────────────────
print("\nLoading Ameli monthly data (13 months)...")
ameli = {}  # cip13 → {'months':[13 ints], 'atc2': str}
MONTHS_LABEL = ['2025-01','2025-02','2025-03','2025-04','2025-05','2025-06',
                '2025-07','2025-08','2025-09','2025-10','2025-11','2025-12','2026-01']

ameli_files = [
    ('2025-01-a-06_medic-am-par-classe-atc_serie-mensuelle-labellisee.xlsx', '2025_cip_total', list(range(6))),
    ('2025-07-a-12_medic-am-par-classe-atc_serie-mensuelle.xlsx',            '2025_cip_total', list(range(6,12))),
    ('2026-01_medic-am-par-classe-atc_serie-mensuelle.xlsx',                 '2026_cip_total', [12]),
]

for fname, shname, month_indices in ameli_files:
    path = os.path.join(BASE, fname)
    if not os.path.exists(path): print(f"  MISSING: {fname}"); continue
    wb4 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if shname not in wb4.sheetnames: wb4.close(); continue
    ws4 = wb4[shname]

    # Find header row
    header = None
    for i, row in enumerate(ws4.iter_rows(max_row=10, values_only=True), 1):
        if row and row[0] == 'CIP13': header = list(row); break
    if header is None: wb4.close(); continue

    # Find boites columns (pattern: "Nombre de boites remboursées\nYYYY-MM")
    boites_col_map = {}  # month_label → col_idx
    for ci, col in enumerate(header):
        if col and 'Nombre de boites' in str(col):
            for ml in MONTHS_LABEL:
                if ml in str(col):
                    boites_col_map[ml] = ci

    count = 0
    for row in ws4.iter_rows(min_row=7, values_only=True):
        if not row or row[0] is None: continue
        try:
            cip = str(int(float(str(row[0]).replace(' ',''))))
        except: continue
        if len(cip) < 10: continue

        atc2 = str(row[3]).strip() if len(row)>3 and row[3] else ''

        if cip not in ameli:
            ameli[cip] = {'months': [0]*13, 'atc2': atc2}
        elif not ameli[cip]['atc2'] and atc2:
            ameli[cip]['atc2'] = atc2

        for mi in month_indices:
            ml = MONTHS_LABEL[mi]
            if ml in boites_col_map:
                ci = boites_col_map[ml]
                val = row[ci] if len(row)>ci else None
                if val is not None:
                    try:
                        ameli[cip]['months'][mi] += int(float(str(val)))
                    except: pass
        count += 1

    wb4.close()
    print(f"  {fname[:40]}: {count} CIP rows, months {[MONTHS_LABEL[i] for i in month_indices]}")

print(f"  Total CIP in Ameli: {len(ameli)}")

# ── 5. ENRICH records with Ameli data ────────────────────────────────────────
ameli_matched = 0
for rec in records.values():
    cip = rec['cip13']
    if cip and cip in ameli:
        months = ameli[cip]['months']
        jan25 = months[0]; jan26 = months[12]
        rot_jan26 = jan26 / N_PHARMA
        yoy = round((jan26 - jan25) / jan25 * 100, 1) if jan25 > 0 else None
        rec.update({
            'has_ameli': True,
            'ameli_months': months,
            'ameli_jan26': jan26,
            'rot_pharma_jan26': round(rot_jan26, 2),
            'ameli_total': sum(months),
            'yoy_jan': yoy,
            'atc2': ameli[cip]['atc2'],
        })
        ameli_matched += 1
    else:
        rec.update({
            'has_ameli': False, 'ameli_months': [0]*13,
            'ameli_jan26': 0, 'rot_pharma_jan26': 0.0,
            'ameli_total': 0, 'yoy_jan': None, 'atc2': '',
        })

print(f"\nAmeli matched: {ameli_matched} / {len(records)}")

# ── 6. SORT & WRITE ──────────────────────────────────────────────────────────
recs = sorted(records.values(), key=lambda r: r['ip_qty'], reverse=True)

def js_str(s):
    return str(s or '').replace('\\','\\\\').replace('"','\\"')

today = datetime.now().strftime('%Y-%m-%d')
lines = [
    f'// Benchmark Marché — Intégral Pharma vs Ameli France',
    f'// Généré le {today} — V2 avec 13 mois Ameli + prix IP',
    f'// {len(recs)} produits IP · {ameli_matched} matchés Ameli · base 19 000 pharmacies',
    'const BENCHMARK = [',
]

for r in recs:
    months_js = '[' + ','.join(str(v) for v in r['ameli_months']) + ']'
    yoy = str(r['yoy_jan']) if r['yoy_jan'] is not None else 'null'
    froid = 'true' if r['is_froid'] else 'false'
    has   = 'true' if r['has_ameli'] else 'false'
    lines.append(
        f'  {{designation:"{js_str(r["designation"])}",categorie:"{r["categorie"]}",'
        f'ip_qty:{r["ip_qty"]},ip_ca:{r["ip_ca"]:.2f},'
        f'ip_rank_qty:{r["ip_rank_qty"]},ip_rank_ca:{r["ip_rank_ca"]},'
        f'cip13:"{r["cip13"]}",prix_ht:{r["prix_ht"]:.4f},prix_ip:{r["prix_ip"]:.4f},'
        f'remise_pct:{r["remise_pct"]:.2f},offre_ip:{r["offre_ip"]:.4f},'
        f'is_froid:{froid},has_ameli:{has},'
        f'ameli_months:{months_js},ameli_jan26:{r["ameli_jan26"]},'
        f'rot_pharma_jan26:{r["rot_pharma_jan26"]:.2f},'
        f'ameli_total:{r["ameli_total"]},yoy_jan:{yoy},atc2:"{r["atc2"]}",'
        f'artnature:"{r.get("artnature","")}"}},'
    )

lines.append('];')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines) + '\n')

sz = os.path.getsize(OUT) / 1024 / 1024
print(f"\nWritten {OUT}")
print(f"Records: {len(recs)} | Ameli matched: {ameli_matched} | File size: {sz:.1f} MB")
print(f"Top 5 by IP qty:")
for r in recs[:5]:
    print(f"  {r['designation'][:45]:<45} qty={r['ip_qty']:>8,} cat={r['categorie']} yoy={r['yoy_jan']}")
