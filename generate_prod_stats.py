#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stats reseau PAR PRODUIT (CIP13) -> crm/v2/prod-stats-data.js (window.PROD_STATS)
Pour chaque produit reellement vendu par le reseau (>=3 pharmacies) :
  - rota   : rotation moyenne (boites) par pharmacie / an  (qte / nbMois * 12 / nbPharma)
  - marge  : abandon de marge Intégral / pharmacie / an (PRINCEPS remboursables uniquement —
             jamais sur un générique ni un biosimilaire). Ce n'est PAS la marge MDL de
             l'officine : confondre les deux fausse tout argumentaire (cf. ROBOT.md §10).
  - remise : remise Integral (PPHT - prix net achat) / pharmacie / an
  - ca     : CA d'achat HT / pharmacie / an
Source prix + libelle + statut remb : STATS/stock et prix 22 06 2026.xlsx
Source ventes : crm/v2/wml-officines-data.js (WML_SALES compact)
"""
import openpyxl, re, json, os
from collections import defaultdict

ROOT = os.path.dirname(os.path.abspath(__file__))
def dernier_prix():
    """Le fichier de prix/stock LE PLUS RÉCENT (STOCKS/ puis STATS/), daté par son NOM.
    Même correction que generate_stock.py : un chemin en dur fige le script sur un vieux
    millésime sans rien signaler."""
    import glob as _g, datetime as _dt
    cands = []
    for d in ('STOCKS', 'STATS'):
        for f in _g.glob(os.path.join(ROOT, d, '*.xlsx')):
            b = os.path.basename(f)
            if b.startswith('~$') or 'stock' not in b.lower():
                continue
            m = re.search(r'(\d{2})[ _-](\d{2})[ _-](\d{4})', b)
            if not m:
                continue
            try:
                cands.append((_dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1))), f))
            except ValueError:
                pass
    if not cands:
        raise SystemExit('[prod-stats] aucun fichier de prix trouvé')
    cands.sort(key=lambda x: x[0])
    print(f'[prod-stats] prix et statuts lus dans : {os.path.basename(cands[-1][1])} '
          f'(inventaire du {cands[-1][0]})')
    return cands[-1][1]


XLSX = dernier_prix()
WML  = os.path.join(ROOT, 'crm', 'v2', 'wml-officines-data.js')
OUT  = os.path.join(ROOT, 'crm', 'v2', 'prod-stats-data.js')

BENCH = os.path.join(ROOT, 'crm', 'benchmark-data.js')

# Produits "froid" (chaine du froid) : extraits du benchmark par CIP (la seule source dispo)
froid = set()
try:
    bt = open(BENCH, encoding='utf-8').read()
    for m in re.finditer(r'\{[^{}]*\}', bt):
        o = m.group(0)
        cm = re.search(r'cip13:"(\d{12,14})"', o)
        if cm and re.search(r'is_froid:true', o):
            froid.add(cm.group(1))
except Exception:
    pass

ws = openpyxl.load_workbook(XLSX, read_only=True, data_only=True).active
hh = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
ci = hh.index('artcodebarre'); di = hh.index('artdesignation')
pi = hh.index('ppht');         ai = hh.index('afmcode')
ni = hh.index('artnature') if 'artnature' in hh else None
info = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    c = str(r[ci] or '')
    if c.isdigit() and len(c) >= 12:
        info[c] = {
            'd': (str(r[di]).strip() if r[di] else ''),
            'ppht': (float(r[pi]) if isinstance(r[pi], (int, float)) else 0.0),
            'remb': str(r[ai] or '') == 'REMBSS',
            'nat': (str(r[ni]).strip() if (ni is not None and r[ni]) else ''),
        }

# Filet de sécurité : quand `artnature` manque ou dit "Referent" à tort, un générique
# se retrouve classé princeps et reçoit un abandon de marge qu'il ne devrait PAS avoir
# (Intégral n'abandonne rien sur les génériques, ce sont les génériqueurs qui remisent).
# Mesuré le 04/08/2026 : 21 produits dans ce cas — SITAGLIPTINE BGR, ROSUVASTATINE BGR,
# SOLIFENACINE BGR, PARACETAMOL SDZ, VENLAFAXINE SDZ, ROPINIROLE EG, METRONIDAZOLE ARW…
# Ces suffixes sont des abréviations de génériqueurs telles qu'elles apparaissent en fin
# de désignation. Sous-estimer notre offre est moins grave que d'annoncer au pharmacien
# un abandon qui n'existe pas : en cas de doute, on classe en générique.
#
# ⚠️ NE PAS réduire cette liste aux génériqueurs PARTENAIRES d'Intégral
# (EG, Zentiva, Zydus, Teva — confirmés par Will le 04/08/2026). Ce sont deux
# choses différentes : ici on détecte TOUT générique, partenaire ou non, parce
# qu'Intégral n'abandonne de marge sur AUCUN d'eux. Retirer Biogaran, Mylan ou
# Sandoz reclasserait leurs produits en princeps et leur collerait un abandon
# fictif — exactement le bug que ce filet existe pour éviter.
#
# 14/08/2026 — quatre suffixes ajoutés (BGA, ARL, CRT, ARG) après les avoir trouvés
# par leur volume dans les ventes nationales : des présentations de paracétamol BGA
# à très gros volume étaient classées à tort en princeps.
# Chacun a été contrôlé contre `artnature` du catalogue avant d'être ajouté :
#   BGA 794 réf. → 743 génériques · ARL 168 → 131 · CRT 359 → 339 · ARG 69 → 54,
#   et AUCUN « Referent » sauf LEVONORGESTREL BGA (produit Biogaran, doute levé
#   dans le sens générique, conformément à la règle ci-dessus).
# ⚠️ Deux candidats ont été REJETÉS par ce même contrôle, ne pas les rajouter :
#   « REF » attrape les numéros de référence des dispositifs (MEPILEX … T REF 58122,
#   CONTOUR XT … REF 9000807) et « FRF » du sérum physiologique Fresenius —
#   aucun des deux n'est un génériqueur.
SUFFIXES_GENERIQUEURS = [
    "BGR", "EG", "TEVA", "MYL", "SDZ", "ZTL", "ARW", "CRS", "ZDS", "ACD", "KRK",
    "ALM", "EVP", "SUN", "BIOGARAN", "VIATRIS", "SANDOZ", "ZENTIVA", "ARROW",
    "CRISTERS", "MYLAN", "ACCORD", "ZYDUS", "ALMUS", "EVOLUPHARM", "SUBSTIPHARM",
    "REDDY", "EUGIA", "KRKA", "BGA", "ARL", "CRT", "ARG",
]
# Le point et la barre oblique comptent comme séparateurs : sans eux, les désignations
# du type « ATOVAQ/PROGUAN.EG » ou « AMLODIP/VALS.ARW » passaient pour des princeps.
# Effet mesuré des deux corrections réunies : 1 760 références nouvellement détectées,
# dont 1 759 confirmées génériques par le catalogue.
RX_GENERIQUEUR = re.compile(
    r"(?:^|[\s\-./])(" + "|".join(SUFFIXES_GENERIQUEURS) + r")(?:[\s\-./]|$)")

RATTRAPES = []   # journal des reclassements, affiché en fin de génération


def famille(cip, ppht, net, remb, nat, designation=''):
    """Catégories voulues (Will) : NR · Génériques+biosimilaires · Princeps par tranche de prix.
    Nature depuis artnature du fichier prix (Referent=princeps ; Generique/Generique
    Partenaire/Biosimilaire = genbio), avec rattrapage par la désignation."""
    if not remb:
        return 'nr'
    nl = (nat or '').lower()
    if 'biosimilaire' in nl: return 'biosim'
    if 'generique' in nl or 'générique' in nl: return 'gen'

    m = RX_GENERIQUEUR.search((designation or '').upper())
    if m:                                   # artnature dit princeps, la désignation dit générique
        RATTRAPES.append((cip, (designation or '').strip()[:60], m.group(1)))
        return 'gen'

    p = ppht if ppht > 0 else net           # princeps / référent → tranche de prix
    if p <= 4.33: return 'pr_low'
    if p <= 468:  return 'pr_mid'
    return 'pr_high'


def abandon_ip(p):
    """ABANDON DE MARGE Intégral sur un princeps remboursable — PAS la marge MDL de
    l'officine. (Confusion historique : ces chiffres ont longtemps été nommés « MDL ».
    La MDL officine a 5 tranches de PFHT et ses taux sont en réforme.)
    Barème : 0,18 €/boîte ≤ 4,33 € · taux sur le cœur de gamme · 19,50 €/boîte > 468 €."""
    if p <= 0: return 0
    if p <= 4.33: return 0.18
    if p <= 468: return p * TAUX_COEUR
    return 19.50


# Taux du cœur de gamme. Deux chiffres circulent pour le MÊME abandon, sur deux bases :
#   4,16 % du PFHT (prix fabricant)  = 60 % de la marge grossiste réglementée de 6,93 %
#   3,89 % du PGHT (prix grossiste)  = 4,16 ÷ 1,0693 — la base que voit le pharmacien
# Ne jamais appliquer l'un sur la base de l'autre.
#
# Vérifié le 04/08/2026 sur les ventes réelles : la colonne `ppht` de
# STATS/stock et prix*.xlsx est le TARIF GROSSISTE. L'écart mesuré PPHT → prix net
# effectivement payé, sur 1 596 princeps du cœur de gamme, a une médiane de 3,70 %
# (quartiles 3,50–3,80) — cohérent avec 3,89 %, incompatible avec 4,16 %.
# C'est donc bien 3,89 % qui s'applique ici. L'ancienne valeur 0.039 était un arrondi.
# Aligné sur `abandonBareme()` / `V2.bestPrice()`, la source de vérité du CRM.
TAUX_COEUR = 0.0389

mdl = abandon_ip   # rétrocompatibilité — ancien nom, trompeur

def charger_ventes():
    """Les ventes réseau, décodées, quel que soit leur découpage.

    ⚠️ CE SCRIPT ÉTAIT CASSÉ DEPUIS LE 15/08/2026 et personne ne l'a vu : les ventes ont
    quitté `wml-officines-data.js` pour 11 fichiers `wml-ventes-NN.js` (un iPhone ne peut
    pas lire 13 Mo d'un tenant), et la recherche de `WML_SALES = [...]` ne trouvait plus
    rien. Le fichier déjà publié continuait d'être servi, donc l'app avait l'air normale.
    Un générateur qui ne peut plus tourner ne se signale jamais tout seul.

    ⚠️ Et les lignes sont ENCODÉES : officine, commercial et produit sont des index dans
    les dictionnaires de `wml-officines-data.js`. Sans décodage, aucun CIP ne correspond
    à rien — sans la moindre erreur, juste un résultat vide.
    """
    import glob as _glob
    src = open(WML, encoding='utf-8', errors='ignore').read()
    def dico(nom):
        m = re.search(nom + r'\s*=\s*(\[[^\]]*\])', src, re.S)
        return json.loads(m.group(1)) if m else None
    dPro = dico('WML_D_PRODUITS'); dOff = dico('WML_D_OFFICINES'); dCom = dico('WML_D_COMMERCIAUX')

    morceaux = sorted(_glob.glob(os.path.join(ROOT, 'crm', 'v2', 'wml-ventes-*.js')))
    lignes = []
    for f in morceaux:
        tx = open(f, encoding='utf-8', errors='ignore').read()
        for m in re.finditer(r'\[(-?[\d.]+(?:,-?[\d.]+){6})\]', tx):
            lignes.append([float(x) for x in m.group(1).split(',')])
    if not lignes:                      # repli : ancien format d'un seul tenant
        m = re.search(r'WML_SALES\s*=\s*(\[.*?\]);', src, re.S)
        if m:
            lignes = json.loads(m.group(1))
    if not lignes:
        raise SystemExit('[prod-stats] aucune vente trouvée — ni dans wml-ventes-*.js '
                         'ni dans wml-officines-data.js. Rien ne sera écrit.')

    # décodage sur place, exactement comme v2-boot.js l.222
    if dPro and dOff and dCom and isinstance(lignes[0][0], float):
        for s in lignes:
            try:
                s[0] = dOff[int(s[0])]; s[2] = dCom[int(s[2])]; s[3] = dPro[int(s[3])]
            except (IndexError, ValueError):
                s[3] = ''
    print(f'[prod-stats] {len(lignes):,} lignes de vente lues dans {len(morceaux)} fichier(s)')
    return lignes


sales = charger_ventes()
NM = len(set(s[1] for s in sales if len(s) > 1 and s[1])) or 5   # nb de mois distincts, compté
print(f'[prod-stats] {NM} mois distincts couverts')

ag = defaultdict(lambda: {'qte': 0.0, 'ph': set(), 'ca': 0.0, 'tarif': 0.0, 'marge': 0.0, 'pw': 0.0, 'pn': 0.0})
for s in sales:                       # [pharmacyId, mois, comm, cip13, qte, puNet, mntNetHt]
    c = str(s[3]); nfo = info.get(c)
    if not nfo: continue
    qte = s[4] or 0; ca = s[6] or 0; pu = s[5] or 0; a = ag[c]
    a['qte'] += qte; a['ph'].add(str(s[0])); a['ca'] += ca
    if qte > 0 and pu > 0:            # prix net remisé = puNet pondéré, RETOURS EXCLUS (fiable)
        a['pw'] += pu * qte; a['pn'] += qte
    if nfo['ppht'] > 0:
        a['tarif'] += nfo['ppht'] * qte
        if nfo['remb']: a['marge'] += mdl(nfo['ppht']) * qte

rows = []
for c, a in ag.items():
    nph = len(a['ph'])
    if nph < 2: continue             # >=2 pharmacies (large : tous commerciaux)
    k = 12.0 / NM / nph               # -> par pharmacie / an
    net = (a['pw'] / a['pn']) if a['pn'] > 0 else ((a['ca'] / a['qte']) if a['qte'] else 0)
    ppht = round(info[c]['ppht'], 2)
    net_u = round(net, 2)                          # prix net remisé réel
    stale = 1 if (ppht > 0 and net_u > ppht) else 0   # PPHT du fichier périmé (< net réel)
    pp_eff = ppht if (ppht > 0 and not stale) else net_u
    rem_pct = round((pp_eff - net_u) / pp_eff * 100, 1) if (pp_eff > 0 and 0 < net_u <= pp_eff) else 0
    fam = famille(c, info[c]['ppht'], net, info[c]['remb'], info[c]['nat'], info[c]['d'])
    # Intégral n'abandonne de marge sur AUCUN générique ni biosimilaire : seuls les
    # princeps (familles pr_*) en portent. Le front dit déjà exactement cela
    # (v2-boot.js : « seuls eux ont l'abandon de marge »), mais l'accumulation plus
    # haut créditait tout produit remboursable sans regarder sa famille — 2 413
    # génériques sur 2 518 affichaient un abandon fictif (9 802 € au total).
    aband = a['marge'] if fam.startswith('pr_') else 0.0
    rows.append({
        'c': c, 'd': info[c]['d'][:46], 'n': nph,
        'f': fam,
        'ppht': ppht,                              # PPHT tarif grossiste
        'net': net_u,                              # prix net remisé (réel achat, hors retours)
        'rpct': rem_pct,                           # % de remise (PPHT → net)
        'stale': stale,                            # 1 = PPHT fichier à rafraîchir
        'rota': round(a['qte'] * k),
        'marge': round(aband * k),                 # abandon de marge Intégral — princeps seuls
        'remise': round(max(0, a['tarif'] - a['ca']) * k),
        'ca': round(a['ca'] * k),
    })
rows.sort(key=lambda r: (-r['n'], -r['rota']))     # classé par nb de pharmacies

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('// Stats reseau PAR PRODUIT (CIP) — rotation/marge/remise/CA par pharmacie/an — generate_prod_stats.py\n')
    f.write('window.PROD_STATS = ' + json.dumps(rows, ensure_ascii=False, separators=(',', ':')) + ';\n')
print('OK %d produits (CIP, nph>=3) sur %d mois -> %s' % (len(rows), NM, OUT))

if RATTRAPES:
    print('\n⚠️  %d produits reclassés PRINCEPS -> GÉNÉRIQUE d\'après leur désignation'
          % len(RATTRAPES))
    print('    (artnature les disait "Referent" : ils auraient reçu un abandon de marge'
          ' qu\'Intégral n\'accorde pas sur les génériques)')
    for cip, des, suf in sorted(RATTRAPES, key=lambda x: x[1]):
        print('      %-14s %-62s [%s]' % (cip, des, suf))
    print('    → si l\'un de ces produits est un VRAI princeps, retirer son suffixe de'
          ' SUFFIXES_GENERIQUEURS et le signaler à Will.')

# ── 03/09/2026 — LE FILET : jamais de conditions commerciales dans un fichier
# public. Ce script vient d'écrire un fichier suivi par git ; on découpe les
# colonnes sensibles AVANT de rendre la main. proteger_conditions.py échoue
# fort si quelque chose subsiste — un échec ici est une fuite évitée.
import subprocess as _sp, sys as _sys, os as _os
_r = _sp.run([_sys.executable, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'proteger_conditions.py')])
if _r.returncode != 0:
    _sys.exit('ARRÊT : proteger_conditions.py a échoué — ne pas committer.')
