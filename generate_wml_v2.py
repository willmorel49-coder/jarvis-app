#!/usr/bin/env python3
# ⚠️ APRÈS AVOIR LANCÉ CE SCRIPT : lancer aussi `python3 compacter_wml.py`.
# Ce générateur écrit encore le format long (le code officine, le nom du
# commercial et le code produit en toutes lettres, 437 848 fois). Le compacteur
# les remplace par des numéros de rang : 27 Mo -> 16,6 Mo, sans qu'un chiffre
# bouge, et il refuse d'écrire si un seul total diffère. Sans lui, le fichier
# repasse à 27 Mo — l'app continue de marcher (le lecteur accepte les deux
# formats), mais le premier écran redevient lent pour toute l'équipe.
# -*- coding: utf-8 -*-
"""
Génère crm/v2/wml-officines-data.js pour le CRM V2.
Source : STATS/WML_pharmacies.xlsx (officines) + STATS/WML_01..07_2026.xlsx (ventes).
Ne garde que les officines AYANT des ventes (clients actifs).
Clé produit = ARTCODEBARRE (CIP13) pour matcher BENCHMARK.cip13.
Python 3.9 compatible.
"""
import json
import os
import glob
import sqlite3
import openpyxl

# Mapping CIP/nom -> groupement issu du scraping (projet GROUPEMENTS)
import re as _re, unicodedata as _ud
GRP_DB = '/Users/williammorel/JARVIS/GROUPEMENTS/data/output/pharmacies.sqlite'
def _norm_name(s):
    s = _ud.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode().upper()
    s = _re.sub(r'\bPHARMACIE\b|\bPHIE\b|\bSELARL\b|\bSARL\b', '', s)
    return _re.sub(r'[^A-Z0-9]+', ' ', s).strip()
def _cipkey(code):
    try:
        return str(int(float(code)))
    except (TypeError, ValueError):
        m = _re.match(r'\d+', str(code or ''))
        return m.group(0) if m else str(code or '').strip()
def load_enseignes():
    """CIP -> enseigne/groupement depuis les fichiers géoloc (colonne TIRENSEIGNE)."""
    import glob
    m = {}
    # on prend le géoloc le plus récent par commercial
    files = {}
    for p in glob.glob(os.path.join(STATS, '*_geolocalisation_*.xlsx')):
        pref = os.path.basename(p).split('_geoloc')[0]
        if pref not in files or p > files[pref]:
            files[pref] = p
    for p in files.values():
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True); ws = wb.active
            it = ws.iter_rows(values_only=True); h = next(it)
            hi = {n: i for i, n in enumerate(h)}
            ti, ei = hi.get('TIRCODE'), hi.get('TIRENSEIGNE')
            if ti is None or ei is None:
                wb.close(); continue
            for r in it:
                code = r[ti] if ti < len(r) else None
                ens = r[ei] if ei < len(r) else None
                ens = (str(ens).strip() if ens else '')
                if not code or not ens:
                    continue
                m[_cipkey(code)] = ens
            wb.close()
        except Exception as e:
            print('  [ens] err', p, e)
    print('  [ens] {} CIP -> enseigne (géoloc)'.format(len(m)))
    return m


def _cpfmt(v):
    if v is None or v == '':
        return ''
    try:
        return str(int(float(v))).zfill(5)
    except (TypeError, ValueError):
        return str(v).strip()


def load_geoloc_addr():
    """CIP -> {cp, ville} depuis les fichiers géoloc (ADRCODEPOSTAL / ADRVILLE)."""
    import glob
    m = {}
    files = {}
    for p in glob.glob(os.path.join(STATS, '*_geolocalisation_*.xlsx')):
        pref = os.path.basename(p).split('_geoloc')[0]
        if pref not in files or p > files[pref]:
            files[pref] = p
    for p in files.values():
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True); ws = wb.active
            it = ws.iter_rows(values_only=True); h = next(it)
            hi = {n: i for i, n in enumerate(h)}
            ti = hi.get('TIRCODE'); ci = hi.get('ADRCODEPOSTAL'); vi = hi.get('ADRVILLE')
            if ti is None:
                wb.close(); continue
            for r in it:
                code = r[ti] if ti < len(r) else None
                if not code:
                    continue
                k = _cipkey(code)
                if k in m:
                    continue
                cp = r[ci] if (ci is not None and ci < len(r)) else None
                ville = r[vi] if (vi is not None and vi < len(r)) else None
                m[k] = {'cp': _cpfmt(cp), 'ville': (str(ville).strip() if ville else '')}
            wb.close()
        except Exception as e:
            print('  [addr] err', p, e)
    print('  [addr] {} CIP -> adresse (géoloc)'.format(len(m)))
    return m


def geocode_officines(officines):
    """Ajoute lat/lng aux officines via l'API BAN (gratuite, sans clé), avec cache."""
    import subprocess, csv as _csv2, io
    GEO_CACHE = os.path.join(STATS, 'geocode_cache.json')
    cache = {}
    if os.path.exists(GEO_CACHE):
        try:
            cache = json.load(open(GEO_CACHE, encoding='utf-8'))
        except Exception:
            cache = {}
    # adresses à géocoder (cp + ville présents, pas déjà en cache)
    todo = {}
    for o in officines:
        cp, ville = (o.get('cp') or ''), (o.get('ville') or '')
        if not cp and not ville:
            continue
        q = (cp + ' ' + ville).strip()
        if q and q not in cache:
            todo[q] = 1
    qs = list(todo.keys())
    if qs:
        print('  [geo] géocodage de {} adresses (BAN)...'.format(len(qs)))
        # CSV d'entrée
        buf = io.StringIO(); w = _csv2.writer(buf); w.writerow(['adresse'])
        for q in qs:
            w.writerow([q])
        inpath = os.path.join(STATS, '_geo_in.csv')
        open(inpath, 'w', encoding='utf-8').write(buf.getvalue())
        try:
            out = subprocess.run(
                ['curl', '-s', '-X', 'POST', '-F', 'data=@' + inpath, '-F', 'columns=adresse',
                 'https://data.geopf.fr/geocodage/search/csv/', '--max-time', '120'],
                capture_output=True, timeout=150)
            txt = out.stdout.decode('utf-8', 'ignore')
            rd = _csv2.DictReader(io.StringIO(txt))
            n = 0
            for row in rd:
                q = row.get('adresse', '')
                lat, lng = row.get('latitude', ''), row.get('longitude', '')
                try:
                    if lat and lng:
                        cache[q] = [round(float(lat), 5), round(float(lng), 5)]; n += 1
                except ValueError:
                    pass
            print('  [geo] {} adresses géocodées'.format(n))
            json.dump(cache, open(GEO_CACHE, 'w', encoding='utf-8'), ensure_ascii=False)
        except Exception as e:
            print('  [geo] err', e)
        finally:
            try:
                os.remove(inpath)
            except OSError:
                pass
    # attache lat/lng
    hit = 0
    for o in officines:
        q = ((o.get('cp') or '') + ' ' + (o.get('ville') or '')).strip()
        c = cache.get(q)
        if c:
            o['lat'], o['lng'] = c[0], c[1]; hit += 1
    print('  [geo] {}/{} officines avec coordonnées'.format(hit, len(officines)))

def load_groupements():
    cipm, namem = {}, {}
    if not os.path.exists(GRP_DB):
        print('  [grp] base absente :', GRP_DB); return cipm, namem
    try:
        c = sqlite3.connect(GRP_DB)
        for code, nom, g25, gact in c.execute(
                "select code_phirst, nom, groupement_2025, groupement_actuel from pharmacies"):
            g = (g25 or gact or '').strip()
            if not g:
                continue
            k = _cipkey(code)
            if k and k not in cipm:
                cipm[k] = g
            nm = _norm_name((nom or '').split('|')[0])
            if nm and len(nm) >= 4 and nm not in namem:
                namem[nm] = g
        c.close()
    except Exception as e:
        print('  [grp] erreur :', e)
    print('  [grp] {} CIP + {} noms -> groupement'.format(len(cipm), len(namem)))
    return cipm, namem

BASE = '/Users/williammorel/JARVIS/APP'
STATS = os.path.join(BASE, 'STATS')
# Tous les masters pharmacies (WML + MD + futurs commerciaux) — WML d'abord (prioritaire)
PHARM_FILES = sorted(glob.glob(os.path.join(STATS, '*_pharmacies.xlsx')),
                     key=lambda p: (not os.path.basename(p).startswith('WML_'), p))
# (commercial, préfixe fichier) — chaque source : les mois de MONTHS_NUM réellement déposés
SOURCES = [
    ('Will', 'WML'),
    ('Pauline G.', 'PGN'),   # Pauline Guillaumin
    ('Karine', 'KV'),
    ('Pauline S.', 'PSA'),   # Pauline Soldevila
    ('Manon', 'MD'),         # Manon
    ('Florent', 'FML'),      # Florent
    ('Morgane', 'MDC'),      # Morgane
    ('Arthur', 'ALH'),       # Arthur Lehouerou
]
MONTHS_NUM = [1, 2, 3, 4, 5, 6, 7]
NB_MOIS = len(MONTHS_NUM)
MOIS_ABBR = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin', 'Juil',
             'Août', 'Sep', 'Oct', 'Nov', 'Déc']
OUT = os.path.join(BASE, 'crm', 'v2', 'wml-officines-data.js')
PALETTE = ['#1E9E6A', '#0050E6', '#C7791A', '#6D4FC4', '#00B5D8',
           '#E0556E', '#0034A0', '#13794F', '#A65F12', '#4F3A99']


def s(v):
    return '' if v is None else str(v).strip()


def num(v):
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return 0.0


def cip13(v):
    """ARTCODEBARRE -> chaîne CIP13 propre (sans .0)."""
    if v is None:
        return ''
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return s(v)


# ── 1. Master officines (tous les *_pharmacies.xlsx) : code CIP -> infos ──
pharm = {}
for pf in PHARM_FILES:
    wb = openpyxl.load_workbook(pf, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {name: i for i, name in enumerate(hdr)}

    def col(row, name, _idx=idx):
        i = _idx.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    added = 0
    for r in rows:
        code = s(col(r, 'Code CIP'))
        if not code:
            continue
        try:
            code = str(int(float(code)))
        except (TypeError, ValueError):
            pass
        rec = {
            'name': s(col(r, 'Nom abrégé')) or ('Officine ' + code),
            'ville': s(col(r, 'Ville')),
            'cp': s(col(r, 'Code Postal')),
            'tel': s(col(r, 'Téléphone')),
            'groupement': s(col(r, 'Groupement Partenaire')),
            'grossiste': s(col(r, 'Grossiste Principal')),
            'potentiel': col(r, 'Potentiel'),
        }
        if code not in pharm:
            pharm[code] = rec
            added += 1
        else:                              # fusion : on complète les champs vides
            ex = pharm[code]
            for k in ('ville', 'cp', 'tel', 'groupement', 'grossiste'):
                if not ex.get(k) and rec.get(k):
                    ex[k] = rec[k]
            if ex.get('potentiel') in (None, '') and rec.get('potentiel') not in (None, ''):
                ex['potentiel'] = rec['potentiel']
            if (not ex.get('name') or ex['name'].startswith('Officine ')) and not rec['name'].startswith('Officine '):
                ex['name'] = rec['name']
    wb.close()
    print('  [master] %s : +%d officines (total %d)' % (os.path.basename(pf), added, len(pharm)))

# ── 1bis. Carte ARTCODE/PROCODE -> ARTCODEBARRE ──
# Certains exports bruts (ex. Karine / KV) n'ont pas la colonne ARTCODEBARRE (EAN).
# On la reconstruit à partir des fichiers déjà enrichis (98,9% des lignes KV couvertes).
ART2EAN = {}
for _p in sorted(_f for _m in MONTHS_NUM
                 for _f in glob.glob(os.path.join(STATS, '*_%02d_2026*.xlsx' % _m))):
    _wb = openpyxl.load_workbook(_p, read_only=True, data_only=True)
    _ws = _wb.active
    _it = _ws.iter_rows(values_only=True)
    _h = next(_it)
    _hi = {name: i for i, name in enumerate(_h)}
    _ib = _hi.get('ARTCODEBARRE')
    if _ib is None:
        _wb.close()
        continue
    _ia, _ip = _hi.get('ARTCODE'), _hi.get('PROCODE')
    for _r in _it:
        _ean = _r[_ib] if _ib < len(_r) else None
        if _ean in (None, ''):
            continue
        _ean = str(_ean).split('.')[0]
        if _ia is not None and _ia < len(_r) and _r[_ia] not in (None, ''):
            ART2EAN.setdefault(str(_r[_ia]).split('.')[0], _ean)
        if _ip is not None and _ip < len(_r) and _r[_ip] not in (None, ''):
            ART2EAN.setdefault(str(_r[_ip]).split('.')[0], _ean)
    _wb.close()
print('  [map] ARTCODE/PROCODE -> EAN : %d entrees' % len(ART2EAN))

# ── 2. Ventes (tous commerciaux, 6 mois chacun) ──
sales = []
active = {}  # code -> nom (fallback)
comms = {}   # code -> set des commerciaux
for comm, prefix in SOURCES:
    for mois in MONTHS_NUM:
        path = os.path.join(STATS, '%s_%02d_2026.xlsx' % (prefix, mois))
        if not os.path.exists(path):
            print('  [skip] absent :', path)
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        h = next(it)
        hi = {name: i for i, name in enumerate(h)}

        def c(row, name, _hi=hi):
            i = _hi.get(name)
            return row[i] if (i is not None and i < len(row)) else None

        n = 0
        for r in it:
            tir = c(r, 'TIRCODE')
            if tir is None:
                continue
            try:
                code = str(int(float(tir)))
            except (TypeError, ValueError):
                code = s(tir)
            if not code:
                continue
            if code not in active:
                active[code] = s(c(r, 'TIRSOCIETE'))
            comms.setdefault(code, set()).add(comm)
            # EAN : colonne ARTCODEBARRE si présente, sinon backfill via ARTCODE/PROCODE (exports bruts type KV)
            _bc = c(r, 'ARTCODEBARRE')
            if _bc in (None, ''):
                for _cn in ('ARTCODE', 'PROCODE'):
                    _v = c(r, _cn)
                    if _v not in (None, '') and str(_v).split('.')[0] in ART2EAN:
                        _bc = ART2EAN[str(_v).split('.')[0]]
                        break
            # format compact (tableau) : [pharmacyId, mois, comm, cip13, qte, puNet, mntNetHt]
            sales.append([code, mois, comm, cip13(_bc),
                          num(c(r, 'PLVQTE')), num(c(r, 'PLVPUNET')), num(c(r, 'PLVMNTNETHT'))])
            n += 1
        wb.close()
        print('  {} ({}) : {} lignes'.format(os.path.basename(path), comm, n))

# ── 3. Officines actives (avec ventes), taguées commercial + groupement ──
# Corrections manuelles (CIP -> groupement), trouvées dans le scraping par nom/ville
OVERRIDE = {
    '2128227': 'Giphar',        # Pharmacie de Canclaux (Nantes)
    '2088724': 'Pharm-Upp',     # Pharmacie de Beauvallon (26800)
    '2035185': 'Positive Pharma',  # Pharmacie Vitton (Zola), Lyon 6 — à confirmer
}
enseignes = load_enseignes()
addr = load_geoloc_addr()
grp_cip, grp_name = load_groupements()
# CA par officine (somme du montant net HT sur toute la période)
ca_by_code = {}
for srow in sales:
    ca_by_code[srow[0]] = ca_by_code.get(srow[0], 0) + (srow[6] or 0)
officines = []
nb_grp = 0
for i, code in enumerate(sorted(active.keys())):
    info = pharm.get(code, {})
    nm = info.get('name') or active[code] or ''
    # priorité : enseigne géoloc (officiel) > override scraping > scraping CIP/nom > WML_pharmacies
    grp = enseignes.get(_cipkey(code)) or OVERRIDE.get(_cipkey(code)) or grp_cip.get(_cipkey(code)) or grp_name.get(_norm_name(nm)) or info.get('groupement', '') or ''
    if grp:
        nb_grp += 1
    officines.append({
        'id': code,
        'code': code,
        'name': info.get('name') or active[code] or ('Officine ' + code),
        'ville': info.get('ville', '') or (addr.get(_cipkey(code), {}).get('ville', '')),
        'cp': info.get('cp', '') or (addr.get(_cipkey(code), {}).get('cp', '')),
        'tel': info.get('tel', ''),
        'groupement': grp,
        'potentiel': info.get('potentiel'),
        'ca': round(ca_by_code.get(code, 0)),
        'comms': sorted(comms.get(code, [])),
        'color': PALETTE[i % len(PALETTE)],
    })
print('  officines avec groupement : {}/{}'.format(nb_grp, len(officines)))

# ── 3bis. Normalisation / fusion des groupements ──
# Fusionne : casse+accents+ponctuation (GIPHAR=Giphar), alias officiels,
# + règles métier (Aelia=Co&Pharm, Normandie/Bretagne Pharma=OPSO Santé).
import csv as _csv
ALIAS_CSV = '/Users/williammorel/JARVIS/GROUPEMENTS/config/groupements_alias.csv'
def _gkey(s):
    s = _ud.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode().upper()
    return _re.sub(r'[^A-Z0-9]', '', s)
MERGE = {  # règles métier (clé normalisée -> nom canonique)
    'AELIA': 'Co&Pharm', 'COPHARM': 'Co&Pharm',
    'NORMANDIEPHARMA': 'OPSO Santé', 'BRETAGNEPHARMA': 'OPSO Santé',
    'BRETAGNENORMANDIEPHARMA': 'OPSO Santé', 'BRETAGNEETNORMANDIEPHARMA': 'OPSO Santé',
    'OPSOSANTE': 'OPSO Santé',
    'UPP07': 'UPP', 'UPP38': 'UPP', 'UPP26': 'UPP', 'UPPCOTESDURHONE': 'UPP',
}
alias = {}
try:
    with open(ALIAS_CSV, encoding='utf-8') as f:
        rdr = _csv.reader(f, delimiter=';'); next(rdr, None)
        for row in rdr:
            if not row or not row[0].strip():
                continue
            off = row[0].strip()
            for v in [off] + ((row[1] if len(row) > 1 else '').split('|')):
                v = v.strip()
                if v:
                    alias.setdefault(_gkey(v), off)
except Exception as e:
    print('  [alias] err', e)
# label d'affichage par clé normalisée (depuis nos données)
disp = {}
for o in officines:
    g = o.get('groupement')
    if not g:
        continue
    n = _gkey(g)
    if n in MERGE:
        continue
    if n in alias:
        disp[n] = alias[n]
    else:
        cur = disp.get(n)
        if cur is None or (cur.isupper() and not g.isupper()):
            disp[n] = g  # préfère une casse non tout-majuscule
for o in officines:
    g = o.get('groupement')
    if not g:
        continue
    n = _gkey(g)
    o['groupement'] = MERGE.get(n) or disp.get(n) or g
from collections import Counter as _C
_c = _C(o['groupement'] for o in officines if o.get('groupement'))
print('  groupements après fusion : {} distincts'.format(len(_c)))

# ── 3ter. Logos par groupement (même canonicalisation appliquée aux noms de logos) ──
def _canon(g):
    n = _gkey(g)
    return MERGE.get(n) or alias.get(n) or disp.get(n) or g
LOGOS_JSON = '/Users/williammorel/JARVIS/GROUPEMENTS/data/sources/logos.json'
grp_logos = {}
try:
    raw_logos = json.load(open(LOGOS_JSON, encoding='utf-8'))
    # Matching conservateur : clé normalisée (_gkey) + retrait d'un mot
    # générique en préfixe/suffixe + match par token entier (>=5 car.).
    # AUCUNE inclusion approximative (un mauvais logo est pire que pas de
    # logo). Jamais un mot générique seul.
    _GENERIC = {'PHARMACIE', 'PHARMACIES', 'PHARMA', 'GROUPE', 'GROUPEMENT', 'SANTE', 'RESEAU', 'LES'}
    _STOP = {'PHARMA', 'PHARMACIE', 'PHARMACIES', 'SANTE', 'GROUPE', 'GROUPEMENT'}

    def _logo_variants(gk):
        vs = {gk}
        for n in _GENERIC:
            if gk.endswith(n) and len(gk) - len(n) >= 4:
                vs.add(gk[:-len(n)])
            if gk.startswith(n) and len(gk) - len(n) >= 4:
                vs.add(gk[len(n):])
        return vs

    logo_idx = {}
    for name, b64 in raw_logos.items():
        for v in _logo_variants(_gkey(name)):
            if v and v not in _STOP:
                logo_idx.setdefault(v, b64)

    grps_all = set(o['groupement'] for o in officines if o.get('groupement'))
    for g in grps_all:
        hit = None
        for v in _logo_variants(_gkey(g)):
            if v in logo_idx and v not in _STOP:
                hit = logo_idx[v]; break
        if not hit:
            for tok in _re.split(r'[\s&/,\-]+', g):
                tk = _gkey(tok)
                if len(tk) >= 5 and tk not in _STOP and tk in logo_idx:
                    hit = logo_idx[tk]; break
        if hit:
            grp_logos[g] = hit
    print('  logos rattachés : {}/{} groupements'.format(len(grp_logos), len(grps_all)))
except Exception as e:
    print('  [logos] err', e)

# ── 3quater. Géocodage des officines (carte de secteur) ──
geocode_officines(officines)

# ── 3quinquies. Détail par officine pour la FICHE de la carte Copilote ──
# (CA mensuel + top produits + nb produits + potentiel), fichier compact chargé à la demande.
def load_benchmark_names():
    names = {}
    try:
        txt = open(os.path.join(BASE, 'crm', 'benchmark-data.js'), encoding='utf-8').read()
        for m in _re.finditer(r'designation:"([^"]*)"[^}]*?cip13:"(\d+)"', txt):
            names[m.group(2)] = m.group(1)
    except Exception as e:
        print('  [bench] err', e)
    return names

name_by_cip = load_benchmark_names()
pot_by_code = {o['id']: o.get('potentiel') for o in officines}
_det = {}
for srow in sales:
    code, mois, comm, cip, qte, pu, mnt = srow
    d = _det.setdefault(code, {'m': [0] * NB_MOIS, 'prod': {}})
    if isinstance(mois, int) and 1 <= mois <= NB_MOIS:
        d['m'][mois - 1] += (mnt or 0)
    d['prod'][cip] = d['prod'].get(cip, 0) + (mnt or 0)
CARTE_DETAIL = {}
for code, d in _det.items():
    top = sorted(d['prod'].items(), key=lambda x: -x[1])[:6]
    pot = pot_by_code.get(code)
    CARTE_DETAIL[code] = {
        'm': [round(x) for x in d['m']],
        'top': [[name_by_cip.get(c, c), round(v)] for c, v in top if v > 0],
        'np': len(d['prod']),
        'pot': (round(pot) if isinstance(pot, (int, float)) else None),
    }
DET_OUT = os.path.join(BASE, 'crm', 'v2', 'carte-detail.js')
with open(DET_OUT, 'w', encoding='utf-8') as f:
    f.write('// Détail par officine pour la fiche de la carte Copilote (CA mensuel, top produits).\n')
    f.write('// {id: {m:[%s..%s], top:[[nom,ca]...], np:nbProduits, pot:potentiel}}\n'
        % (MOIS_ABBR[MONTHS_NUM[0] - 1].lower(), MOIS_ABBR[MONTHS_NUM[-1] - 1].lower()))
    f.write('window.CARTE_DETAIL=' + json.dumps(CARTE_DETAIL, ensure_ascii=False, separators=(',', ':')) + ';\n')
print('  [fiche] carte-detail.js : {} officines, {:.0f} Ko'.format(len(CARTE_DETAIL), os.path.getsize(DET_OUT) / 1024))

# ── 4. Écriture JS ──
months_lbl = '%s-%s 2026' % (MOIS_ABBR[MONTHS_NUM[0] - 1], MOIS_ABBR[MONTHS_NUM[-1] - 1])
WML_MOIS = ['2026-%02d' % m for m in MONTHS_NUM]   # période couverte, lue par le front
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('// WML Officines + Ventes — source de vérité CRM V2\n')
    f.write('// {} officines actives · {} lignes de ventes · {}\n'.format(
        len(officines), len(sales), months_lbl))
    f.write('// Généré par generate_wml_v2.py depuis STATS/WML_pharmacies + WML_%02d..%02d_2026\n'
        % (MONTHS_NUM[0], MONTHS_NUM[-1]))
    f.write('// WML_SALES format compact : [pharmacyId, mois, commercial, cip13, qte, puNet, mntNetHt]\n')
    f.write('const WML_OFFICINES = ' + json.dumps(officines, ensure_ascii=False, separators=(',', ':')) + ';\n')
    f.write('const WML_SALES = ' + json.dumps(sales, ensure_ascii=False, separators=(',', ':')) + ';\n')
    f.write('const GRP_LOGOS = ' + json.dumps(grp_logos, ensure_ascii=False, separators=(',', ':')) + ';\n')
    f.write('const WML_MOIS = ' + json.dumps(WML_MOIS) + ';\n')
    f.write('try{window.WML_OFFICINES=WML_OFFICINES;window.WML_SALES=WML_SALES;window.GRP_LOGOS=GRP_LOGOS;window.WML_MOIS=WML_MOIS;}catch(e){}\n')

size = os.path.getsize(OUT)
print('\nOK -> {}'.format(OUT))
print('  {} officines · {} ventes · {:.0f} Ko'.format(len(officines), len(sales), size / 1024))

# ── 5. Découpage OBLIGATOIRE ──────────────────────────────────────────────
# ⚠️ 15/08/2026 — ne PAS retirer cet appel.
# Le fichier écrit ci-dessus fait ~17 Mo, dont 13 Mo de ventes sur UNE SEULE
# LIGNE. Mesuré sur l'iPhone de Will : Safari n'arrive pas à le lire jusqu'au
# bout, abandonne en route, et comme les données ne sont publiées qu'à la
# dernière ligne, il ne reste RIEN — l'app affichait alors 22 officines au lieu
# de 690, en puisant dans de vieilles tables, sans le moindre message.
# JARVIS a été inutilisable en mobilité pendant deux jours à cause de ça.
#
# `decouper_wml.py` casse ce fichier en tranches de 1,5 Mo et sort les logos.
# Sans lui, la panne revient à la première régénération.
import sys, subprocess  # noqa: E402  (import tardif : ce script est un outil, pas une lib)
# ⚠️ Le COMPACTAGE passe AVANT le découpage : il a besoin du fichier ENTIER.
# Sans lui on republie 27 Mo de ventes en clair — la panne iPhone des 13-14/08.
# Il manquait à ce pipeline : seul un lancement à la main le faisait, donc le
# robot `stats-commercial` produisait un fichier non compacté sans rien signaler.
# `compacter_wml.py` lit un chemin RELATIF, d'où le cwd explicite.
print('\n── Compactage (officine/commercial/produit → rang) ──')
subprocess.run([sys.executable, os.path.join(BASE, 'compacter_wml.py')], cwd=BASE, check=True)

print('\n── Découpage pour les téléphones ──')
subprocess.run([sys.executable, os.path.join(BASE, 'decouper_wml.py'), OUT], check=True)
