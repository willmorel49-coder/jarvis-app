#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Génère base-biosimilaires.xlsx depuis biosimilaires-export.json.
Onglets : Synthèse molécules · Biosimilaires (détail) · Substituables officine.
Aucune dépendance hors openpyxl. python3 generate_biosim_xlsx.py
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
data = json.load(open(os.path.join(ROOT, "biosimilaires-export.json"), encoding="utf-8"))
meta, molecules = data["meta"], data["molecules"]

NAVY = "0F2A47"
ORANGE = "E8722B"
GREEN = "1E9E5A"
GREY = "EEF1F5"
WHITE = "FFFFFF"

hfont = Font(bold=True, color=WHITE, size=11, name="Calibri")
hfill = PatternFill("solid", fgColor=NAVY)
sub_fill = PatternFill("solid", fgColor="E9F7EF")
part_fill = PatternFill("solid", fgColor="FDECDD")
thin = Side(style="thin", color="D5DBE2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------- Onglet 1 : Synthèse molécules ----------------
ws = wb.active
ws.title = "Synthèse molécules"
cols = ["DCI (molécule)", "ATC", "Aire thérapeutique", "Princeps réf.", "Labo princeps",
        "Canal", "Substituable officine", "Date arrêté", "Nb biosim.",
        "Réf. chez Intégral", "Partenaire dispo", "Pénétration biosim. (%)",
        "Boîtes biosim. (Ameli)", "CA/pharma/an (molécule)", "Note"]
ws.append(cols)
style_header(ws, len(cols))
for m in molecules:
    ms = m.get("mol_stats") or {}
    ws.append([
        m["dci"], m["atc"], m["aire"], m["reference"], m["reference_labo"],
        m["canal"], "OUI" if m["substituable"] else "non",
        m.get("substituable_date") or "", m["nb_biosim"],
        "OUI" if m["referenced_ip"] else "hors périmètre",
        "OUI" if m["has_partenaire"] else "",
        m["penetration"] if m["penetration"] is not None else "",
        m["biosim_ameli_boxes"] or "",
        ms.get("ca_pharma_an", "") if ms else "",
        m.get("note") or "",
    ])
    r = ws.max_row
    if m["substituable"]:
        ws.cell(row=r, column=7).fill = sub_fill
        ws.cell(row=r, column=7).font = Font(bold=True, color=GREEN)
    if m["has_partenaire"]:
        ws.cell(row=r, column=11).fill = part_fill
        ws.cell(row=r, column=11).font = Font(bold=True, color=ORANGE)
autosize(ws, [22, 9, 26, 20, 20, 9, 13, 12, 9, 15, 12, 13, 15, 16, 40])
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols)):
    for c in row:
        c.border = border

# ---------------- Onglet 2 : Biosimilaires (détail) ----------------
ws2 = wb.create_sheet("Biosimilaires (détail)")
cols2 = ["DCI", "ATC", "Aire", "Substituable", "Canal", "Type", "Marque", "Laboratoire",
         "Année", "Partenaire IP", "Acteur majeur", "Dispo Intégral",
         "Boîtes Ameli (France)", "CA Ameli (€)", "Qté vendue IP", "Stock dispo",
         "Prix PPHT (€)", "Prix net IP (€)", "Boîtes/pharma/an", "Labo réel (données IP)", "CIP13 liés"]
ws2.append(cols2)
style_header(ws2, len(cols2))


def add_line(m, typ, name, labo, annee, e):
    ws2.append([
        m["dci"], m["atc"], m["aire"], "OUI" if m["substituable"] else "non", m["canal"],
        typ, name, labo, annee or "",
        "OUI" if e.get("partenaire") else "", "OUI" if e.get("acteur_majeur") else "",
        "OUI" if e.get("disponible_ip") else "",
        e.get("ameli_boxes") or "", e.get("ameli_ca") or "",
        e.get("ip_qty") or "", e.get("stock_dispo") or "",
        e.get("prix_ppht") if e.get("prix_ppht") is not None else "",
        e.get("prix_ip") if e.get("prix_ip") is not None else "",
        e.get("boites_par_pharma_an") or "",
        " / ".join(e.get("labos_reels") or []),
        ", ".join(e.get("cips") or []),
    ])
    r = ws2.max_row
    if typ == "Princeps":
        for c in range(1, len(cols2) + 1):
            ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREY)
    if e.get("partenaire"):
        ws2.cell(row=r, column=10).fill = part_fill
        ws2.cell(row=r, column=10).font = Font(bold=True, color=ORANGE)


for m in molecules:
    add_line(m, "Princeps", m["reference"], m["reference_labo"], "", m["reference_enrich"])
    for b in m["biosimilaires"]:
        add_line(m, "Biosimilaire", b["nom"], b["labo"], b.get("annee"), b)
autosize(ws2, [20, 8, 22, 11, 8, 12, 22, 22, 7, 12, 12, 12, 16, 14, 12, 11, 12, 12, 14, 24, 26])
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(cols2)):
    for c in row:
        c.border = border

# ---------------- Onglet 3 : Substituables officine (angle commercial) ----------------
ws3 = wb.create_sheet("Substituables officine")
ws3.append(["Les 11 groupes substituables en officine — arrêté du 10 avril 2026 (+ historiques)"])
ws3.cell(row=1, column=1).font = Font(bold=True, size=13, color=NAVY)
ws3.append([])
cols3 = ["DCI", "ATC", "Aire", "Princeps", "Date arrêté", "Nb biosim.",
         "Réf. Intégral", "Biosim. partenaires (Zentiva/EG/Teva)", "Pénétration (%)", "Boîtes biosim. (Ameli)"]
ws3.append(cols3)
style_header(ws3, len(cols3), row=3)
subs = [m for m in molecules if m["substituable"]]
subs.sort(key=lambda m: (0 if m["has_partenaire"] else 1, -(m["biosim_ameli_boxes"] or 0)))
for m in subs:
    parts = sorted({b["labo"] for b in m["biosimilaires"] if b.get("partenaire")} |
                   {l for b in m["biosimilaires"] for l in (b.get("labos_reels") or [])
                    if any(p.upper() in l.upper() for p in meta["partenaires_ip"])})
    ws3.append([
        m["dci"], m["atc"], m["aire"], m["reference"], m.get("substituable_date") or "",
        m["nb_biosim"], "OUI" if m["referenced_ip"] else "hors périmètre",
        ", ".join(parts) if parts else "—",
        m["penetration"] if m["penetration"] is not None else "",
        m["biosim_ameli_boxes"] or "",
    ])
    r = ws3.max_row
    if m["has_partenaire"]:
        for c in range(1, len(cols3) + 1):
            ws3.cell(row=r, column=c).fill = part_fill
autosize(ws3, [20, 9, 24, 18, 14, 9, 15, 34, 13, 18])
for row in ws3.iter_rows(min_row=4, max_row=ws3.max_row, max_col=len(cols3)):
    for c in row:
        c.border = border

out = os.path.join(ROOT, "base-biosimilaires.xlsx")
wb.save(out)
print(f"✓ {out}")
print(f"  {meta['nb_molecules']} molécules · {meta['nb_biosimilaires']} biosimilaires · "
      f"{meta['nb_substituables']} substituables · {len(subs)} lignes onglet substituables")
