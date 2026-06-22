# -*- coding: utf-8 -*-
"""Prix PPHT officiel (prix pharmacien HT = tarif grossiste) par CIP13, pour les
produits NON remboursables (NR). Source : STATS/stock et prix <date>.xlsx.
→ crm/v2/ppht-data.js  (window.PPHT = {cip13: ppht})
Sert à corriger le prix NR partout (benchmark prix_ht au chargement + catalogue NR).
Python 3.9.
"""
import re
import json
import glob
import os
import openpyxl

# fichier stock+prix le plus récent
cands = sorted(glob.glob('STATS/stock et prix*.xlsx'), key=os.path.getmtime, reverse=True)
SRC = cands[0]
OUT = 'crm/v2/ppht-data.js'
print('source:', SRC)

ws = openpyxl.load_workbook(SRC, read_only=True, data_only=True).active
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {h: i for i, h in enumerate(hdr) if h}


def g(row, name):
    i = idx.get(name)
    return row[i] if (i is not None and i < len(row)) else None


ppht = {}
allp = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    cip = str(g(r, 'artcodebarre') or '').strip()
    p = g(r, 'ppht')
    afm = str(g(r, 'afmcode') or '')
    if not (cip.isdigit() and len(cip) >= 12):
        continue
    if not isinstance(p, (int, float)) or p <= 0:
        continue
    p = round(float(p), 2)
    allp[cip] = p
    if afm != 'REMBSS':                 # NR uniquement
        ppht[cip] = p

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('// Prix PPHT (tarif grossiste HT) par CIP13 — produits NR — generate_ppht.py\n')
    f.write('window.PPHT = ' + json.dumps(ppht, ensure_ascii=False, separators=(',', ':')) + ';\n')

print('OK -> %s : %d prix NR (sur %d prix total)' % (OUT, len(ppht), len(allp)))
