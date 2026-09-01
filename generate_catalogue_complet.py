#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LE CATALOGUE INTÉGRAL PHARMA EN ENTIER -> crm/v2/catalogue-complet-data.js

Pourquoi ce fichier existe : l'écran « Produits » ne connaissait que les
6 292 références que le réseau a réellement achetées (PROD_STATS, seuil
>= 3 pharmacies). Le catalogue en compte 16 305 (16 439 lignes, dédoublonnées
par code-barres). Les 10 013 autres — dont
tout ce qu'on référence sans encore le vendre — étaient INVISIBLES dans
l'app : impossible de les proposer à une officine, impossible d'en faire
une fiche marketing.

Sources (toutes locales, hors dépôt) :
  · STATS/stock et prix 22 06 2026.xlsx  — LE catalogue : désignation, labo,
    molécule, nature (Referent/Generique/Biosimilaire), AFMCODE, PPHT, stock.
  · crm/v2/prod-stats-data.js            — stats réseau (nb pharmacies, net réel).

⚠️ STATS/prix-recents.json a été ESSAYÉ comme source de prix, puis ÉCARTÉ le
01/09/2026. Sa colonne « ppht » vient de PLVPUBRUT des fichiers de ventes, qui
porte le prix RÉELLEMENT FACTURÉ — donc le prix d'offre quand une offre
laboratoire s'applique. Mesuré sur DOLIPRANE 1000MG BT8 : il annonçait un tarif
de 0,78 € là où le tarif grossiste est 1,06 € (0,78 = l'offre Opella). Un tarif
faux déplace le produit de tranche et fausse l'abandon de marge derrière. Le
fichier « stock et prix » reste la seule source du TARIF.
  · crm/v2/generiqueurs-data.js          — répertoire officiel des génériques
    (BDPM), 11 381 CIP13. C'est LUI qui dit qu'un produit est un générique
    quand la colonne `artnature` du fichier de prix est vide — 3 692 des
    références remboursables sont dans ce cas.

⚠️ Les règles de classement et d'abandon de marge sont celles de
`generate_prod_stats.py`, qui reste la source de vérité. Elles sont reprises
ici parce que ce script tourne seul — et un CONTRÔLE final les rejoue sur les
6 292 références dont le réseau connaît déjà la réponse. Si le calcul y fait
passer un générique pour un princeps, il lui collerait un abandon de marge
inventé : au-delà de 3 cas, le script refuse de publier.

⚠️ LIMITE CONNUE, mesurée : 49 références portent `artnature = Referent` chez
un laboratoire qui ne vend que du générique et sont absentes du répertoire
BDPM. Il y a des deux dans le lot — de vrais princeps distribués par un
génériqueur (LYRICA, PARLODEL chez Biogaran) et de vrais génériques mal
étiquetés (LEVONORGESTREL BGA). Aucune règle automatique ne les sépare : les
classer en bloc ferait perdre un abandon réel sur les premiers. Elles restent
donc au classement du fichier de prix.
"""
import json
import os
import re
import sys
from collections import Counter

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, 'STATS', 'stock et prix 22 06 2026.xlsx')
PROD_STATS = os.path.join(ROOT, 'crm', 'v2', 'prod-stats-data.js')
GENERIQUEURS = os.path.join(ROOT, 'crm', 'v2', 'generiqueurs-data.js')
OUT = os.path.join(ROOT, 'crm', 'v2', 'catalogue-complet-data.js')

# ── Règles métier (miroir de generate_prod_stats.py) ────────────────────
TAUX_COEUR = 0.0389          # abandon du cœur de gamme, base PGHT
SUFFIXES_GENERIQUEURS = [
    "BGR", "EG", "TEVA", "MYL", "SDZ", "ZTL", "ARW", "CRS", "ZDS", "ACD", "KRK",
    "ALM", "EVP", "SUN", "BIOGARAN", "VIATRIS", "SANDOZ", "ZENTIVA", "ARROW",
    "CRISTERS", "MYLAN", "ACCORD", "ZYDUS", "ALMUS", "EVOLUPHARM", "SUBSTIPHARM",
    "REDDY", "EUGIA", "KRKA",
]
# ⚠️ La borne de generate_prod_stats.py ne connaît que l'espace et le tiret.
# Le fichier de prix écrit aussi « BIMATOPR.SDZ » et « HYDROCHLOROT.ARW » : le
# point et la barre oblique séparent tout autant, et sans eux ces génériques
# repartaient en princeps — donc avec un abandon de marge qui n'existe pas.
RX_GENERIQUEUR = re.compile(
    r"(?:^|[\s\-./])(" + "|".join(SUFFIXES_GENERIQUEURS) + r")(?:[\s\-./]|$)")


def famille(cip, ppht, remb, nat, designation='', repertoire=None, labo='',
            labos_generiqueurs=()):
    """NR · Génériques · Biosimilaires · Princeps par tranche de prix."""
    if not remb:
        return 'nr'
    nl = (nat or '').lower()
    if 'biosimilaire' in nl:
        return 'biosim'
    if 'generique' in nl or 'générique' in nl:
        return 'gen'
    # Le répertoire officiel des génériques prime sur toute déduction : quand
    # la BDPM dit qu'un CIP est un générique, il n'y a rien à interpréter.
    if repertoire and cip in repertoire:
        return 'gen'
    if RX_GENERIQUEUR.search((designation or '').upper()):
        return 'gen'
    # Dernier filet, quand `artnature` est vide : le laboratoire. Ces maisons-là
    # ne vendent QUE du générique (mesuré : 96 à 100 % de leurs remboursables
    # sont au répertoire BDPM). Le reste est du générique trop récent ou trop
    # ancien pour y figurer — et devant le doute, on classe en générique :
    # sous-estimer notre offre est moins grave qu'annoncer au pharmacien un
    # abandon de marge qui n'existe pas.
    if not nl and labo in labos_generiqueurs:
        return 'gen'
    if ppht <= 4.33:
        return 'pr_low'
    if ppht <= 468:
        return 'pr_mid'
    return 'pr_high'


def abandon_ip(p):
    """ABANDON DE MARGE Intégral sur un princeps remboursable — jamais la MDL."""
    if p <= 0:
        return 0.0
    if p <= 4.33:
        return 0.18
    if p <= 468:
        return p * TAUX_COEUR
    return 19.50


def porte_abandon(f):
    return f.startswith('pr_')


def r2(x):
    return round(float(x) + 1e-9, 2)


def txt(v):
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ('#N/A', 'None', 'nan') else s


def nombre(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def lire_catalogue():
    ws = openpyxl.load_workbook(XLSX, read_only=True, data_only=True).active
    hh = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {n: hh.index(n) for n in
           ('artcodebarre', 'artdesignation', 'artcollection', 'artmol',
            'artnature', 'afmcode', 'ppht', 'stockdispo', 'artclasse')}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        cip = str(r[col['artcodebarre']] or '').strip()
        if not (cip.isdigit() and len(cip) >= 12):
            continue
        d = txt(r[col['artdesignation']])
        if not d:
            continue
        # Un CIP peut revenir sur deux lignes : on garde celle qui a du stock.
        stock = int(nombre(r[col['stockdispo']]))
        if cip in out and out[cip]['s'] >= stock:
            continue
        out[cip] = {
            'c': cip,
            'd': d,
            'labo': txt(r[col['artcollection']]),
            'mol': txt(r[col['artmol']]),
            'nat': txt(r[col['artnature']]),
            'remb': txt(r[col['afmcode']]) == 'REMBSS',
            'afm': txt(r[col['afmcode']]),
            'ppht': r2(nombre(r[col['ppht']])),
            's': stock,
            'mitm': txt(r[col['artclasse']]).upper() == 'MITM',
        }
    return out


def lire_prod_stats():
    s = open(PROD_STATS, encoding='utf-8').read()
    i = s.index('[')
    j = s.rindex(']')
    return {str(o['c']): o for o in json.loads(s[i:j + 1])}


def lire_repertoire():
    s = open(GENERIQUEURS, encoding='utf-8').read()
    return set(json.loads(s[s.index('{'):s.rindex('}') + 1]).keys())


def labos_generiqueurs(cat, repertoire, part=0.80, mini=20):
    """Les laboratoires qui ne vendent QUE du générique, DÉDUITS des données :
    part de leurs remboursables présents au répertoire officiel. Rien n'est
    écrit en dur — si Intégral change de fournisseurs, la liste suit."""
    tot, dedans = Counter(), Counter()
    for o in cat.values():
        if not o['remb'] or not o['labo']:
            continue
        tot[o['labo']] += 1
        if o['c'] in repertoire:
            dedans[o['labo']] += 1
    return {l: (tot[l], dedans[l]) for l in tot
            if tot[l] >= mini and dedans[l] / tot[l] >= part}


def main():
    for p in (XLSX, PROD_STATS, GENERIQUEURS):
        if not os.path.exists(p):
            sys.exit('Source manquante : ' + p)

    cat = lire_catalogue()
    ps = lire_prod_stats()
    repertoire = lire_repertoire()
    generiqueurs = labos_generiqueurs(cat, repertoire)
    print('catalogue : %d références · stats réseau : %d · répertoire génériques : %d'
          % (len(cat), len(ps), len(repertoire)))
    print('laboratoires 100 %% générique déduits des données : %d (%s…)'
          % (len(generiqueurs),
             ', '.join(sorted(generiqueurs, key=lambda l: -generiqueurs[l][0])[:6])))

    labos, mols = {}, {}
    rows = []
    fam_compte = Counter()
    faux_abandons, ecarts_fam = [], []

    for cip in sorted(cat):
        o = cat[cip]
        ppht = o['ppht']
        f = calcule = famille(cip, ppht, o['remb'], o['nat'], o['d'],
                              repertoire, o['labo'], generiqueurs)
        # Quand le réseau connaît déjà le produit, sa famille EST celle que
        # l'app affiche partout ailleurs. On ne la recalcule pas : deux écrans
        # qui classent le même produit différemment, c'est un bug visible.
        if cip in ps and ps[cip].get('f'):
            f = ps[cip]['f']
            if f != calcule:
                ecarts_fam.append((cip, o['d'][:40], f, calcule))
                # Le seul sens dangereux : le réseau dit générique/biosimilaire
                # et mon calcul dit princeps. Ça collerait un abandon de marge
                # fictif à un produit qui n'en porte aucun.
                if not porte_abandon(f) and porte_abandon(calcule):
                    faux_abandons.append((cip, o['d'][:40], f, calcule))
        fam_compte[f] += 1

        # NET : le net réseau d'abord — c'est CELUI QUE L'APP AFFICHE déjà
        # partout ailleurs, et deux écrans qui donnent deux prix pour le même
        # produit, c'est un bug visible. Le barème ensuite, pour tout ce que le
        # réseau n'a jamais acheté ; et il NE S'APPLIQUE QU'AUX PRINCEPS.
        net = 0.0
        if cip in ps and nombre(ps[cip].get('net')) > 0:
            # ⚠️ On NE corrige PAS ce chiffre, même quand il dépasse le tarif du
            # 22/06 : c'est celui que l'app affiche déjà partout. Le « corriger »
            # posait un second prix pour 494 des 6 292 produits du réseau (jusqu'à
            # 18 € d'écart sur XGEVA) — deux vérités dans la même app. Un net
            # au-dessus du tarif veut dire que le tarif a bougé depuis, pas que
            # la facture est fausse ; les colonnes d'abandon savent se taire
            # toutes seules dans ce cas.
            net = r2(ps[cip]['net'])
        elif ppht > 0:
            net = r2(ppht - abandon_ip(ppht)) if porte_abandon(f) else ppht


        li = labos.setdefault(o['labo'], len(labos)) if o['labo'] else -1
        mi = mols.setdefault(o['mol'], len(mols)) if o['mol'] else -1
        st = ps.get(cip)
        rows.append([
            cip, o['d'], li, mi, f, ppht, net, o['s'],
            1 if o['mitm'] else 0,
            int(st['n']) if st and st.get('n') else 0,      # nb pharmacies acheteuses
        ])

    # ── Le contrôle qui compte ─────────────────────────────────────────
    # Les 6 292 produits que le réseau achète sont les seuls dont on connaisse
    # déjà la réponse. Si mon classement en fait des princeps là où l'app dit
    # générique, la même erreur frappe en silence les 10 000 autres — et
    # chacun ressort avec un abandon de marge inventé.
    # Au-delà de 3, ce n'est plus un accident de saisie mais une règle cassée :
    # la même erreur frappe alors en silence les 10 000 références que le
    # réseau ne connaît pas. (Mesuré : la borne de mot sans le point en
    # laissait passer 26 d'un coup.)
    if faux_abandons:
        print('\n⚠️  %d produit(s) que le réseau dit générique/biosimilaire et que '
              'le calcul dit princeps :' % len(faux_abandons))
        for e in faux_abandons[:12]:
            print('   %s %-40s réseau=%s calcul=%s' % e)
        print('   (le classement du réseau est retenu pour eux ; la cause est une '
              'colonne `artnature` fausse dans le fichier de prix)')
    if len(faux_abandons) > 3:
        sys.exit('Abandon de marge fictif en série — corriger la règle avant de publier.')

    labos_l = [''] * len(labos)
    for k, v in labos.items():
        labos_l[v] = k
    mols_l = [''] * len(mols)
    for k, v in mols.items():
        mols_l[v] = k

    meta = {
        'source': 'STATS/stock et prix 22 06 2026.xlsx',
        'arrete': '2026-06-22',
        'n': len(rows),
        'enStock': sum(1 for r in rows if r[7] > 0),
        'connusReseau': sum(1 for r in rows if r[9] > 0),
        'cols': ['cip', 'd', 'labo', 'mol', 'fam', 'ppht', 'net', 'stock', 'mitm', 'nbPharm'],
    }
    # Une ligne JS par référence : Safari n'avale pas un fichier d'un seul bloc.
    corps = ',\n'.join(json.dumps(r, ensure_ascii=False, separators=(',', ':')) for r in rows)
    js = ('/* Catalogue Intégral Pharma COMPLET — généré par generate_catalogue_complet.py.\n'
          '   Ne pas éditer à la main. %d références au %s. */\n'
          'window.CATALOGUE_COMPLET={\n'
          'meta:%s,\n'
          'labos:%s,\n'
          'mols:%s,\n'
          'rows:[\n%s\n]};\n'
          % (len(rows), meta['arrete'],
             json.dumps(meta, ensure_ascii=False),
             json.dumps(labos_l, ensure_ascii=False, separators=(',', ':')),
             json.dumps(mols_l, ensure_ascii=False, separators=(',', ':')),
             corps))
    open(OUT, 'w', encoding='utf-8').write(js)

    print('\n→ %s · %.2f Mo' % (OUT, os.path.getsize(OUT) / 1048576))
    print('   %d références · %d en stock · %d déjà commandées par le réseau'
          % (meta['n'], meta['enStock'], meta['connusReseau']))
    print('   familles : ' + ' · '.join('%s %d' % kv for kv in fam_compte.most_common()))
    print('   %d laboratoires · %d molécules' % (len(labos_l), len(mols_l)))
    print('   contrôle : %d CIP communs avec le réseau · %d familles reprises '
          'du réseau · 0 abandon fictif en série'
          % (sum(1 for r in rows if r[0] in ps), len(ecarts_fam)))


if __name__ == '__main__':
    main()
