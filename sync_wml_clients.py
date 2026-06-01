#!/usr/bin/env python3
"""
Synchronise les VRAIS clients depuis les fichiers WML_XX_2026.xlsx (factures mensuelles).

Règles :
- Un CIP qui apparaît dans WML 2026 = CLIENT (ca2023 = CA cumulé 2026 de la pharmacie)
- Un CIP qui n'apparaît PAS dans WML = PROSPECT (ca2023 forcé à 0, garde son potentielGx)
- Si un client WML n'est pas dans clients-data.js : on l'ajoute

Met à jour : crm/clients-data.js
Backup : crm/clients-data.bak.js (écrasé)
"""

import openpyxl
import re
import shutil
from pathlib import Path
from collections import defaultdict
from glob import glob

CRM_JS_PATH = Path('crm/clients-data.js')


def clean_int_str(v):
    """Convertit valeur Excel en string entière sans '.0' final."""
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def read_wml_clients():
    """Lit tous les WML_*_2026.xlsx et agrège le CA par CIP."""
    files = sorted(glob('WML_*_2026.xlsx'))
    print(f'[sync] Fichiers WML trouvés : {len(files)} → {[Path(f).name for f in files]}')
    clients_ca = defaultdict(float)
    clients_nom = {}
    line_count = 0
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb['Sheet1']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is None or row[11] is None:
                continue
            cip = clean_int_str(row[2])
            if not cip.isdigit():
                continue
            try:
                montant = float(row[11])
            except (ValueError, TypeError):
                continue
            clients_ca[cip] += montant
            nom = row[4]
            if cip not in clients_nom and nom:
                clients_nom[cip] = str(nom).strip()
            line_count += 1
        print(f'[sync]   {Path(f).name} → {ws.max_row - 1} lignes lues')
    print(f'[sync] Total lignes parsées : {line_count}')
    print(f'[sync] Clients distincts : {len(clients_ca)}')
    return clients_ca, clients_nom


def parse_crm_records(content):
    """Parse les records JS de clients-data.js — retourne dict cip → record dict."""
    records = re.findall(r'\{cip:"(\d+)"[^}]*\}', content)
    # On ne peut pas reconstruire parfaitement avec un regex simple,
    # on retourne juste les CIPs présents.
    return set(records)


def replace_field(rec_str, field, new_value):
    """Remplace ou ajoute un champ dans un record JS-like {…}."""
    if isinstance(new_value, str):
        new_value = f'"{new_value}"'
    else:
        new_value = str(new_value)
    pattern = rf'({field}:)[^,}}]*'
    if re.search(pattern, rec_str):
        return re.sub(pattern, rf'\g<1>{new_value}', rec_str, count=1)
    # Pas trouvé : on l'ajoute avant la dernière `}`
    return rec_str[:-1] + f',{field}:{new_value}' + '}'


def update_existing_records(content, clients_ca):
    """Met à jour ca2023 pour chaque record selon présence dans WML."""
    def repl(m):
        rec = m.group(0)
        cip_match = re.search(r'cip:"(\d+)"', rec)
        if not cip_match:
            return rec
        cip = cip_match.group(1)
        if cip in clients_ca:
            ca = int(round(clients_ca[cip]))
            return replace_field(rec, 'ca2023', ca)
        # Pas client : force ca2023 à 0
        return replace_field(rec, 'ca2023', 0)
    return re.sub(r'\{cip:"\d+"[^}]*\}', repl, content)


def lookup_ip_wm_pharma(cip):
    """Cherche une pharma par CIP dans BASE DE DONNÉE IP WM (peu importe le territoire)."""
    if not Path('BASE DE DONNÉE IP WM.xlsx').exists():
        return None
    wb = openpyxl.load_workbook('BASE DE DONNÉE IP WM.xlsx', read_only=True, data_only=True)
    ws = wb['BASE DONNÉE IP']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[11] is None:
            continue
        row_cip = clean_int_str(row[11])
        if row_cip == cip:
            cp_val = row[5] if row[5] else row[4]
            cp = clean_int_str(cp_val) if cp_val else ''
            if cp.isdigit() and len(cp) < 5:
                cp = cp.zfill(5)
            return {
                'nom': str(row[0] or '').strip(),
                'groupement': str(row[1] or '').strip(),
                'adresse': str(row[3] or '').strip(),
                'cp': cp[:5] if cp else '',
                'ville': str(row[8] or '').strip(),
                'email': str(row[9] or '').strip(),
                'tel': clean_int_str(row[10]),
            }
    return None


def build_new_client_record(cip, nom, ca):
    nom_safe = (nom or f'CIP {cip}').replace('\\', '\\\\').replace('"', '\\"')
    # Lookup IP WM pour récupérer adresse/CP/ville/groupement
    extra = lookup_ip_wm_pharma(cip) or {}
    adresse = (extra.get('adresse') or '').replace('\\', '\\\\').replace('"', '\\"')
    cp = extra.get('cp') or ''
    ville = (extra.get('ville') or '').replace('"', '\\"')
    email = (extra.get('email') or '').replace('"', '\\"')
    tel = extra.get('tel') or ''
    groupement = (extra.get('groupement') or '').replace('"', '\\"')
    return (
        f'  {{cip:"{cip}",nom:"{nom_safe}",adresse:"{adresse}",cp:"{cp}",ville:"{ville}",'
        f'email:"{email}",tel:"{tel}",potentielGx:0,ca2023:{int(round(ca))},'
        f'prochaineVisite:null,commentaire:"Client WML 2026 (auto)",'
        f'pelgraz:"",pelmeg:"",ecodage:"{groupement}",gros1:"",gros2:""}},'
    )


def add_missing_clients(content, missing, clients_ca, clients_nom):
    """Insère les nouveaux clients absents du CRM avant le `];` final."""
    if not missing:
        return content
    records = [build_new_client_record(cip, clients_nom.get(cip, ''), clients_ca[cip]) for cip in missing]
    last = content.rfind('];')
    pre = content[:last].rstrip()
    if pre.endswith('}'):
        pre += ','
    return pre + '\n  // ── Clients WML ajoutés auto (absents du CRM) ─────────\n' + '\n'.join(records).rstrip(',') + '\n];'


def main():
    print('[sync] Lecture des fichiers WML...')
    clients_ca, clients_nom = read_wml_clients()

    print('[sync] Top 10 clients WML 2026 (CA cumulé) :')
    for cip, ca in sorted(clients_ca.items(), key=lambda x: -x[1])[:10]:
        print(f'  CIP {cip:8s} · {clients_nom.get(cip, "?"):40s} · {ca:>12.0f} €')

    print('[sync] Lecture crm/clients-data.js...')
    content = CRM_JS_PATH.read_text(encoding='utf-8')
    crm_cips = parse_crm_records(content)
    print(f'[sync] {len(crm_cips)} pharmacies dans le CRM')

    matched = [c for c in clients_ca if c in crm_cips]
    missing = [c for c in clients_ca if c not in crm_cips]
    print(f'[sync] Clients WML matchés dans CRM : {len(matched)}')
    print(f'[sync] Clients WML ABSENTS du CRM (ajouter) : {len(missing)}')
    if missing:
        for cip in missing:
            print(f'    + {cip} · {clients_nom.get(cip, "?")}')

    # Backup
    backup = CRM_JS_PATH.with_suffix('.bak.js')
    shutil.copy(CRM_JS_PATH, backup)
    print(f'[sync] Backup → {backup}')

    # Update ca2023 pour tous les records existants
    new_content = update_existing_records(content, clients_ca)

    # Ajout des clients manquants
    new_content = add_missing_clients(new_content, missing, clients_ca, clients_nom)

    CRM_JS_PATH.write_text(new_content, encoding='utf-8')

    # Vérif
    total = len(re.findall(r'cip:"(\d+)"', new_content))
    actifs = len(re.findall(r'ca2023:(\d+)', new_content))
    actifs_nonzero = sum(1 for m in re.finditer(r'ca2023:(\d+)', new_content) if int(m.group(1)) > 0)
    print(f'[sync] Écrit crm/clients-data.js — {total} pharmacies, {actifs_nonzero} clients (ca2023 > 0)')


if __name__ == '__main__':
    main()
