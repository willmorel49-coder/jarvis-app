# -*- coding: utf-8 -*-
"""Statistiques RÉSEAU par molécule (artmol), pour montrer au pharmacien ce qu'une
molécule génère : rotation moyenne/pharmacie, marge pharmacien (MDL, remboursables),
remise Intégral (PPHT→net) et CA d'achat. Précalculé depuis les ventes WML + le
fichier stock/prix (molécule + PPHT + afmcode).
→ crm/v2/mol-stats-data.js  (window.MOL_STATS = [ {m, n, rota, marge, remise, ca}... ])
Python 3.9.
"""
import re
import json
import glob
import os
from collections import defaultdict
import openpyxl

OUT = 'crm/v2/mol-stats-data.js'

# ── 1. CIP -> molécule / PPHT / remboursable (fichier stock+prix le plus récent) ──
src = sorted(glob.glob('STATS/stock et prix*.xlsx'), key=os.path.getmtime, reverse=True)[0]
ws = openpyxl.load_workbook(src, read_only=True, data_only=True).active
h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
ci, mi, pi, ai = h.index('artcodebarre'), h.index('artmol'), h.index('ppht'), h.index('afmcode')
info = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    c = str(r[ci] or '')
    if not (c.isdigit() and len(c) >= 12):
        continue
    mol = str(r[mi]).strip().upper() if (r[mi] and str(r[mi]) != '#N/A') else ''
    ppht = float(r[pi]) if isinstance(r[pi], (int, float)) else 0.0
    info[c] = {'mol': mol, 'ppht': ppht, 'remb': str(r[ai] or '') == 'REMBSS'}


def mdl(p):                       # barème MDL France (remboursables)
    if p <= 0:
        return 0.0
    if p <= 4.33:
        return 0.18
    if p <= 468:
        return p * 0.039
    return 19.50


# ── 2. Ventes WML : [pharmacyId, mois, comm, cip13, qte, puNet, mntNetHt] ──
t = open('crm/v2/wml-officines-data.js', encoding='utf-8').read()
sales = json.loads(re.search(r'WML_SALES\s*=\s*(\[.*?\]);', t, re.S).group(1))
months = set(s[1] for s in sales if len(s) > 1 and s[1])
NM = len(months) or 5

agg = defaultdict(lambda: {'qte': 0.0, 'ph': set(), 'ca': 0.0, 'tarif': 0.0, 'marge': 0.0})
for s in sales:
    cip = str(s[3]); nfo = info.get(cip)
    if not nfo or not nfo['mol']:
        continue
    qte = s[4] or 0; ca = s[6] or 0; ppht = nfo['ppht']
    a = agg[nfo['mol']]
    a['qte'] += qte; a['ph'].add(str(s[0])); a['ca'] += ca
    if ppht > 0:
        a['tarif'] += ppht * qte
        if nfo['remb']:
            a['marge'] += mdl(ppht) * qte

# ── 3. Ramené à 1 pharmacie / an (moyenne réseau) ──
rows = []
for mol, a in agg.items():
    nph = len(a['ph'])
    if nph < 3:
        continue
    k = 12.0 / NM / nph
    remise = max(0.0, a['tarif'] - a['ca'])     # valeur tarif - payé = remise en €
    rows.append({
        'm': mol,
        'n': nph,                                # nb pharmacies du réseau
        'rota': round(a['qte'] * k),             # boîtes / pharmacie / an
        'marge': round(a['marge'] * k),          # marge pharmacien (MDL) € / pharmacie / an
        'remise': round(remise * k),             # remise Intégral (PPHT→net) € / pharmacie / an
        'ca': round(a['ca'] * k),                # CA achat HT € / pharmacie / an
    })
rows.sort(key=lambda r: -r['marge'])

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('// Stats réseau par molécule (rotation/marge/remise/CA par pharmacie/an) — generate_mol_stats.py\n')
    f.write('window.MOL_STATS = ' + json.dumps(rows, ensure_ascii=False, separators=(',', ':')) + ';\n')

print('OK -> %s : %d molécules (sur %d mois de ventes)' % (OUT, len(rows), NM))
for r in rows[:8]:
    print('  %-28s %4d phies · rota %5d/an · marge %5d€ · remise %5d€ · CA %6d€' % (r['m'][:28], r['n'], r['rota'], r['marge'], r['remise'], r['ca']))
