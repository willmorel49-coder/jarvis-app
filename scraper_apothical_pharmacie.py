#!/usr/bin/env python3
"""
Scraper Apothical — Sous-domaine pharmacie (Parapharmacie)
===========================================================
Cible : https://pharmacie-{slug}.apothical.fr/medicaments-parapharmacie/parapharmacie
Moteur : requests + BeautifulSoup (HTML server-rendered, SANS Selenium)
Sortie : benchmark_apothical_pharmacie_{slug}.xlsx

Structure découverte :
  - Catégories  : /medicaments-parapharmacie/{id}-{slug}
  - Page 2+     : /medicaments-parapharmacie/{id}-{slug}/{n}
  - Produits    : div.product.mb-4 → a.image-wrapper (title, href, data-product-id)
  - Prix affiché: span.catalog-discount-price-discounted | span.price > premier span
  - 24 produits / page
  - Pagination  : ul.pagination → liens href

Usage :
    python3 scraper_apothical_pharmacie.py
    # ou modifier PHARMACY_URL ci-dessous
"""

import time
import re
import unicodedata
import logging
import sys
from pathlib import Path
from typing import Optional, List, Dict, Tuple

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ══════════════════════════════════════════════════════════════════
#  CONFIG — modifier ici pour changer de pharmacie
# ══════════════════════════════════════════════════════════════════
PHARMACY_URL = "https://pharmacie-montlouis-sur-loire.apothical.fr"
# Autre exemple: "https://pharmacie-aixenprovence-centre.apothical.fr"

DELAY       = 1.0   # secondes entre requêtes
MAX_RETRIES = 3
TIMEOUT     = 20

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ══════════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("apothical-pharma")


# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════
def normalize(text: str) -> str:
    """Minuscules, sans accents, alphanum+espaces."""
    if not text:
        return ""
    text = text.lower().strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_price(raw: str) -> Optional[float]:
    """'29,98 €' → 29.98 | None si invalide."""
    if not raw:
        return None
    clean = re.sub(r"[^\d,\.]", "", raw.replace("\xa0", ""))
    clean = clean.replace(",", ".")
    try:
        return float(clean)
    except ValueError:
        return None


def clean_name(title: str) -> str:
    """
    'Acheter Forcapil Anti-chute Comprimés à Montlouis-sur-Loire'
    → 'Forcapil Anti-chute Comprimés'
    """
    name = re.sub(r"^acheter\s+", "", title.strip(), flags=re.IGNORECASE)
    # Retirer " à {ville}" en fin
    name = re.sub(r"\s+à\s+[\w\s\-]+$", "", name, flags=re.IGNORECASE)
    return name.strip()


def get_slug(url: str) -> str:
    """'https://pharmacie-montlouis-sur-loire.apothical.fr' → 'montlouis-sur-loire'"""
    host = re.sub(r"https?://", "", url).split(".")[0]
    return re.sub(r"^pharmacie-", "", host)


def get_page(session: requests.Session, url: str) -> Optional[BeautifulSoup]:
    """Fetch avec retry."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except requests.RequestException as e:
            wait = attempt * 3
            log.warning(f"Tentative {attempt}/{MAX_RETRIES} ({e}) – attente {wait}s")
            time.sleep(wait)
    log.error(f"Échec définitif : {url}")
    return None


# ══════════════════════════════════════════════════════════════════
#  STEP 1 — Récupérer toutes les catégories de parapharmacie
# ══════════════════════════════════════════════════════════════════
def get_categories(session: requests.Session, base: str) -> List[Tuple[str, str]]:
    """
    Retourne [(label, url), ...] pour toutes les catégories + sous-catégories
    trouvées sur la page /medicaments-parapharmacie/parapharmacie.
    """
    url = f"{base}/medicaments-parapharmacie/parapharmacie"
    soup = get_page(session, url)
    if not soup:
        return []

    cats = {}
    # Les catégories sont dans le menu de navigation et dans la grille d'accueil
    for a in soup.find_all("a", href=True):
        href = a["href"]
        # Pattern: /medicaments-parapharmacie/{digits}-{slug}
        m = re.match(r"^/medicaments-parapharmacie/(\d+)-([a-z0-9\-]+)$", href)
        if m:
            full_url = base + href
            label = a.get_text(strip=True) or m.group(2).replace("-", " ").title()
            cats[href] = (label, full_url)

    result = list(cats.values())
    log.info(f"Catégories trouvées : {len(result)}")
    for label, u in result[:5]:
        log.info(f"  {label:40s}  {u}")
    return result


# ══════════════════════════════════════════════════════════════════
#  STEP 2 — Détecter la dernière page d'une catégorie
# ══════════════════════════════════════════════════════════════════
def get_last_page(soup: BeautifulSoup) -> int:
    """Lit ul.pagination et retourne le plus grand numéro de page."""
    pag = soup.find("ul", class_="pagination")
    if not pag:
        return 1
    nums = []
    for a in pag.find_all("a", href=True):
        m = re.search(r"/(\d+)$", a["href"])
        if m:
            nums.append(int(m.group(1)))
    return max(nums) if nums else 1


# ══════════════════════════════════════════════════════════════════
#  STEP 3 — Extraire les produits d'une page
# ══════════════════════════════════════════════════════════════════
def parse_products(soup: BeautifulSoup, cat_label: str, cat_url: str) -> List[Dict]:
    """Extrait tous les produits d'une page listage."""
    products = []
    containers = soup.find_all("div", class_="product")

    for div in containers:
        # ── Lien & nom ──────────────────────────────────────────
        link = div.find("a", class_="image-wrapper")
        if not link:
            continue

        title  = link.get("title", "")
        name   = clean_name(title)
        href   = link.get("href", "")
        url    = PHARMACY_URL + href if href.startswith("/") else href
        prod_id = link.get("data-product-id", "")

        # ── Prix ────────────────────────────────────────────────
        price_raw  = None
        prix_type  = "standard"

        # Prix barré (prix original avant remise)
        orig_el = div.find("span", class_="catalog-discount-price-original")
        prix_original = parse_price(orig_el.get_text(strip=True)) if orig_el else None

        # Prix affiché (remisé ou normal)
        disc_el = div.find("span", class_="catalog-discount-price-discounted")
        if disc_el:
            price_raw = disc_el.get_text(strip=True)
            prix_type = "promo"
        else:
            # Prix normal : premier span dans span.price
            price_span = div.find("span", class_="price")
            if price_span:
                first_span = price_span.find("span")
                if first_span:
                    price_raw = first_span.get_text(strip=True)
                else:
                    price_raw = price_span.get_text(strip=True)

        prix_affiche  = parse_price(price_raw)

        # ── Stock / disponibilité ────────────────────────────────
        is_ordonnance = bool(div.find(class_=lambda c: c and "drug" in c.lower() if c else False))
        disponible    = not bool(div.find(class_=lambda c: c and "out-of-stock" in c.lower() if c else False))

        products.append({
            "id":              prod_id,
            "nom":             name,
            "nom_normalise":   normalize(name),
            "categorie":       cat_label,
            "prix_affiche":    prix_affiche,
            "prix_original":   prix_original,
            "type_prix":       prix_type,
            "disponible":      disponible,
            "url":             url,
            "categorie_url":   cat_url,
        })

    return products


# ══════════════════════════════════════════════════════════════════
#  STEP 4 — Export Excel
# ══════════════════════════════════════════════════════════════════
def export_excel(products: List[Dict], path: Path, slug: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apothical Parapharmacie"

    cols = [
        ("ID",            "id"),
        ("Nom",           "nom"),
        ("Catégorie",     "categorie"),
        ("Prix affiché",  "prix_affiche"),
        ("Prix original", "prix_original"),
        ("Type prix",     "type_prix"),
        ("Disponible",    "disponible"),
        ("Nom normalisé", "nom_normalise"),
        ("URL",           "url"),
        ("Catégorie URL", "categorie_url"),
    ]

    hdr_fill = PatternFill("solid", fgColor="FF6B35")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (label, _) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    for row_i, p in enumerate(products, 2):
        for col_i, (_, key) in enumerate(cols, 1):
            val = p[key]
            if key == "url":
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.hyperlink = val
                cell.font = Font(color="0057FF", underline="single")
            elif key == "disponible":
                ws.cell(row=row_i, column=col_i, value="Oui" if val else "Non")
            elif key in ("prix_affiche", "prix_original") and val is not None:
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.number_format = '#,##0.00 "€"'
            else:
                ws.cell(row=row_i, column=col_i, value=val)

    widths = [10, 45, 30, 14, 14, 12, 10, 50, 80, 60]
    for col_i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col_i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(path)
    sz = path.stat().st_size / 1024
    log.info(f"Export : {path}  ({sz:.0f} KB, {len(products)} produits)")


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    base  = PHARMACY_URL.rstrip("/")
    slug  = get_slug(base)
    out   = Path(__file__).parent / f"benchmark_apothical_pharmacie_{slug}.xlsx"

    log.info("=" * 60)
    log.info(f"  Scraper Apothical Pharmacie — {slug}")
    log.info(f"  Cible : {base}")
    log.info("=" * 60)

    session = requests.Session()

    # ── Catégories ───────────────────────────────────────────────
    categories = get_categories(session, base)
    if not categories:
        log.error("Aucune catégorie trouvée. Vérifier l'URL.")
        return

    all_products: List[Dict] = []
    seen_ids = set()

    # ── Parcours catégories → pages → produits ────────────────────
    for cat_i, (cat_label, cat_url) in enumerate(categories, 1):
        log.info(f"\nCatégorie {cat_i}/{len(categories)} : {cat_label}")

        soup = get_page(session, cat_url)
        if not soup:
            continue
        time.sleep(DELAY)

        last_page = get_last_page(soup)
        log.info(f"  Pages : {last_page}")

        for page_num in range(1, last_page + 1):
            if page_num == 1:
                page_url = cat_url
            else:
                page_url = f"{cat_url}/{page_num}"
                soup = get_page(session, page_url)
                if not soup:
                    continue
                time.sleep(DELAY)

            products = parse_products(soup, cat_label, cat_url)

            new = 0
            for p in products:
                uid = p["id"] or p["nom_normalise"]
                if uid not in seen_ids:
                    seen_ids.add(uid)
                    all_products.append(p)
                    new += 1

            log.info(f"  Page {page_num}/{last_page} : {len(products)} produits ({new} nouveaux) | Total : {len(all_products)}")

    # ── Résumé ───────────────────────────────────────────────────
    log.info("\n" + "=" * 60)
    log.info(f"TOTAL PRODUITS UNIQUES : {len(all_products)}")
    avec_prix = sum(1 for p in all_products if p["prix_affiche"] is not None)
    promos    = sum(1 for p in all_products if p["type_prix"] == "promo")
    log.info(f"Avec prix              : {avec_prix}")
    log.info(f"En promotion           : {promos}")

    if all_products:
        export_excel(all_products, out, slug)
        log.info(f"\n✓ Fichier : {out}")
    else:
        log.error("Aucun produit extrait.")


if __name__ == "__main__":
    main()
