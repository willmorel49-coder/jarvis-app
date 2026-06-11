#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère opso/opso-stats-data.js depuis les fichiers d'achats par officine
déposés dans opso/STATISTIQUES/ (*.xlsx hors WML_*).
Chaque fichier = une officine OPSO (identifiée par son nom de fichier),
rattachée au listing officiel par CIP. Sortie : window.OPSO_STATS_SALES
(lignes produit par officine) + window.OPSO_STATS_META (récap officines).
"""
import json, re, glob, os
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
STATS_DIR = os.path.join(ROOT, "opso", "STATISTIQUES")
LISTING_JS = os.path.join(ROOT, "opso", "opso-listing-2026.js")
OUT = os.path.join(ROOT, "opso", "opso-stats-data.js")


def norm(s):
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


def load_listing():
    txt = open(LISTING_JS, encoding="utf-8").read()
    return json.loads(txt[txt.index("["):txt.rindex("]") + 1])


def match_cip(filename, listing):
    """Rattache un nom de fichier d'officine au CIP du listing."""
    base = re.sub(r"\.xlsx$", "", os.path.basename(filename), flags=re.I)
    base = re.sub(r"opso\s*sant.*$", "", base, flags=re.I).strip()
    nb = norm(base).replace("phie", "pharmacie")
    # nom complet candidat
    cand = nb if nb.startswith("pharmacie") else ("pharmacie " + nb)
    # 1) match exact sur nom normalisé
    for e in listing:
        if norm(e["nom"]) == cand:
            return e
    # 2) tous les tokens significatifs présents → on prend le nom le plus court
    stop = {"phie", "pharmacie", "de", "du", "des", "la", "le", "l", "d", "aux", "au"}
    toks = [t for t in nb.split() if t not in stop]
    best = [e for e in listing if toks and all(t in norm(e["nom"]) for t in toks)]
    best.sort(key=lambda e: len(e["nom"]))
    return best[0] if best else None


def num(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except Exception:
        return 0.0


def main():
    listing = load_listing()
    files = [f for f in glob.glob(os.path.join(STATS_DIR, "*.xlsx"))
             if not os.path.basename(f).upper().startswith("WML")]
    files.sort()

    sales = []
    meta = []
    for f in files:
        e = match_cip(f, listing)
        if not e:
            print("  ⚠️  AUCUN MATCH pour", os.path.basename(f), "— ignoré")
            continue
        cip = str(e["cip"])
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # entêtes -> index par nom
        head = [str(c).strip().upper() if c is not None else "" for c in rows[0]]
        idx = {h: i for i, h in enumerate(head)}

        def g(r, key):
            i = idx.get(key)
            return r[i] if (i is not None and i < len(r)) else None

        nb_ref = 0
        ca_off = 0.0
        for r in rows[1:]:
            desig = g(r, "ARTDESIGNATION")
            code = g(r, "ARTCODE")
            if not desig and not code:
                continue
            qte = num(g(r, "PLVQTE"))
            pubrut = num(g(r, "PLVPUBRUT_MOY"))
            punet = num(g(r, "PLVPUNET_MOY"))
            mnt = num(g(r, "PLVMNTNETHT_MOY"))
            if mnt == 0 and qte and punet:
                mnt = round(qte * punet, 2)
            sales.append({
                "cip": cip,
                "artCode": str(code).strip() if code is not None else "",
                "artDesignation": str(desig).strip() if desig is not None else "",
                "qte": round(qte, 2),
                "puBrut": round(pubrut, 4),
                "puNet": round(punet, 4),
                "mntNetHt": round(mnt, 2),
            })
            nb_ref += 1
            ca_off += mnt
        meta.append({
            "cip": cip, "nom": e["nom"], "ville": e.get("ville", ""),
            "perimetre": e.get("perimetre", ""), "nbRef": nb_ref,
            "ca": round(ca_off, 2),
        })
        print("  OK %-32s -> CIP %-9s %-32s %4d réf · %9.2f € HT"
              % (os.path.basename(f), cip, e["nom"], nb_ref, ca_off))

    payload_sales = json.dumps(sales, ensure_ascii=False)
    payload_meta = json.dumps(meta, ensure_ascii=False)
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write("// Achats par officine OPSO Santé — déposés dans opso/STATISTIQUES/\n")
        fh.write("// Généré par generate_opso_stats.py — ne pas éditer à la main\n")
        fh.write("// %d officines · %d lignes produit\n" % (len(meta), len(sales)))
        fh.write("window.OPSO_STATS_SALES = " + payload_sales + ";\n")
        fh.write("window.OPSO_STATS_META = " + payload_meta + ";\n")
    print("\n✅ %s — %d officines, %d lignes" % (os.path.relpath(OUT, ROOT), len(meta), len(sales)))


if __name__ == "__main__":
    main()
