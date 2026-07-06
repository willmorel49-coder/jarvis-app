#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_zone.py — Copilote / feed #3 : POTENTIEL DE ZONE (axe OFFICINE).

Pour chaque officine du réseau (fichiers STATS/*_geolocalisation_*.xlsx : TIRCODE +
code postal + ville), retrouve sa commune + population + département via l'API publique
geo.api.gouv.fr (par code postal), et écrit crm/v2/zone-data.js (clé = id officine).

v1 : commune, code INSEE, département, population de la commune (indicatif).
Source : geo.api.gouv.fr (Etalab) — gratuit, sans clé. Python 3.9. openpyxl.
"""
import os
import re
import sys
import glob
import json
import time
import unicodedata
import urllib.request

import openpyxl

ROOT = os.path.dirname(__file__)
OUT = os.path.join(ROOT, "crm", "v2", "zone-data.js")
API = "https://geo.api.gouv.fr/communes?codePostal=%s&fields=nom,population,codeDepartement,code&format=json"


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", s.lower())


def load_officines():
    """Combine les geoloc par commercial → {tircode: {cp, ville}} (le plus récent gagne)."""
    files = sorted(glob.glob(os.path.join(ROOT, "STATS", "*_geolocalisation_*.xlsx")))
    officines = {}
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True)
        sh = wb[wb.sheetnames[0]]
        it = sh.iter_rows(values_only=True)
        hdr = next(it, None)
        if not hdr:
            continue
        idx = {str(h).strip(): i for i, h in enumerate(hdr)}
        ic, icp, iv = idx.get("TIRCODE"), idx.get("ADRCODEPOSTAL"), idx.get("ADRVILLE")
        if ic is None or icp is None:
            continue
        for r in it:
            code = str(r[ic]).strip() if r[ic] is not None else ""
            cp = re.sub(r"\D", "", str(r[icp] or ""))
            if not code or not cp:
                continue
            cp = cp.zfill(5)
            ville = str(r[iv] or "").strip() if iv is not None else ""
            officines[code] = {"cp": cp, "ville": ville}
    return officines, len(files)


_cache = {}
def commune_by_cp(cp, ville):
    if cp in _cache:
        arr = _cache[cp]
    else:
        arr = None
        for attempt in range(3):
            try:
                req = urllib.request.Request(API % cp, headers={"User-Agent": "JARVIS/1.0"})
                with urllib.request.urlopen(req, timeout=20) as rr:
                    arr = json.loads(rr.read().decode("utf-8"))
                break
            except Exception:
                time.sleep(0.6 * (attempt + 1))
        _cache[cp] = arr or []
        arr = _cache[cp]
        time.sleep(0.04)
    if not arr:
        return None
    # une ville par CP possible : on prend celle qui matche le nom, sinon la plus peuplée
    nv = norm(ville)
    best = None
    for c in arr:
        if nv and norm(c.get("nom")) == nv:
            best = c; break
    if not best:
        best = max(arr, key=lambda c: c.get("population") or 0)
    return {"c": best.get("nom", ""), "cc": best.get("code", ""),
            "dep": best.get("codeDepartement", ""), "pop": int(best.get("population") or 0)}


def main():
    officines, nf = load_officines()
    print("Officines (geoloc, %d fichiers) : %d" % (nf, len(officines)))
    data = {}
    for i, (code, o) in enumerate(officines.items()):
        z = commune_by_cp(o["cp"], o["ville"])
        if z and z["pop"] > 0:
            data[code] = z
        if (i + 1) % 150 == 0:
            print("  … %d/%d (%d résolues, %d CP en cache)" % (i + 1, len(officines), len(data), len(_cache)))

    if not data:
        sys.exit("Aucune zone résolue.")

    pops = sorted(z["pop"] for z in data.values())
    deps = {}
    for z in data.values():
        deps[z["dep"]] = deps.get(z["dep"], 0) + 1
    top = sorted(deps.items(), key=lambda x: -x[1])[:5]
    print("Zones : %d · pop médiane ~%d · top dép : %s" % (len(data), pops[len(pops) // 2], top))

    items = ",".join(
        '"%s":{c:"%s",cc:"%s",dep:"%s",pop:%d}'
        % (code, z["c"].replace('"', "'"), z["cc"], z["dep"], z["pop"])
        for code, z in sorted(data.items())
    )
    js = (
        "// Copilote — potentiel de zone (axe officine). Commune + population par officine.\n"
        "// Source: geo.api.gouv.fr (Etalab), par code postal (STATS/*_geolocalisation). Indicatif.\n"
        "// Généré par generate_zone.py — ne pas éditer.\n"
        "window.ZONE={meta:{n:%d,source:\"geo.api.gouv.fr\"},data:{%s}};\n"
    ) % (len(data), items)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(js)
    print("Écrit %s (%d Ko, %d officines)" % (OUT, os.path.getsize(OUT) // 1024, len(data)))


if __name__ == "__main__":
    main()
