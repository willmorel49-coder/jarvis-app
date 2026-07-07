# -*- coding: utf-8 -*-
"""Marketing · source « Ventes NR réelles » (hors médicament remboursable).

Croise les VRAIES ventes Intégral hors-remboursable (fichiers *_NETTOYE_agregation.xlsx,
sell-out CA + quantités) avec le prix de référence (PPHT) tiré des fichiers de stock
par établissement, et produit crm/v2/mkt-nr-data.js = window.MKT_NR :
un catalogue de catégories NR classées par ventes réelles, prêt pour la machine de
catalogues du CRM (sélecteur d'établissement = prix + stock à jour via ETAB_PRICES).

Périmètre NR = tout sauf REMBSS (remboursable sécu) : MED010 (OTC/conseil),
DM / DM_20 (dispositifs médicaux), PARA (parapharmacie), MED021 (autres NR).

100% local, Python 3.9. Sources gratuites (exports ERP Intégral).
"""
import openpyxl
import glob
import os
import re
import json

AGG_DIR = 'STATS'                                        # *_NETTOYE_agregation.xlsx (sell-out)
STOCK_DIR = '/Users/williammorel/JARVIS/PRIX ET STOCKS ETABLISSEMENTS'   # *_extrait.xlsx (prix/stock)
OUT = 'crm/v2/mkt-nr-data.js'

# familles NR -> (clé catégorie, libellé vendeur, ordre d'affichage)
FAM = {
    'MED010': ('otc',  'Médicaments conseil (OTC non remboursables)', 1),
    'DM':     ('dm',   'Dispositifs médicaux', 2),
    'DM_20':  ('dm',   'Dispositifs médicaux', 2),
    'PARA':   ('para', 'Parapharmacie', 3),
    'MED021': ('autre', 'Autres non remboursables', 4),
}
TOP_HEADLINE = 80      # taille de la catégorie « Top ventes hors-remboursable »
TOP_PER_FAM = 120      # taille par catégorie famille (marge pour « plus de NR » côté catalogue)

# Produits sortis du périmètre NR (devenus remboursables) — à exclure du catalogue.
# Mounjaro & Wegovy : plus NR depuis le 15/06/2026.
EXCLUDE_NAMES = ('MOUNJARO', 'WEGOVY')


def ean_of(v):
    if v is None:
        return ''
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return re.sub(r'[^0-9]', '', str(v))


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def cap(s):
    s = str(s or '').strip()
    return s[:1].upper() + s[1:].lower() if s else ''


def build_price_map():
    """EAN -> meilleur PPHT (>0) sur les 7 établissements de stock."""
    pmap = {}
    for fn in sorted(glob.glob(os.path.join(STOCK_DIR, '*_extrait.xlsx'))):
        wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        hdr = [str(c) if c is not None else '' for c in next(it)]
        ix = {h: i for i, h in enumerate(hdr)}
        ci, pi = ix.get('ARTCODEBARRE'), ix.get('PPHT')
        for r in it:
            ean = ean_of(r[ci]) if ci is not None else ''
            if not ean:
                continue
            p = round(num(r[pi]), 2) if pi is not None else 0.0
            if p > 0 and (ean not in pmap or p < pmap[ean]):
                pmap[ean] = p
        wb.close()
    return pmap


def main():
    price = build_price_map()

    prod = {}   # ean -> {d, lab, fam, ca, vol}
    etabs = set()
    for fn in sorted(glob.glob(os.path.join(AGG_DIR, '*_NETTOYE_agregation.xlsx'))):
        code = re.match(r'([A-Za-z]{2,4})', os.path.basename(fn))
        etabs.add(code.group(1).upper() if code else '?')
        wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c) if c is not None else '' for c in next(it)]
        ix = {h: i for i, h in enumerate(hdr)}
        for r in it:
            afm = str(r[ix['AFMCODE']] or '').strip()
            if afm == 'REMBSS' or afm not in FAM:      # on ne garde que le NR connu
                continue
            ean = ean_of(r[ix['ARTCODEBARRE']])
            if not ean or len(ean) < 8:
                continue
            nom_u = str(r[ix['PLVDESIGNATION']] or '').upper()
            if any(x in nom_u for x in EXCLUDE_NAMES):   # sortis du NR (remboursables)
                continue
            p = prod.get(ean)
            if p is None:
                p = prod[ean] = {'d': '', 'lab': '', 'fam': afm, 'ca': 0.0, 'vol': 0}
            p['ca'] += num(r[ix['CA_NET_HT']])
            p['vol'] += int(num(r[ix['QTE_TOTALE']]))
            nm = r[ix['PLVDESIGNATION']]
            if nm and not p['d']:
                p['d'] = str(nm).strip()
            lb = r[ix['ARTCOLLECTION']]
            if lb and not p['lab']:
                p['lab'] = cap(lb)
        wb.close()

    def row(ean, p):
        # prix de référence « Tous » : PPHT du stock si dispo, sinon prix net réel
        # moyen (CA vendu ÷ quantité) — toujours un prix cohérent à afficher.
        pr = price.get(ean, 0.0)
        if pr <= 0 and p['vol'] > 0:
            pr = round(p['ca'] / p['vol'], 2)
        return {
            'd': p['d'], 'cip': ean, 'p': pr,
            'vol': p['vol'], 'ca': int(round(p['ca'])),
            'lab': p['lab'],
        }

    items = [(ean, p) for ean, p in prod.items() if p['vol'] > 0]

    # Catégorie phare : top ventes toutes familles NR confondues
    headline = sorted(items, key=lambda x: -x[1]['vol'])[:TOP_HEADLINE]
    cats = [{'cat': 'Top ventes hors-remboursable', 'rows': [row(e, p) for e, p in headline]}]

    # Catégories par famille (ordre défini par FAM)
    buckets = {}
    for ean, p in items:
        key, label, order = FAM[p['fam']]
        buckets.setdefault(key, {'label': label, 'order': order, 'rows': []})
        buckets[key]['rows'].append((ean, p))
    for key in sorted(buckets, key=lambda k: buckets[k]['order']):
        b = buckets[key]
        rows = sorted(b['rows'], key=lambda x: -x[1]['vol'])[:TOP_PER_FAM]
        cats.append({'cat': b['label'], 'rows': [row(e, p) for e, p in rows]})

    total_ca = int(round(sum(p['ca'] for _, p in items)))
    data = {
        'meta': {
            'source': 'Ventes Intégral (sell-out) · ' + '/'.join(sorted(etabs)),
            'etabs': sorted(etabs),
            'nProduits': len(items),
            'caTotal': total_ca,
        },
        'total': len(etabs),
        'cats': cats,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as fh:
        fh.write('// Marketing — ventes NR réelles Intégral (hors remboursable) — generate_mkt_nr.py\n')
        fh.write('// Sell-out (*_NETTOYE_agregation) x prix PPHT (stocks établissements). Ne pas éditer.\n')
        fh.write('window.MKT_NR = ' + json.dumps(data, ensure_ascii=False, separators=(',', ':')) + ';\n')

    print('OK ->', OUT)
    print('  établissements sell-out :', ', '.join(sorted(etabs)))
    print('  produits NR vendus      :', len(items))
    print('  CA NR total             : %s €' % format(total_ca, ',d').replace(',', ' '))
    for c in cats:
        print('   - %-45s %3d produits' % (c['cat'], len(c['rows'])))
    print('  taille fichier : %.0f Ko' % (os.path.getsize(OUT) / 1024))


if __name__ == '__main__':
    main()
