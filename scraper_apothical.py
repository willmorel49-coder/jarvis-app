#!/usr/bin/env python3
"""
Scraper Apothical — Catégorie Médicaments
==========================================
Site : https://www.apothical.fr/medicaments-parapharmacie/produits/page/N/categorie/2-medicaments
Moteur : requests + BeautifulSoup (HTML server-rendered, pas de Selenium nécessaire)
Pages  : 287 pages × 36 produits ≈ 10 332 produits
Prix   : non publiés sur les pages listing (le site est un réseau click-and-collect,
         les prix sont par pharmacie). Le script extrait les prix quand ils sont présents,
         sinon marque "N/D".

Sortie : benchmark_apothical_medicaments.xlsx
"""

import time
import re
import unicodedata
import logging
import sys
from pathlib import Path
from typing import Optional, List, Dict

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ──────────────────────────────────────────────────────────────────
BASE_URL    = "https://www.apothical.fr"
CAT_URL     = "/medicaments-parapharmacie/produits/page/{page}/categorie/2-medicaments"
OUTPUT_FILE = Path(__file__).parent / "benchmark_apothical_medicaments.xlsx"
DELAY       = 1.2          # secondes entre chaque requête
MAX_RETRIES = 3
TIMEOUT     = 20

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.apothical.fr/",
}

# ── LOGGING ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("apothical")


# ── HELPERS ─────────────────────────────────────────────────────────────────
def normalize(text: str) -> str:
    """Minuscules, sans accents, sans caractères spéciaux."""
    if not text:
        return ""
    text = text.lower().strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_price(raw: str) -> Optional[float]:
    """Convertit '12,50 €' ou '12.50' en float, None si impossible."""
    if not raw:
        return None
    raw = raw.replace("\xa0", "").replace(" ", "").replace("€", "")
    raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def get_page(session: requests.Session, url: str) -> Optional[BeautifulSoup]:
    """Fetch + parse avec retry exponentiel."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")
        except requests.RequestException as e:
            wait = attempt * 3
            log.warning(f"Tentative {attempt}/{MAX_RETRIES} échouée ({e}). Attente {wait}s…")
            time.sleep(wait)
    log.error(f"Impossible de charger : {url}")
    return None


# ── DÉTECTION DERNIÈRE PAGE ─────────────────────────────────────────────────
def get_last_page(session: requests.Session) -> int:
    """Lit la pagination de la page 1 et retourne le numéro de la dernière page."""
    url = BASE_URL + CAT_URL.format(page=1)
    soup = get_page(session, url)
    if not soup:
        return 1

    max_page = 1
    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = re.search(r"/produits/page/(\d+)/categorie", href)
        if m:
            max_page = max(max_page, int(m.group(1)))

    log.info(f"Dernière page détectée : {max_page}")
    return max_page


# ── EXTRACTION PRODUITS SUR UNE PAGE ────────────────────────────────────────
def parse_products(soup: BeautifulSoup) -> List[Dict]:
    """Extrait tous les produits d'une page listage."""
    products = []
    articles = soup.find_all("article", class_="product-hover")

    for article in articles:
        # ── Nom ──
        h3 = article.find("h3", class_="product-title")
        if not h3:
            continue

        # Le nom principal est dans le <a> direct, on enlève les sous-spans (DCI, etc.)
        a_tag = h3.find("a")
        if not a_tag:
            continue

        # Texte brut du lien (inclut DCI dans <span>)
        name_full = a_tag.get_text(separator=" ", strip=True)
        name_full = re.sub(r"\s+", " ", name_full).strip()

        # Nom court = texte direct du <a> avant le premier <span>
        name_short = ""
        for node in a_tag.children:
            if hasattr(node, "name"):
                break
            name_short += str(node)
        name_short = name_short.strip().rstrip(",").strip()
        if not name_short:
            name_short = name_full

        # ── URL ──
        href = a_tag.get("href", "")
        url = BASE_URL + href if href.startswith("/") else href

        # ── ID produit (extrait de l'URL) ──
        id_match = re.search(r"/(\d+)-", href)
        product_id = id_match.group(1) if id_match else ""

        # ── Sous-titre (contenance/format) ──
        h4_list = article.find_all("h4", class_="product-subtitle")
        subtitle = " | ".join(h.get_text(strip=True) for h in h4_list if h.get_text(strip=True))

        # ── Prix (non disponible sans sélection de pharmacie, mais on tente) ──
        price_raw = None
        for price_cls in ["product-price", "price", "prix"]:
            el = article.find(class_=lambda c: c and price_cls in c.lower() if c else False)
            if el:
                price_raw = el.get_text(strip=True)
                break
        # Fallback : chercher '€' dans tout l'article
        if not price_raw:
            text = article.get_text(" ", strip=True)
            m = re.search(r"(\d+[,\.]\d{1,2})\s*€", text)
            if m:
                price_raw = m.group(0)

        prix_numerique = parse_price(price_raw)

        # ── Badge (médicament, parapharmacie…) ──
        ribbon = article.find(class_=lambda c: c and "product-ribbon" in c if c else False)
        badge = ribbon.get_text(strip=True) if ribbon else ""

        products.append({
            "id":              product_id,
            "nom":             name_short,
            "nom_complet":     name_full,
            "sous_titre":      subtitle,
            "badge":           badge,
            "prix_affiche":    price_raw or "N/D",
            "prix_numerique":  prix_numerique,
            "nom_normalise":   normalize(name_short),
            "url":             url,
        })

    return products


# ── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def export_excel(products: List[Dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apothical Médicaments"

    headers = [
        "ID",
        "Nom",
        "Nom complet",
        "Sous-titre",
        "Badge",
        "Prix affiché",
        "Prix numérique",
        "Nom normalisé",
        "URL",
    ]

    # Style en-tête
    hdr_fill = PatternFill("solid", fgColor="0057FF")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Données
    for row_idx, p in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=p["id"])
        ws.cell(row=row_idx, column=2, value=p["nom"])
        ws.cell(row=row_idx, column=3, value=p["nom_complet"])
        ws.cell(row=row_idx, column=4, value=p["sous_titre"])
        ws.cell(row=row_idx, column=5, value=p["badge"])
        ws.cell(row=row_idx, column=6, value=p["prix_affiche"])
        ws.cell(row=row_idx, column=7, value=p["prix_numerique"])
        ws.cell(row=row_idx, column=8, value=p["nom_normalise"])
        url_cell = ws.cell(row=row_idx, column=9, value=p["url"])
        url_cell.hyperlink = p["url"]
        url_cell.font = Font(color="0057FF", underline="single")

    # Largeurs colonnes
    widths = [8, 35, 50, 30, 14, 14, 16, 45, 80]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(path)
    log.info(f"Export Excel : {path} ({len(products)} produits)")


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    session = requests.Session()

    log.info("=== Scraper Apothical — Catégorie Médicaments ===")
    last_page = get_last_page(session)

    all_products: List[Dict] = []
    empty_pages = 0

    for page_num in range(1, last_page + 1):
        url = BASE_URL + CAT_URL.format(page=page_num)
        log.info(f"Page {page_num:>4}/{last_page}  →  {url}")

        soup = get_page(session, url)
        if not soup:
            empty_pages += 1
            if empty_pages >= 5:
                log.error("5 pages consécutives en erreur — arrêt.")
                break
            continue

        products = parse_products(soup)

        if not products:
            log.warning(f"Page {page_num} : 0 produits extraits (fin de catalogue ?).")
            empty_pages += 1
            if empty_pages >= 3:
                log.info("3 pages vides consécutives — fin du catalogue détectée.")
                break
        else:
            empty_pages = 0
            all_products.extend(products)
            log.info(f"  → {len(products)} produits extraits  |  Total : {len(all_products)}")

        time.sleep(DELAY)

    log.info(f"\n{'='*50}")
    log.info(f"TOTAL PRODUITS SCRAPPÉS : {len(all_products)}")
    avec_prix = sum(1 for p in all_products if p["prix_numerique"] is not None)
    log.info(f"Avec prix numériques    : {avec_prix}")
    log.info(f"Sans prix (N/D)         : {len(all_products) - avec_prix}")
    log.info(
        "Note : les prix Apothical sont par pharmacie (réseau click-and-collect).\n"
        "       Pour obtenir des prix réels, sélectionner une pharmacie sur le site\n"
        "       et utiliser Selenium avec un cookie de session pharmacie."
    )

    if all_products:
        export_excel(all_products, OUTPUT_FILE)
        log.info(f"\n✓ Fichier généré : {OUTPUT_FILE}")
    else:
        log.error("Aucun produit extrait — vérifier la connectivité ou la structure du site.")


if __name__ == "__main__":
    main()
