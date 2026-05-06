#!/usr/bin/env python3
"""
Generate benchmark-data.js for Intégral Pharma CRM
Sources:
  - WML_01/02/03_2026.xlsx  → name → CIP13 bridge
  - TOP rotations France 2026.xlsx → IP top sellers
  - ventes france stats ameli.xlsx → national Ameli data
Output: crm/benchmark-data.js
"""

import warnings
warnings.filterwarnings('ignore')

import openpyxl
from datetime import datetime
import os
import sys
from pathlib import Path

BASE_DIR = str(Path(__file__).parent)
OUT_FILE = os.path.join(BASE_DIR, 'crm', 'benchmark-data.js')
N_PHARMACIES = 19000

# ── 1. BUILD name→CIP13 lookup from WML files ──────────────────────────────
print("Loading WML files...")
name_to_cip = {}
wml_files = ['WML_01_2026.xlsx', 'WML_02_2026.xlsx', 'WML_03_2026.xlsx']

for wml_file in wml_files:
    path = os.path.join(BASE_DIR, wml_file)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

        # Find column indices
        try:
            desig_idx = headers.index('PLVDESIGNATION')
        except ValueError:
            print(f"  WARNING: PLVDESIGNATION not found in {wml_file}, headers: {headers[:15]}")
            wb.close()
            continue

        try:
            cip_idx = headers.index('ARTCODEBARRE')
        except ValueError:
            print(f"  WARNING: ARTCODEBARRE not found in {wml_file}")
            wb.close()
            continue

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            desig = row[desig_idx]
            cip_raw = row[cip_idx]
            if desig is None or cip_raw is None:
                continue
            try:
                name = str(desig).strip().upper()
                cip = str(int(float(str(cip_raw).replace(',', '.'))))
                if name and cip and len(cip) >= 10:
                    name_to_cip[name] = cip
                    count += 1
            except (ValueError, TypeError):
                pass

        wb.close()
        print(f"  {wml_file}: {count} rows processed, {len(name_to_cip)} unique names so far")
    except Exception as e:
        print(f"  ERROR loading {wml_file}: {e}")

print(f"Total name→CIP13 mappings: {len(name_to_cip)}")

# ── 2. LOAD AMELI national totals per CIP13 ────────────────────────────────
print("\nLoading Ameli data...")
ameli = {}  # cip13 -> {'nom': str, 'boites_jan26': int, 'boites_total': int}

AMELI_FILE = os.path.join(BASE_DIR, 'ventes france stats ameli.xlsx')
if not os.path.exists(AMELI_FILE):
    print(f'ERREUR : fichier Ameli introuvable → {AMELI_FILE}')
    sys.exit(1)
wb_ameli = openpyxl.load_workbook(AMELI_FILE, read_only=True, data_only=True)

for shname in wb_ameli.sheetnames:
    ws = wb_ameli[shname]
    print(f"  Processing sheet: {shname}")

    # Find the header row (where first cell = 'CIP13')
    header = None
    header_row_num = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row and row[0] == 'CIP13':
            header = list(row)
            header_row_num = i
            break

    if header is None:
        print(f"    WARNING: could not find CIP13 header row in {shname}")
        continue

    # Find "Nombre de boites remboursées" columns
    boites_cols = []
    jan26_col = None
    for col_idx, col_name in enumerate(header):
        if col_name and 'Nombre de boites remboursées' in str(col_name):
            boites_cols.append(col_idx)
            if '2026-01' in str(col_name):
                jan26_col = col_idx

    print(f"    Header row: {header_row_num}, boites cols: {len(boites_cols)}, jan26 col: {jan26_col}")

    # Parse data rows
    cip_col = 0
    nom_col = 1
    row_count = 0

    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        cip_raw = row[cip_col] if len(row) > cip_col else None
        if cip_raw is None:
            continue

        try:
            cip = str(int(float(str(cip_raw).replace(' ', '').replace(',', '.'))))
        except (ValueError, TypeError):
            continue

        nom = str(row[nom_col]).strip() if len(row) > nom_col and row[nom_col] else ''

        # Sum boites across all months
        total = 0
        for ci in boites_cols:
            val = row[ci] if len(row) > ci else None
            if val is not None:
                try:
                    total += int(float(str(val)))
                except (ValueError, TypeError):
                    pass

        # Jan 2026 specifically
        jan26 = 0
        if jan26_col is not None and len(row) > jan26_col and row[jan26_col] is not None:
            try:
                jan26 = int(float(str(row[jan26_col])))
            except (ValueError, TypeError):
                pass

        # Add (sum 100% and non-100%)
        if cip in ameli:
            ameli[cip]['boites_total'] += total
            ameli[cip]['boites_jan26'] += jan26
            if not ameli[cip]['nom'] and nom:
                ameli[cip]['nom'] = nom
        else:
            ameli[cip] = {'nom': nom, 'boites_jan26': jan26, 'boites_total': total}

        row_count += 1

    print(f"    Processed {row_count} rows")

wb_ameli.close()
print(f"Total CIP13 in Ameli: {len(ameli)}")

# ── 3. LOAD TOP ROTATIONS IP DATA ──────────────────────────────────────────
print("\nLoading TOP rotations data...")

TOP_FILE = os.path.join(BASE_DIR, 'TOP rotations France 2026.xlsx')
if not os.path.exists(TOP_FILE):
    print(f'ERREUR : fichier TOP rotations introuvable → {TOP_FILE}')
    sys.exit(1)
sheets_cats = {
    'TOP Froid': 'froid',
    'TOP Med010': 'nr',
    'TOP inf 4,8': 'pp',
    'TOP median': 'mi',
    'TOP sup 480': 'ch',
}

records = []
wb_top = openpyxl.load_workbook(TOP_FILE, read_only=True, data_only=True)

for shname, categorie in sheets_cats.items():
    if shname not in wb_top.sheetnames:
        print(f"  WARNING: sheet '{shname}' not found")
        continue

    ws = wb_top[shname]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row 1
    count = 0

    for row in rows:
        if not row or row[0] is None:
            continue

        designation = str(row[0]).strip()
        if not designation or designation == 'None':
            continue

        # Columns: 0=designation, 1=CA, 2=qty, 3=rank_qty, 4=rank_ca
        try:
            ip_ca = float(row[1]) if row[1] is not None else 0.0
        except (ValueError, TypeError):
            ip_ca = 0.0

        try:
            ip_qty = int(float(row[2])) if row[2] is not None else 0
        except (ValueError, TypeError):
            ip_qty = 0

        try:
            ip_rank_qty = int(float(row[3])) if row[3] is not None else 0
        except (ValueError, TypeError):
            ip_rank_qty = 0

        try:
            ip_rank_ca = int(float(row[4])) if row[4] is not None else 0
        except (ValueError, TypeError):
            ip_rank_ca = 0

        # CIP13 lookup
        nom_upper = designation.strip().upper()
        cip13 = name_to_cip.get(nom_upper, None)

        # Ameli data
        has_ameli = False
        ameli_jan26 = 0
        rot_pharma_jan26 = 0.0
        ameli_total = 0

        if cip13 and cip13 in ameli:
            has_ameli = True
            ameli_jan26 = ameli[cip13]['boites_jan26']
            ameli_total = ameli[cip13]['boites_total']
            rot_pharma_jan26 = ameli_jan26 / N_PHARMACIES

        records.append({
            'designation': designation,
            'categorie': categorie,
            'ip_qty': ip_qty,
            'ip_ca': ip_ca,
            'ip_rank_qty': ip_rank_qty,
            'ip_rank_ca': ip_rank_ca,
            'cip13': cip13 or '',
            'ameli_jan26': ameli_jan26,
            'rot_pharma_jan26': rot_pharma_jan26,
            'ameli_total': ameli_total,
            'has_ameli': has_ameli,
        })
        count += 1

    print(f"  {shname}: {count} products loaded")

wb_top.close()

# Sort by ip_qty descending
records.sort(key=lambda r: r['ip_qty'], reverse=True)

matched = sum(1 for r in records if r['has_ameli'])
print(f"\nTotal products: {len(records)}")
print(f"Matched with Ameli: {matched} / {len(records)} ({100*matched//len(records) if records else 0}%)")

# ── 4. GENERATE benchmark-data.js ──────────────────────────────────────────
print(f"\nGenerating {OUT_FILE}...")

def js_str(s):
    """Escape a string for JS double-quoted string."""
    return s.replace('\\', '\\\\').replace('"', '\\"')

today = datetime.now().strftime('%Y-%m-%d')

lines = []
lines.append('// Benchmark Marché — Intégral Pharma vs Ameli France')
lines.append(f'// Généré le {today}')
lines.append('// Sources: TOP rotations France 2026.xlsx + ventes france stats ameli.xlsx')
lines.append('// Base: 19 000 pharmacies France')
lines.append('const BENCHMARK = [')
lines.append('  // { designation, categorie, ip_qty, ip_ca, ip_rank_qty, ip_rank_ca, cip13, ameli_jan26, rot_pharma_jan26, ameli_total, has_ameli }')

for r in records:
    rot = f'{r["rot_pharma_jan26"]:.2f}'
    has = 'true' if r['has_ameli'] else 'false'
    line = (
        f'  {{ designation: "{js_str(r["designation"])}", '
        f'categorie: "{r["categorie"]}", '
        f'ip_qty: {r["ip_qty"]}, '
        f'ip_ca: {r["ip_ca"]:.2f}, '
        f'ip_rank_qty: {r["ip_rank_qty"]}, '
        f'ip_rank_ca: {r["ip_rank_ca"]}, '
        f'cip13: "{r["cip13"]}", '
        f'ameli_jan26: {r["ameli_jan26"]}, '
        f'rot_pharma_jan26: {rot}, '
        f'ameli_total: {r["ameli_total"]}, '
        f'has_ameli: {has} }},'
    )
    lines.append(line)

lines.append('];')

with open(OUT_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines) + '\n')

print(f"Written {len(records)} records to {OUT_FILE}")

# ── 5. PRINT STATS ──────────────────────────────────────────────────────────
print("\n" + "="*60)
print("STATS")
print("="*60)
print(f"Total products: {len(records)}")
print(f"Matched with Ameli: {matched} / {len(records)}")
print(f"\nTop 5 by IP quantity (with Ameli rotation):")
top5 = [r for r in records[:5]]
for r in top5:
    rot_str = f"{r['rot_pharma_jan26']:.1f}" if r['has_ameli'] else 'N/A'
    print(f"  #{r['ip_rank_qty']:3d} {r['designation'][:45]:<45} "
          f"qty={r['ip_qty']:>10,}  rot/pharma/mois={rot_str}  cat={r['categorie']}")
print("="*60)
