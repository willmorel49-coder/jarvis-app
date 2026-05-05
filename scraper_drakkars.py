#!/usr/bin/env python3
"""
Scraper Pharmacie des Drakkars — Parapharmacie
===============================================
Cible : https://www.pharmaciedesdrakkars.com/parapharmacie/
Moteur : requests + BeautifulSoup (HTML server-rendered)
Sortie : benchmark_drakkars.xlsx

Structure :
  - Catégories : sitemap-categories.xml → URLs /parapharmacie/*/
  - Pagination : ?page=N (32 produits/page)
  - Produits   : div#container-item → <a href title>
  - Prix       : JSON-LD sur page produit (offer.price) — fiable et précis
  - EAN/SKU    : JSON-LD (sku/gtin)

Usage :
    python3 scraper_drakkars.py
"""

import time
import re
import json
import unicodedata
import logging
import sys
from pathlib import Path
from typing import Optional, List, Dict

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ══════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════
PHARMACY_URL = "https://www.pharmaciedesdrakkars.com"

DELAY       = 0.8
MAX_RETRIES = 3
TIMEOUT     = 20

# Mode rapide : ne récupère pas les pages produit (prix depuis listing)
# Mode précis : visite chaque page produit pour le prix JSON-LD
FETCH_PRODUCT_PAGES = True   # False = rapide, True = précis

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ══════════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("drakkars")


# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════
def normalize(text: str) -> str:
    if not text:
        return ""
    text = text.lower().strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_price(raw: str) -> Optional[float]:
    if not raw:
        return None
    clean = re.sub(r"[^\d,\.]", "", raw.replace("\xa0", ""))
    clean = clean.replace(",", ".")
    try:
        return float(clean)
    except ValueError:
        return None


def get_page(session: requests.Session, url: str) -> Optional[BeautifulSoup]:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except requests.RequestException as e:
            wait = attempt * 3
            log.warning(f"Tentative {attempt}/{MAX_RETRIES} ({e}) — attente {wait}s")
            time.sleep(wait)
    log.error(f"Échec : {url}")
    return None


# ══════════════════════════════════════════════════════════════════
#  STEP 1 — Récupérer catégories parapharmacie via sitemap
# ══════════════════════════════════════════════════════════════════
def get_categories(session: requests.Session) -> List[str]:
    """Retourne toutes les URLs de catégories parapharmacie depuis le sitemap."""
    url = f"{PHARMACY_URL}/sitemap-categories.xml"
    log.info(f"Chargement sitemap catégories : {url}")
    try:
        r = session.get(url, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
    except requests.RequestException as e:
        log.error(f"Sitemap inaccessible : {e}")
        return []

    cats = re.findall(r'<loc>(https://[^<]+parapharmacie[^<]+)</loc>', r.text)
    # Dédoublonner, exclure la page racine /parapharmacie/ elle-même (pas de produits directs)
    seen = set()
    result = []
    for c in cats:
        c = c.strip()
        if c not in seen:
            seen.add(c)
            result.append(c)

    # Ajouter /parapharmacie/ elle-même si absente
    root = f"{PHARMACY_URL}/parapharmacie/"
    if root not in seen:
        result.insert(0, root)

    log.info(f"Catégories parapharmacie trouvées : {len(result)}")
    return result


# ══════════════════════════════════════════════════════════════════
#  STEP 2 — Extraire produits d'une page listing
# ══════════════════════════════════════════════════════════════════
def extract_price_from_card(card: BeautifulSoup) -> Optional[str]:
    """
    Extrait le prix courant d'une carte produit du listing.
    Cherche des spans/divs sans classe line-through contenant N,NN €.
    """
    # Chercher tous les textes prix
    price_candidates = []
    for el in card.find_all(True):
        classes = " ".join(el.get("class", []))
        text = el.get_text(strip=True)
        m = re.match(r'^(\d+[,\.]\d{1,2})\s*€$', text)
        if m:
            if "line-through" not in classes and "barr" not in classes:
                price_candidates.append(text)

    # Premier prix non barré trouvé
    return price_candidates[0] if price_candidates else None


def parse_listing_page(soup: BeautifulSoup, cat_url: str, cat_label: str) -> List[Dict]:
    """Extrait les produits d'une page listing (listing rapide)."""
    products = []
    container = soup.find(id="container-item")
    if not container:
        return products

    for a in container.find_all("a", recursive=False):
        href  = a.get("href", "").strip()
        title = a.get("title", "").strip()
        if not href or not title:
            continue

        url = href if href.startswith("http") else PHARMACY_URL + href
        price_raw = extract_price_from_card(a)

        products.append({
            "nom":           title,
            "nom_normalise": normalize(title),
            "marque":        "",
            "ean":           "",
            "prix_affiche":  price_raw or "N/D",
            "prix_numerique": parse_price(price_raw),
            "categorie":     cat_label,
            "categorie_url": cat_url,
            "url":           url,
            "enrichi":       False,
        })
    return products


def get_last_page(soup: BeautifulSoup) -> int:
    """Détecte le numéro de la dernière page via les liens ?page=N."""
    nums = []
    for a in soup.find_all("a", href=True):
        m = re.search(r'\?page=(\d+)', a["href"])
        if m:
            nums.append(int(m.group(1)))
    return max(nums) if nums else 1


def get_cat_label(url: str) -> str:
    """Extrait un label lisible depuis l'URL de catégorie."""
    slug = url.rstrip("/").split("/")[-1]
    return slug.replace("-", " ").title()


# ══════════════════════════════════════════════════════════════════
#  STEP 3 — Enrichissement prix via JSON-LD page produit
# ══════════════════════════════════════════════════════════════════
def enrich_from_product_page(session: requests.Session, product: Dict) -> None:
    """Fetche la page produit et enrichit prix + EAN + marque depuis JSON-LD."""
    soup = get_page(session, product["url"])
    if not soup:
        return

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if not isinstance(data, dict):
                continue
            if data.get("@type") not in ("Product", "product"):
                continue

            # Marque
            brand = data.get("brand", {})
            if isinstance(brand, dict):
                product["marque"] = brand.get("name", "")

            # EAN / SKU
            product["ean"] = str(data.get("gtin", "") or data.get("sku", "") or "")

            # Prix
            offers = data.get("offers", {})
            if isinstance(offers, dict):
                price = offers.get("lowPrice") or offers.get("price")
                if price:
                    product["prix_numerique"] = float(price)
                    product["prix_affiche"]   = f"{float(price):.2f} €".replace(".", ",")
            elif isinstance(offers, list) and offers:
                price = offers[0].get("price")
                if price:
                    product["prix_numerique"] = float(price)
                    product["prix_affiche"]   = f"{float(price):.2f} €".replace(".", ",")

            product["enrichi"] = True
            return
        except (json.JSONDecodeError, ValueError, TypeError):
            continue


# ══════════════════════════════════════════════════════════════════
#  STEP 4 — Export Excel
# ══════════════════════════════════════════════════════════════════
def export_excel(products: List[Dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drakkars Parapharmacie"

    cols = [
        ("Nom",           "nom"),
        ("Marque",        "marque"),
        ("EAN",           "ean"),
        ("Catégorie",     "categorie"),
        ("Prix affiché",  "prix_affiche"),
        ("Prix numérique","prix_numerique"),
        ("Nom normalisé", "nom_normalise"),
        ("URL",           "url"),
        ("Catégorie URL", "categorie_url"),
    ]

    hdr_fill = PatternFill("solid", fgColor="1A56DB")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (label, _) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    for row_i, p in enumerate(products, 2):
        for col_i, (_, key) in enumerate(cols, 1):
            val = p.get(key)
            if key == "url":
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.hyperlink = val
                cell.font = Font(color="0057FF", underline="single")
            elif key == "prix_numerique" and val is not None:
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.number_format = '#,##0.00 "€"'
            else:
                ws.cell(row=row_i, column=col_i, value=val)

    widths = [50, 25, 16, 30, 14, 16, 55, 80, 65]
    for col_i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col_i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(path)
    sz = path.stat().st_size / 1024
    log.info(f"Export : {path}  ({sz:.0f} KB, {len(products)} produits)")


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    out = Path(__file__).parent / "benchmark_drakkars.xlsx"

    log.info("=" * 60)
    log.info("  Scraper Pharmacie des Drakkars — Parapharmacie")
    log.info(f"  Mode : {'précis (JSON-LD)' if FETCH_PRODUCT_PAGES else 'rapide (listing)'}")
    log.info("=" * 60)

    session = requests.Session()

    # ── Catégories ───────────────────────────────────────────────
    categories = get_categories(session)
    if not categories:
        log.error("Aucune catégorie trouvée.")
        return

    all_products: List[Dict] = []
    seen_urls = set()

    # ── Parcours catégories → pages → produits ────────────────────
    for cat_i, cat_url in enumerate(categories, 1):
        cat_label = get_cat_label(cat_url)
        log.info(f"\nCatégorie {cat_i}/{len(categories)} : {cat_label}")
        log.info(f"  URL : {cat_url}")

        soup = get_page(session, cat_url)
        if not soup:
            continue
        time.sleep(DELAY)

        last_page = get_last_page(soup)
        log.info(f"  Pages : {last_page}")

        for page_num in range(1, last_page + 1):
            if page_num > 1:
                page_url = f"{cat_url}?page={page_num}"
                soup = get_page(session, page_url)
                if not soup:
                    continue
                time.sleep(DELAY)

            products = parse_listing_page(soup, cat_url, cat_label)

            new = 0
            for p in products:
                if p["url"] not in seen_urls:
                    seen_urls.add(p["url"])
                    all_products.append(p)
                    new += 1

            log.info(f"  Page {page_num}/{last_page} : {len(products)} produits ({new} nouveaux) | Total : {len(all_products)}")

    log.info(f"\nListing terminé — {len(all_products)} produits uniques")

    # ── Enrichissement JSON-LD (optionnel) ────────────────────────
    if FETCH_PRODUCT_PAGES:
        log.info("Enrichissement via pages produit (prix + EAN + marque)…")
        for i, p in enumerate(all_products, 1):
            if i % 50 == 0:
                log.info(f"  Enrichi {i}/{len(all_products)}…")
            enrich_from_product_page(session, p)
            time.sleep(DELAY)

        enriched = sum(1 for p in all_products if p["enrichi"])
        log.info(f"  Enrichissement : {enriched}/{len(all_products)} produits")

    # ── Résumé ───────────────────────────────────────────────────
    log.info("\n" + "=" * 60)
    log.info(f"TOTAL PRODUITS UNIQUES : {len(all_products)}")
    avec_prix = sum(1 for p in all_products if p["prix_numerique"] is not None)
    log.info(f"Avec prix              : {avec_prix}")

    if all_products:
        export_excel(all_products, out)
        log.info(f"\n✓ Fichier : {out}")
    else:
        log.error("Aucun produit extrait.")


if __name__ == "__main__":
    main()
