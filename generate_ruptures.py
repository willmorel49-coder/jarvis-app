#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_ruptures.py — Copilote / feed #2 : RUPTURES & TENSIONS (ANSM).

Récupère la liste ANSM des médicaments signalés en rupture / risque de rupture de
stock (fichier annuel, avec CIP13 + DCI + date), et écrit crm/v2/ruptures-data.js
que le CRM croise à ton catalogue (par CIP13) : « ce produit est en tension ».

Source : ANSM — « Médicaments ayant fait l'objet d'un signalement de rupture ou de
risque de rupture de stock ». Fichiers XLSX annuels. Gratuit, aucune clé.
On découvre l'URL du dernier fichier en lisant la page (les URLs sont datées).

Python 3.9. Dépend de openpyxl.
"""
import io
import os
import re
import sys
import datetime
import urllib.request

import openpyxl

OUT = os.path.join(os.path.dirname(__file__), "crm", "v2", "ruptures-data.js")
PAGE = "https://ansm.sante.fr/page/medicaments-ayant-fait-lobjet-dun-signalement-de-rupture-ou-de-risque-de-rupture-de-stock"
HOST = "https://ansm.sante.fr"
MONTHS_WINDOW = 18   # on garde les signalements des ~18 derniers mois


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read()


def cip13(v):
    try:
        return str(int(float(v))).zfill(13)
    except Exception:
        s = re.sub(r"\D", "", str(v or ""))
        return s.zfill(13) if s else None


def find_files():
    """Découvre les XLSX 'liste-ruptures-medicaments-YYYY' sur la page ANSM,
    renvoie les URLs des 2 années les plus récentes (année en cours + précédente)."""
    html = fetch(PAGE).decode("utf-8", "ignore")
    hits = re.findall(r'href="([^"]*liste[-_]?ruptures?[-_]?medicaments?[-_](\d{4})\.xlsx)"', html, re.I)
    if not hits:
        # repli : tout xlsx contenant 'rupture'
        hits = [(h, (re.search(r"(\d{4})", h) or ["0", "0"])[1]) for h in re.findall(r'href="([^"]*rupture[^"]*\.xlsx)"', html, re.I)]
    seen = {}
    for href, yr in hits:
        url = href if href.startswith("http") else HOST + href
        seen[yr] = url  # une URL par année (la dernière rencontrée)
    years = sorted(seen.keys(), reverse=True)[:2]
    return [seen[y] for y in years]


def parse(url, rows):
    data = fetch(url)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    sh = wb[wb.sheetnames[0]]
    it = sh.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return
    idx = {}
    for i, h in enumerate(header):
        hn = re.sub(r"\s+", " ", str(h or "")).strip().lower()
        if hn.startswith("date"): idx["date"] = i
        elif hn == "cip" or "cip" in hn: idx["cip"] = i
        elif hn == "dci" or hn.startswith("dci"): idx["dci"] = i
        elif hn == "nom": idx["nom"] = i
        elif "laborat" in hn: idx["labo"] = i
    if "cip" not in idx:
        return
    for r in it:
        cip = cip13(r[idx["cip"]]) if idx.get("cip") is not None else None
        if not cip or len(cip) != 13:
            continue
        d = r[idx["date"]] if idx.get("date") is not None else None
        ds = str(d)[:10] if d else ""
        rec = {
            "dci": str(r[idx["dci"]] or "").strip() if idx.get("dci") is not None else "",
            "nom": str(r[idx["nom"]] or "").strip() if idx.get("nom") is not None else "",
            "date": ds,
            "labo": str(r[idx["labo"]] or "").strip() if idx.get("labo") is not None else "",
        }
        # garde le signalement le PLUS RÉCENT par CIP
        if cip not in rows or ds > rows[cip]["date"]:
            rows[cip] = rec


def main():
    print("Découverte des fichiers ANSM…")
    urls = find_files()
    if not urls:
        sys.exit("Aucun fichier 'liste-ruptures-medicaments' trouvé sur la page ANSM.")
    for u in urls:
        print("  ·", u.split("/")[-1])

    rows = {}
    for u in urls:
        try:
            parse(u, rows)
        except Exception as e:
            sys.stderr.write("  ! %s : %s\n" % (u, e))

    if not rows:
        sys.exit("Aucune ligne exploitable.")

    # fenêtre glissante : ~18 derniers mois
    cutoff = (datetime.date(2020, 1, 1)).isoformat()
    dates = [r["date"] for r in rows.values() if r["date"]]
    if dates:
        latest = max(dates)
        y, m = int(latest[:4]), int(latest[5:7])
        m -= MONTHS_WINDOW
        while m <= 0:
            m += 12; y -= 1
        cutoff = "%04d-%02d-01" % (y, m)
    kept = {c: r for c, r in rows.items() if r["date"] >= cutoff}
    print("Signalements retenus (depuis %s) : %d CIP" % (cutoff, len(kept)))

    # écriture JS compacte
    def esc(s):
        return str(s).replace("\\", "").replace('"', "'")
    items = []
    for c, r in sorted(kept.items()):
        items.append('"%s":{d:"%s",dt:"%s"}' % (c, esc(r["dci"])[:40], r["date"]))
    annee = max((r["date"][:4] for r in kept.values() if r["date"]), default="")
    js = (
        "// ANSM — médicaments signalés en rupture / risque de rupture de stock.\n"
        "// Source: ansm.sante.fr (liste annuelle des signalements) · fenêtre ~%d mois.\n"
        "// Croisé au catalogue par CIP13. Généré par generate_ruptures.py — ne pas éditer.\n"
        "window.RUPTURES={meta:{annee:\"%s\",n:%d,source:\"ANSM\"},data:{%s}};\n"
    ) % (MONTHS_WINDOW, annee, len(kept), ",".join(items))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(js)
    print("Écrit %s (%d Ko, %d CIP)" % (OUT, os.path.getsize(OUT) // 1024, len(kept)))


if __name__ == "__main__":
    main()
