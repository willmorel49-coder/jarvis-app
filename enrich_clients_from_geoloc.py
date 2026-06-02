#!/usr/bin/env python3
"""
Enrichit crm/clients-data.js avec les vraies données depuis STATS/WML_geolocalisation_2026-05-07.xlsx.
Pour chaque CIP matché : met à jour adresse, cp, ville, tel, email, contact pharmacien,
logiciel d'officine (Winpharma/Alliadis/LGPI/Offilog), enseigne, structure (OPS).
"""

import json
import openpyxl
import re
import shutil
from pathlib import Path

SRC = 'STATS/WML_geolocalisation_2026-05-07.xlsx'
CRM = Path('crm/clients-data.js')


def clean_int_str(v):
    if v is None: return ''
    s = str(v).strip()
    if s.endswith('.0'): s = s[:-2]
    return s


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb['Donnees']
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col = {h: i for i, h in enumerate(headers)}

    enriched = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cip = clean_int_str(row[col['TIRCODE']])
        if not cip.isdigit(): continue
        adresse_parts = [str(row[col[c]]).strip() for c in ['ADRL1', 'ADRL2', 'ADRL3'] if row[col[c]]]
        # Logiciel officine (1 actif parmi WINPHARMA/ALLIADIS/LGPI/OFFILOG/PHARMALAND)
        logiciels = []
        for sw in ['WINPHARMA', 'ALLIADIS', 'LGPI', 'OFFILOG', 'PHARMALAND']:
            if str(row[col[sw]] or '').strip().upper() == 'O':
                logiciels.append(sw.title())

        enriched[cip] = {
            'cip': cip,
            'nom': str(row[col['TIRSOCIETE']] or '').strip(),
            'adresse': ' · '.join(adresse_parts),
            'cp': clean_int_str(row[col['ADRCODEPOSTAL']]).zfill(5)[:5] if row[col['ADRCODEPOSTAL']] else '',
            'ville': str(row[col['ADRVILLE']] or '').strip(),
            'tel': clean_int_str(row[col['ADRTEL']]),
            'portable': clean_int_str(row[col['ADRPORTABLE']]),
            'email': str(row[col['ADRMAIL']] or '').strip(),
            'contact_nom': str(row[col['ADRCONTACTNOM']] or '').strip(),
            'contact_prenom': str(row[col['ADRCONTACTPRENOM']] or '').strip(),
            'enseigne': str(row[col['TIRENSEIGNE']] or '').strip(),
            'logiciels': logiciels,
            'structure': str(row[col['Structure']] or '').strip(),
            'uga': str(row[col['UGA']] or '').strip(),
            'siren': clean_int_str(row[col['TIRSIREN']]),
            'origine': str(row[col['TIRORIGINE']] or '').strip(),
            'cible1': str(row[col['TIRCIBLE1']] or '').strip(),
            'cible2': str(row[col['TIRCIBLE2']] or '').strip(),
            'is_actif': str(row[col['TIRISACTIF']] or '').strip() == 'O',
        }
    print(f'[enrich] {len(enriched)} clients lus dans WML géoloc')

    # Charge le CRM existant
    content = CRM.read_text(encoding='utf-8')
    backup = CRM.with_suffix('.bak.js')
    shutil.copy(CRM, backup)

    # Enrichit en place : pour chaque CIP du geoloc, update les champs du record JS
    updates = 0
    for cip, data in enriched.items():
        rec_pattern = re.compile(r'\{cip:"' + re.escape(cip) + r'"[^}]*\}')
        match = rec_pattern.search(content)
        if not match:
            continue
        rec = match.group(0)

        def upd(key, val):
            nonlocal rec
            if val is None or val == '':
                return
            val_str = str(val).replace('\\', '\\\\').replace('"', '\\"')
            pat = re.compile(rf'({key}:")[^"]*(")')
            if pat.search(rec):
                rec = pat.sub(rf'\g<1>{val_str}\g<2>', rec, count=1)
            else:
                # Ajout avant le } final
                rec = rec[:-1] + f',{key}:"{val_str}"' + '}'

        upd('nom', data['nom'])
        upd('adresse', data['adresse'])
        upd('cp', data['cp'])
        upd('ville', data['ville'])
        upd('tel', data['tel'])
        upd('email', data['email'])
        # Champs nouveaux
        upd('contact', f'{data["contact_prenom"]} {data["contact_nom"]}'.strip())
        upd('enseigne', data['enseigne'])
        upd('logiciel', ' / '.join(data['logiciels']))
        upd('structure', data['structure'])
        upd('uga', data['uga'])

        content = content[:match.start()] + rec + content[match.end():]
        updates += 1

    CRM.write_text(content, encoding='utf-8')
    print(f'[enrich] {updates} pharmacies enrichies dans crm/clients-data.js')
    print(f'[enrich] Backup → {backup}')


if __name__ == '__main__':
    main()
