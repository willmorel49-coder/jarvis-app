#!/usr/bin/env python3
"""
generate_offilog.py
===================
Lit : OFFILOG/offilog_x_maxipara_3517.xlsx  (feuille "Croisement Complet")
       benchmark_drakkars.xlsx               (si disponible → enrichit prix_drakkars)
       benchmark_cap3000.xlsx                (si disponible → enrichit prix_cap3000)
Écrit: crm/offilog-data.js

Format JS :
  const OFFILOG = [{rang, produit, produit_norm, ean, marque, univers, saison,
                    prix_maxi, dans_offilog, marque_off, prix_offilog, ecart,
                    marge_pct, potentiel, role, prix_drakkars, prix_cap3000}, ...]

Stratégie de matching (par ordre de priorité) :
  1. Par EAN exact (si ean_str non vide) — sources : Drakkars, Cap3000
  2. Par nom normalisé (fallback) — toutes les sources
"""
import re
import unicodedata
import json
from datetime import datetime
from pathlib import Path

import openpyxl

BASE      = Path(__file__).parent
SRC       = BASE / 'OFFILOG' / 'offilog_x_maxipara_3517.xlsx'
DRAKKARS  = BASE / 'benchmark_drakkars.xlsx'
CAP3000   = BASE / 'benchmark_cap3000.xlsx'
LIVE_JSON  = BASE / 'offilog_live.json'
BEST_JSON  = BASE / 'bestsellers_offilog.json'
IMG_JSON   = BASE / 'images_by_ean.json'   # lookup consolidé toutes sources
PHARMA     = BASE / 'benchmark_apothical_pharmacie_montlouis-sur-loire.xlsx'
LECLERC_JSON = BASE / 'leclerc_prices.json'
OUT        = BASE / 'crm' / 'offilog-data.js'
# ⚠️ 03/09/2026 — nos CONDITIONS COMMERCIALES sortent du fichier public.
# `jarvis-app` est un dépôt PUBLIC servi par GitHub Pages : tout ce qui
# est écrit dans crm/offilog-data.js est téléchargeable sans mot de passe.
# prix_offilog, ecart et marge_pct partent donc dans un fichier séparé,
# hors dépôt (.gitignore) et servi par adresse signée Supabase.
# Les TROIS ensemble, jamais l'un sans les autres : prix_offilog se
# recalcule exactement par `prix_maxi - ecart` (vérifié 2000/2000) et par
# `prix_maxi x (1 - marge_pct/100)`. N'en retirer qu'un est un faux correctif.
OUT_COND   = BASE / 'crm' / 'offilog-conditions.js'


# ── HELPERS ─────────────────────────────────────────────────────────────────
def norm(text: str) -> str:
    """Normalize: minuscules, sans accents, alphanum+espace uniquement."""
    if not text:
        return ''
    text = str(text).lower().strip()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def js_str(v) -> str:
    if v is None:
        return 'null'
    s = (str(v)
         .replace('\\', '\\\\')
         .replace('"', '\\"')
         .replace('&', '&amp;')
         .replace('<', '&lt;')
         .replace('>', '&gt;'))
    return f'"{s}"'


def js_num(v) -> str:
    if v is None:
        return 'null'
    try:
        f = float(v)
        return f'{f:.4f}'.rstrip('0').rstrip('.')
    except (TypeError, ValueError):
        return 'null'


def js_bool(v) -> str:
    return 'true' if v else 'false'


# ── LOAD OFFILOG EXCEL ───────────────────────────────────────────────────────
import sys
if not SRC.exists():
    print(f'ERREUR : fichier source introuvable → {SRC}')
    print('Vérifiez que OFFILOG/offilog_x_maxipara_3517.xlsx est présent.')
    sys.exit(1)

print(f'Lecture : {SRC}')
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb['Croisement Complet']
rows = list(ws.iter_rows(values_only=True))
wb.close()

HEADER = rows[0]
print(f'Colonnes : {HEADER}')
print(f'Produits : {len(rows) - 1}')

# Mapping colonnes
COL = {name: i for i, name in enumerate(HEADER) if name}

def get(row, col_name):
    idx = COL.get(col_name)
    if idx is None:
        return None
    v = row[idx]
    if isinstance(v, str):
        v = v.strip()
        if v == '':
            return None
    return v


# ── LOAD DRAKKARS (optionnel) ────────────────────────────────────────────────
# Colonnes: Nom, Marque, EAN, Catégorie, Prix affiché, Prix numérique,
#           Nom normalisé, URL, Catégorie URL
# Matching : priorité EAN, fallback nom normalisé
drakkars_by_norm: dict = {}
drakkars_by_ean:  dict = {}
if DRAKKARS.exists():
    print(f'Enrichissement Drakkars depuis : {DRAKKARS}')
    wb4 = openpyxl.load_workbook(DRAKKARS, read_only=True, data_only=True)
    ws4 = wb4.active
    rows4 = list(ws4.iter_rows(values_only=True))
    wb4.close()
    hdr4 = rows4[0]
    col4 = {n: i for i, n in enumerate(hdr4) if n}
    ean_col4  = col4.get('EAN')
    prix_col4 = col4.get('Prix numérique')
    norm_col4 = col4.get('Nom normalisé')
    nom_col4  = col4.get('Nom')
    for r in rows4[1:]:
        prix = r[prix_col4] if prix_col4 is not None else None
        prix_f = None
        try:
            prix_f = float(prix) if prix is not None else None
        except (TypeError, ValueError):
            pass
        # Index par EAN
        ean_raw = r[ean_col4] if ean_col4 is not None else None
        if ean_raw is not None:
            ean_str_d = str(int(ean_raw)) if isinstance(ean_raw, float) else str(ean_raw).strip()
            if ean_str_d.isdigit() and len(ean_str_d) >= 8:
                if ean_str_d not in drakkars_by_ean:
                    drakkars_by_ean[ean_str_d] = prix_f
        # Index par nom normalisé (colonne du scraper ou calculé)
        nom_norm_raw = r[norm_col4] if norm_col4 is not None else None
        if nom_norm_raw:
            k = str(nom_norm_raw).strip()
        else:
            nom_raw = r[nom_col4] if nom_col4 is not None else None
            k = norm(str(nom_raw)) if nom_raw else ''
        if k and k not in drakkars_by_norm:
            drakkars_by_norm[k] = prix_f
    print(f'  → {len(drakkars_by_norm)} produits Drakkars par nom | {len(drakkars_by_ean)} par EAN')
else:
    print(f'Fichier Drakkars non trouvé ({DRAKKARS}) — prix_drakkars = null')


# ── LOAD CAP3000 (optionnel) ─────────────────────────────────────────────────
# EAN disponible directement (extrait du dump PHP du site)
cap3000_by_ean:  dict = {}
cap3000_by_norm: dict = {}
if CAP3000.exists():
    print(f'Enrichissement Cap3000 depuis : {CAP3000}')
    wb5 = openpyxl.load_workbook(CAP3000, read_only=True, data_only=True)
    ws5 = wb5.active
    rows5 = list(ws5.iter_rows(values_only=True))
    wb5.close()
    hdr5 = rows5[0]
    col5 = {str(n).strip(): i for i, n in enumerate(hdr5) if n}
    ean_col5  = col5.get('EAN')
    prix_col5 = col5.get('Prix numérique')
    norm_col5 = col5.get('Nom normalisé')
    nom_col5  = col5.get('Nom')
    for r in rows5[1:]:
        prix = r[prix_col5] if prix_col5 is not None else None
        prix_f = None
        try:
            prix_f = float(prix) if prix is not None else None
        except (TypeError, ValueError):
            pass
        ean_raw = r[ean_col5] if ean_col5 is not None else None
        if ean_raw is not None:
            ean_s = str(int(ean_raw)) if isinstance(ean_raw, float) else str(ean_raw).strip()
            if ean_s.isdigit() and len(ean_s) >= 8 and ean_s not in cap3000_by_ean:
                cap3000_by_ean[ean_s] = prix_f
        nom_norm_raw = r[norm_col5] if norm_col5 is not None else None
        if nom_norm_raw:
            k = str(nom_norm_raw).strip()
        else:
            nom_raw = r[nom_col5] if nom_col5 is not None else None
            k = norm(str(nom_raw)) if nom_raw else ''
        if k and k not in cap3000_by_norm:
            cap3000_by_norm[k] = prix_f
    print(f'  → {len(cap3000_by_norm)} produits Cap3000 par nom | {len(cap3000_by_ean)} par EAN')
else:
    print(f'Fichier Cap3000 non trouvé ({CAP3000}) — prix_cap3000 = null')


# ── LOAD OFFILOG LIVE (optionnel) ────────────────────────────────────────────
# Données scrapées depuis offilog.fr : prix d'achat réel + images produit
# Matching : EAN prioritaire (tous les produits live ont un EAN)
live_by_ean:  dict = {}   # ean_str → {prix, img}
live_by_norm: dict = {}   # nom_norm → {prix, img}
if LIVE_JSON.exists():
    print(f'Enrichissement Offilog Live depuis : {LIVE_JSON}')
    with open(LIVE_JSON, encoding='utf-8') as _f:
        _live = json.load(_f)
    for _p in _live.get('products', []):
        _ean = str(_p.get('ean', '') or '').strip()
        _prix = _p.get('prix')
        _img  = _p.get('img', '') or ''
        _entry = {'prix': float(_prix) if _prix else None, 'img': _img}
        if _ean and _ean.isdigit() and len(_ean) >= 8:
            if _ean not in live_by_ean:
                live_by_ean[_ean] = _entry
        _nom_n = norm(str(_p.get('nom', '') or ''))
        if _nom_n and _nom_n not in live_by_norm:
            live_by_norm[_nom_n] = _entry
    print(f'  → {len(live_by_ean)} produits Offilog Live par EAN | {len(live_by_norm)} par nom')
else:
    print(f'Fichier Offilog Live non trouvé ({LIVE_JSON}) — prix_live/img = null')


# ── LOAD LECLERC (optionnel) ────────────────────────────────────────────────
leclerc_by_ean: dict = {}
if LECLERC_JSON.exists():
    print(f'Enrichissement Leclerc depuis : {LECLERC_JSON}')
    with open(LECLERC_JSON, encoding='utf-8') as _f:
        _lecl = json.load(_f)
    for _ean, _p in _lecl.get('prices', {}).items():
        leclerc_by_ean[str(_ean)] = float(_p['prix'])
    print(f'  → {len(leclerc_by_ean)} produits Leclerc par EAN')
else:
    print(f'Fichier Leclerc non trouvé ({LECLERC_JSON}) — prix_leclerc = null')


# ── LOAD BESTSELLERS (optionnel) ──────────────────────────────────────────────
# Classement des meilleures ventes Offilog (rang 1 = meilleure vente)
best_by_ean:  dict = {}   # ean_str → rang_vente int
best_by_norm: dict = {}   # nom_norm → rang_vente int
if BEST_JSON.exists():
    print(f'Enrichissement Best-sellers depuis : {BEST_JSON}')
    with open(BEST_JSON, encoding='utf-8') as _f:
        _best = json.load(_f)
    for _p in _best.get('products', []):
        _ean = str(_p.get('ean', '') or '').strip()
        _rang = _p.get('rang')
        if _ean and _ean.isdigit() and len(_ean) >= 8 and _rang:
            best_by_ean[_ean] = int(_rang)
        _nom_n = norm(str(_p.get('nom', '') or ''))
        if _nom_n and _rang and _nom_n not in best_by_norm:
            best_by_norm[_nom_n] = int(_rang)
    print(f'  → {len(best_by_ean)} best-sellers par EAN')
else:
    print(f'Fichier Best-sellers non trouvé ({BEST_JSON}) — rang_vente = null')


# ── LOAD PHARMACIE (optionnel) ───────────────────────────────────────────────
# Prix affiché sur le site Apothical de la pharmacie Montlouis-sur-Loire
# Matching : nom exact, puis préfixe (premiers mots) en fallback
pharma_by_norm:   dict = {}   # nom_norm exact → prix
pharma_by_prefix: dict = {}   # " ".join(premiers 4 mots) → prix
STOPWORDS = {'de', 'du', 'la', 'le', 'les', 'et', 'en', 'a', 'au', 'aux', 'ml', 'mg', 'g', 'l', 'x'}

def prefix_key(name, n=4):
    words = [w for w in name.split() if w not in STOPWORDS and not w.isdigit()]
    return ' '.join(words[:n])

if PHARMA.exists():
    print(f'Enrichissement prix pharmacie depuis : {PHARMA}')
    wb6 = openpyxl.load_workbook(PHARMA, read_only=True, data_only=True)
    ws6 = wb6.active
    rows6 = list(ws6.iter_rows(values_only=True))
    wb6.close()
    hdr6 = rows6[0]
    col6 = {str(n).strip(): i for i, n in enumerate(hdr6) if n}
    nom_col6  = col6.get('Nom normalisé')
    prix_col6 = col6.get('Prix affiché')
    for r in rows6[1:]:
        k = str(r[nom_col6]).strip() if nom_col6 is not None and r[nom_col6] else None
        prix = r[prix_col6] if prix_col6 is not None else None
        try:
            prix_f = float(prix) if prix is not None else None
        except (TypeError, ValueError):
            prix_f = None
        if not k:
            continue
        if k not in pharma_by_norm:
            pharma_by_norm[k] = prix_f
        pk = prefix_key(k)
        if pk and len(pk) > 4 and pk not in pharma_by_prefix:
            pharma_by_prefix[pk] = prix_f
    valides = sum(1 for v in pharma_by_norm.values() if v)
    print(f'  → {valides} prix pharmacie valides sur {len(pharma_by_norm)} produits | {len(pharma_by_prefix)} clés préfixe')
else:
    print(f'Fichier pharmacie non trouvé ({PHARMA}) — prix_pharmacie = null')


# ── LOAD IMAGES CONSOLIDÉES (optionnel, priorité max) ────────────────────────
# Lookup EAN/nom_norm → URL image (toutes sources Offilog + Drakkars + Cap3000)
img_by_ean:  dict = {}
img_by_norm: dict = {}
if IMG_JSON.exists():
    print(f'Enrichissement images consolidées depuis : {IMG_JSON}')
    with open(IMG_JSON, encoding='utf-8') as _f:
        _imgs = json.load(_f)
    img_by_ean  = _imgs.get('by_ean', {})
    img_by_norm = _imgs.get('by_norm', {})
    print(f'  → {len(img_by_ean)} images par EAN · {len(img_by_norm)} par nom')
else:
    print(f'Fichier images consolidées non trouvé ({IMG_JSON}) — fallback live uniquement')


# ── BUILD RECORDS ────────────────────────────────────────────────────────────
records = []
for row in rows[1:]:
    produit    = get(row, 'Produit') or ''
    if not produit:
        continue

    rang       = get(row, 'Rang')
    source     = get(row, 'Source')
    marque     = get(row, 'Marque_Maxi') or ''
    ean        = get(row, 'EAN')
    ean_str    = str(int(ean)) if isinstance(ean, (int, float)) and ean else (str(ean) if ean else '')
    prix_maxi  = get(row, 'Prix_Public')
    dans_off   = str(get(row, 'Dans_Offilog') or '').upper() == 'OUI'
    marque_off = get(row, 'Marque_Offilog') or ''
    univers    = get(row, 'Univers') or ''
    saison     = get(row, 'Saison')
    prix_off   = get(row, 'Prix_Achat_Offilog')
    ecart      = get(row, 'Ecart_Prix')
    marge_pct  = get(row, 'Marge_%')
    potentiel  = get(row, 'Potentiel') or ''
    role       = get(row, 'Role_Recommande') or ''

    produit_n = norm(produit)

    # ── Matching Drakkars : EAN en priorité, nom normalisé en fallback
    prix_drakkars = None
    if ean_str:
        prix_drakkars = drakkars_by_ean.get(ean_str)
    if prix_drakkars is None:
        prix_drakkars = drakkars_by_norm.get(produit_n)

    # ── Matching Cap3000 : EAN en priorité, nom normalisé en fallback
    prix_cap3000 = None
    if ean_str:
        prix_cap3000 = cap3000_by_ean.get(ean_str)
    if prix_cap3000 is None:
        prix_cap3000 = cap3000_by_norm.get(produit_n)

    # ── Matching Offilog Live : EAN en priorité, nom en fallback
    _live_entry = None
    if ean_str:
        _live_entry = live_by_ean.get(ean_str)
    if _live_entry is None:
        _live_entry = live_by_norm.get(produit_n)
    prix_live = _live_entry['prix'] if _live_entry else None
    # Si trouvé dans le live, on confirme dans_offilog=True
    if _live_entry:
        dans_off = True

    # ── Image : lookup consolidé (toutes sources) en priorité, live en fallback
    img = ''
    if img_by_ean and ean_str:
        img = img_by_ean.get(ean_str, '')
    if not img and img_by_norm:
        img = img_by_norm.get(produit_n, '')
    if not img and _live_entry:
        img = _live_entry.get('img', '') or ''

    # ── Matching Best-sellers : EAN en priorité, nom en fallback
    rang_vente = None
    if ean_str:
        rang_vente = best_by_ean.get(ean_str)
    if rang_vente is None:
        rang_vente = best_by_norm.get(produit_n)

    # ── Matching Pharmacie : exact, puis préfixe 4 mots
    prix_pharmacie = pharma_by_norm.get(produit_n)
    if prix_pharmacie is None:
        pk = prefix_key(produit_n)
        if pk and len(pk) > 4:
            prix_pharmacie = pharma_by_prefix.get(pk)

    # ── Matching Leclerc : EAN uniquement
    prix_leclerc = leclerc_by_ean.get(ean_str) if ean_str else None

    records.append({
        'rang':           rang,
        'produit':        produit,
        'produit_norm':   produit_n,
        'ean':            ean_str,
        'marque':         marque,
        'univers':        univers,
        'saison':         saison,
        'prix_maxi':      prix_maxi,
        'dans_offilog':   dans_off,
        'marque_off':     marque_off,
        'prix_offilog':   prix_off,
        'ecart':          ecart,
        'marge_pct':      marge_pct,
        'potentiel':      potentiel,
        'role':           role,
        'prix_drakkars':  prix_drakkars,
        'prix_cap3000':   prix_cap3000,
        'prix_live':      prix_live,
        'img':            img,
        'rang_vente':     rang_vente,
        'prix_pharmacie': prix_pharmacie,
        'prix_leclerc':   prix_leclerc,
    })

print(f'Records valides : {len(records)}')
avec_drakkars = sum(1 for r in records if r['prix_drakkars'] is not None)
avec_cap3000  = sum(1 for r in records if r['prix_cap3000']  is not None)
avec_live     = sum(1 for r in records if r['prix_live']   is not None)
avec_img      = sum(1 for r in records if r['img'])
avec_best     = sum(1 for r in records if r['rang_vente']    is not None)
avec_pharma   = sum(1 for r in records if r['prix_pharmacie'] is not None)
avec_leclerc  = sum(1 for r in records if r['prix_leclerc']  is not None)
print(f'Avec prix Drakkars  : {avec_drakkars}')
print(f'Avec prix Cap3000   : {avec_cap3000}')
print(f'Avec prix E.Leclerc : {avec_leclerc}')
print(f'Avec prix Live      : {avec_live}')
print(f'Avec image          : {avec_img}')
print(f'Best-sellers matchés: {avec_best}')
print(f'Avec prix pharmacie : {avec_pharma}')
dans_off_n = sum(1 for r in records if r['dans_offilog'])
print(f'Dans Offilog : {dans_off_n}')


# ── WRITE JS ─────────────────────────────────────────────────────────────────
today = datetime.now().strftime('%Y-%m-%d')
lines = [
    f'// Intégral Pharma — Offilog × Maxipara × Drakkars × Cap3000 × Live',
    f'// Généré le {today}',
    f'// {len(records)} produits | {dans_off_n} dans Offilog | {avec_live} prix live | {avec_img} images | {avec_drakkars} Drakkars | {avec_cap3000} Cap3000 | {avec_leclerc} Leclerc',
    f'const OFFILOG = [',
]

for r in records:
    line = (
        f'  {{rang:{js_num(r["rang"])},produit:{js_str(r["produit"])},'
        f'produit_norm:{js_str(r["produit_norm"])},ean:{js_str(r["ean"])},'
        f'marque:{js_str(r["marque"])},univers:{js_str(r["univers"])},'
        f'saison:{js_str(r["saison"])},prix_maxi:{js_num(r["prix_maxi"])},'
        f'dans_offilog:{js_bool(r["dans_offilog"])},marque_off:{js_str(r["marque_off"])},'
        # prix_offilog / ecart / marge_pct : voir OUT_COND. Rien ici.
        f'potentiel:{js_str(r["potentiel"])},'
        f'role:{js_str(r["role"])},prix_drakkars:{js_num(r["prix_drakkars"])},'
        f'prix_cap3000:{js_num(r["prix_cap3000"])},'
        f'prix_live:{js_num(r["prix_live"])},img:{js_str(r["img"] or None)},'
        f'rang_vente:{js_num(r["rang_vente"])},'
        f'prix_pharmacie:{js_num(r["prix_pharmacie"])},'
        f'prix_leclerc:{js_num(r["prix_leclerc"])}}},'
    )
    lines.append(line)

lines.append('];')

OUT.parent.mkdir(exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines) + '\n')

sz = OUT.stat().st_size / 1024
print(f'\n✓ Écrit : {OUT}  ({sz:.0f} KB, {len(records)} produits)')

# ── LES CONDITIONS COMMERCIALES, À PART ──────────────────────────────────
# Format volontairement compact : un objet EAN -> [prix_offilog, ecart,
# marge_pct]. Le but est qu'il reste LÉGER. Le 15/08/2026, protéger des
# fichiers lourds avait rendu l'app inutilisable (17 Mo retéléchargés à
# chaque ouverture, adresse signée = pas de cache navigateur). À quelques
# dizaines de Ko, ce reproche ne tient plus.
cond_lines = [
    '// Intégral Pharma — Offilog, CONDITIONS COMMERCIALES',
    f'// Généré le {today}',
    '// ⚠️ NE JAMAIS COMMITER. Fichier servi par adresse signée (Supabase).',
    '// ean -> [prix_offilog, ecart, marge_pct]',
    'const OFFILOG_COND = {',
]
n_cond = 0
for r in records:
    if not r['ean'] or r['prix_offilog'] in (None, ''):
        continue
    cond_lines.append(
        f'{js_str(r["ean"])}:[{js_num(r["prix_offilog"])},'
        f'{js_num(r["ecart"])},{js_num(r["marge_pct"])}],'
    )
    n_cond += 1
cond_lines.append('};')
with open(OUT_COND, 'w', encoding='utf-8') as f:
    f.write('\n'.join(cond_lines) + '\n')
szc = OUT_COND.stat().st_size / 1024
print(f'✓ Écrit : {OUT_COND}  ({szc:.0f} KB, {n_cond} conditions) — NE PAS COMMITER')

# ── GARDE-FOU : le fichier public ne doit plus porter nos conditions ─────
# Un contrôle qui ne s'exécute jamais ne protège rien : celui-ci tourne à
# chaque génération et fait échouer le script plutôt que d'écrire la fuite.
_public = OUT.read_text(encoding='utf-8')
_fuites = [c for c in ('prix_offilog', 'marge_pct', 'ecart') if c + ':' in _public]
if _fuites:
    raise SystemExit(
        f'ARRÊT : {OUT.name} contient encore {_fuites}. '
        'Ce fichier part dans un dépôt PUBLIC — corriger avant de committer.')
print('✓ Contrôle : aucune condition commerciale dans le fichier public.')
